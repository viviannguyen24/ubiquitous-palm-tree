import asyncio
import calendar
from collections import Counter
from datetime import date, datetime, timedelta
from getpass import getpass
import hashlib
import json
from pathlib import Path
import re
import sys
from typing import Literal

import pandas as pd
from pydantic import BaseModel, Field, create_model

try:
    from agents import Agent, Runner, set_default_openai_key
except ImportError as fout:
    raise ImportError(
        "De OpenAI Agents SDK ontbreekt. Activeer eerst .venv en voer uit: "
        "python -m pip install openai-agents"
    ) from fout


# Vaste hoofdprompt. blijft voor iedere fictieve patiënt gelijk.
MAIN_PROMPT = """
Je maakt uitsluitend synthetische Nederlandse huisartsendossiers voor onderzoek.
Alle personen en gebeurtenissen moeten fictief zijn. Neem nooit gegevens van een
bestaande patiënt over.

Schrijf beknopt en realistisch in Nederlandse huisartsentaal. Maak een logisch,
chronologisch dossier zonder medische tegenstrijdigheden. Gebruik bij ieder
deelcontact de SOEP-structuur:
- S: klachten, ervaringen en informatie van de patiënt.
- O: objectieve bevindingen, lichamelijk onderzoek en relevante meetwaarden.
- E: beoordeling of diagnose, zo mogelijk met een passende ICPC-code tussen haken.
- P: beleid, medicatie, controles, adviezen en eventuele verwijzing.

Verdeel de contacten logisch over huisarts, POH-S en doktersassistente. Maak alleen
uitslagen, medicatie en verwijzingen die medisch passen bij de fictieve casus.

Genereer een overzicht met persoons- en achtergrondgegevens. Neem daarin een
rookstatus, alcoholgebruik, familieanamnese, woonsituatie, beroep of dagbesteding
en functionele context op. Laat deze gegevens aansluiten op leeftijd, episodes en
zorgbehoefte, maar vermijd stereotiepe of onnodig gedetailleerde invulling. Als
een achtergrondgegeven in de patiëntenprompt handmatig is opgegeven, moet die
waarde exact inhoudelijk worden gevolgd.

Genereer daarnaast een afzonderlijke episode-/probleemlijst. Neem iedere episode
uit de patiëntenprompt precies één keer op en behoud de opgegeven volgorde en
episodenaam. Vermeld per episode een passende ICPC-code, begin- en eventuele
einddatum, status, attentiewaarde, een korte samenvatting van het beloop en het
actuele of afsluitende beleid. Een actieve episode heeft geen einddatum. Een
afgesloten episode heeft een passende einddatum. Een begindatum mag vóór de
dossierperiode liggen als de aandoening aantoonbaar al eerder bestond.

Genereer daarnaast een medicatielijst die volledig aansluit op het journaal.
Neem actuele medicatie en relevante gestopte of eenmalige medicatie uit de
dossierperiode op. Als medicatie wordt gestart, gewijzigd of gestopt, moet dit
logisch terugkomen in de P-regels van het journaal en in de medicatielijst.
Gebruik geen medicatie zonder passende indicatie.

Genereer daarnaast volledige fictieve verwijsbrieven en bijbehorende
specialistenbrieven wanneer een verwijzing klinisch logisch is. Iedere verwijzing
en de bijbehorende specialistenbrief vormen samen een traject en krijgen hetzelfde
traject-ID, bijvoorbeeld T001. De verwijsbrief is altijd van de huisarts aan de
specialist. De specialistenbrief is van de specialist aan de huisarts en heeft
altijd een latere datum dan de bijbehorende verwijsbrief.

Maak bij een acute SEH-presentatie of ziekenhuisopname via ambulance of
zelfverwijzing géén fictieve voorafgaande huisartsverwijzing. Gebruik daarvoor
één 'Ontslagbrief' van het ziekenhuis aan de huisarts met een eigen traject-ID.
De ontslagbrief heeft een latere datum dan het acute contact of de opnamedatum.
Een acute ontslagbrief vormt dus geen koppel met een verwijsbrief. Gebruik het
type 'Specialistenbrief' uitsluitend als retourbrief binnen een regulier
huisartsverwijstraject.

Een ontslagbrief mag zijn gekoppeld aan een relevante diagnose die tijdens
dezelfde acute opname is vastgesteld, ook als het eerdere acute contact onder
een andere primaire episode staat. Zo kan een contact wegens een acuut CVA
worden gevolgd door een ontslagbrief onder de episode Atriumfibrilleren wanneer
dit ritmeprobleem tijdens die opname is vastgesteld. Maak de klinische relatie
met de opname dan expliciet in het contactplan en in de brief.

Een verwijsbrief bevat in beknopte maar volledige vorm de reden van verwijzing,
relevante voorgeschiedenis, relevante bevindingen en uitslagen, relevante
medicatie en een duidelijke vraagstelling aan de specialist. Een specialistenbrief
bevat in beknopte maar volledige vorm de reden van beoordeling, bevindingen,
eventueel verricht onderzoek, conclusie of diagnose, beleid en vervolgadvies.
Een ontslagbrief bevat de reden van acute opname, het relevante beloop,
onderzoek, ontslagdiagnosen, medicatiebeleid en vervolgafspraken.
Gebruik geen namen, adressen, geboortedata of andere identificerende gegevens in
de brieven.

Schrijf alle correspondentie — verwijsbrieven, specialistenbrieven en
ontslagbrieven — altijd in verzorgde, professioneel-medische taal met correcte
spelling, grammatica, hoofdletters en interpunctie. Gebruik volledige zinnen en
geen telegramstijl, opzettelijke typefouten, storende herhaling of kunstmatig
toegevoegde afkortingen. De ingestelde ruisniveaus gelden uitsluitend voor de
S-, O-, E- en P-regels van het huisartsjournaal en nooit voor correspondentie.

Laat de correspondentie volledig aansluiten op het journaal en de medicatielijst.
Een verwijzing moet terugkomen in de P-regel van een passend deelcontact. Een
relevant advies, diagnose of medicatiewijziging uit een specialistenbrief moet
logisch terugkomen in een later deelcontact en zo nodig in de medicatielijst.

Volg voor allergieën, microbiologie en laboratorium altijd de gekozen modus uit
de patiëntenprompt: 'Automatisch', 'Zelf invoeren' of 'Geen'. Bij 'Zelf
invoeren' moeten de opgegeven items herkenbaar en volledig in de uitvoer worden
verwerkt. Bij 'Geen' moet de betreffende lijst leeg blijven.

Een allergieoverzicht vermeldt het allergeen, het type allergie, de reactie, de
ernst, de registratiedatum en de status. Als een geneesmiddelenallergie aanwezig
is, mag een geneesmiddel uit die groep nergens in het dossier worden
voorgeschreven.

Een microbiologische uitslag vermeldt datum, episode, materiaal, onderzoek,
uitslag, eventuele verwekker, hoeveelheid of groei,
gevoeligheid/resistentie en conclusie. Laat iedere uitslag logisch aansluiten
op de klachten, het journaal en het antibioticabeleid.

Groepeer laboratoriumbepalingen per realistisch prikmoment. Alle bepalingen uit
dezelfde aanvraag krijgen hetzelfde aanvraag-ID, dezelfde datum en dezelfde
aanvraagreden. Een laboratoriumuitslag vermeldt daarnaast episode, bepaling,
waarde, eenheid, referentiewaarde, afwijking en conclusie. Een aanvraag mag
meerdere klinisch samenhangende bepalingen bevatten en hoeft dus niet beperkt
te blijven tot één episode-specifieke waarde. Neem ook normale bepalingen op als
die bij hetzelfde pakket zijn aangevraagd.

Laat laboratoriumonderzoek aansluiten op de episodes, medicatie, controles,
leeftijd en klachten. Overweeg bij een passende cardiovasculaire controle onder
andere nierfunctie, elektrolyten, glucose, lipiden en urine-albuminurie, afhankelijk
van medicatie, risico en controle-interval. Een beperkt algemeen
laboratoriumonderzoek kan passen bij aspecifieke klachten, maar forceer geen
volledig bloedbeeld of breed pakket zonder indicatie. Een aanvraag of relevante
uitslag moet terugkomen in een passende O- of P-regel. Laat een afwijkende
uitslag die beleid vereist logisch doorwerken in een E- en P-regel en eventueel
in de medicatielijst. Gebruik alleen medisch passende combinaties van waarden,
eenheden en referentiewaarden.

Gebruik datums in de vorm JJJJ-MM-DD. Voeg geen uitleg buiten de gevraagde
gestructureerde uitvoer toe.
""".strip()


SETTINGS_BESTANDSNAAM = "patient_settings_uitgebreid.xlsx"

# script gebruikt
# dezelfde settings, maar genereert eerst een blauwdruk en daarna blokken van
# maximaal vijf jaar. De grenzen zijn centraal vastgelegd voor reproduceerbare
# pilotruns.
MODEL_NAAM = "gpt-5.6"
BLOKGROOTTE_JAREN = 5
MAX_HERSTELPOGINGEN = 2
PIPELINE_VERSIE = "iteratief-v1.11"


RuisNiveau = Literal["Geen", "Laag", "Gemiddeld"]


class EpisodeInstelling(BaseModel):
    """Een episode met chronologische volgorde en optionele patiëntspecificatie."""

    volgorde: int
    episode: str
    patient_specifieke_omschrijving: str = ""
    aantal_deelcontacten: int | None = None


class AllergieInstelling(BaseModel):
    """Een door de gebruiker opgegeven allergie uit het instellingenbestand."""

    allergeen: str
    type_allergie: str = "Model bepaalt"
    reactie: str = "Model bepaalt"
    ernst: str = "Model bepaalt"
    status: str = "Actief"


class MicrobiologieInstelling(BaseModel):
    """Een door de gebruiker gewenste microbiologische bepaling of uitslag."""

    episode: str = "Model bepaalt"
    materiaal: str = "Model bepaalt"
    onderzoek: str = "Model bepaalt"
    gewenste_uitslag: str = "Model bepaalt"
    verwekker: str = "Model bepaalt"


class LaboratoriumInstelling(BaseModel):
    """Een door de gebruiker gewenste laboratoriumbepaling."""

    episode: str = "Model bepaalt"
    aanvraagreden: str = "Model bepaalt"
    bepaling: str
    gewenste_waarde: str = "Model bepaalt"
    eenheid: str = "Model bepaalt"


class PatientInstellingen(BaseModel):
    patient_id: str
    geslacht: str
    leeftijd: int
    aantal_jaren: int
    aantal_deelcontacten: int
    ruis_spelfouten: RuisNiveau
    ruis_afkortingen: RuisNiveau
    ruis_telegramstijl: RuisNiveau
    ruis_interpunctie: RuisNiveau
    ruis_herhaling: RuisNiveau
    allergie_modus: Literal["Automatisch", "Zelf invoeren", "Geen"]
    microbiologie_modus: Literal["Automatisch", "Zelf invoeren", "Geen"]
    laboratorium_modus: Literal["Automatisch", "Zelf invoeren", "Geen"]
    rookstatus: str = "Automatisch"
    alcoholgebruik: str = "Automatisch"
    familieanamnese: str = "Automatisch"
    woonsituatie: str = "Automatisch"
    beroep_functionele_context: str = "Automatisch"
    episodes: list[EpisodeInstelling]
    handmatige_allergieen: list[AllergieInstelling]
    handmatige_microbiologie: list[MicrobiologieInstelling]
    handmatig_laboratorium: list[LaboratoriumInstelling]


def waarde_is_leeg(waarde: object) -> bool:
    """Controleer of een cel uit Excel leeg is."""
    return waarde is None or pd.isna(waarde) or str(waarde).strip() == ""


def lees_tekst(waarde: object) -> str:
    """Zet een optionele Excelwaarde veilig om naar tekst."""
    if waarde_is_leeg(waarde):
        return ""
    return str(waarde).strip()


def lees_geheel_getal(waarde: object, naam: str, minimum: int = 0) -> int:
    """Lees en valideer een geheel getal uit het instellingenbestand."""
    if waarde_is_leeg(waarde):
        raise ValueError(f"Instelling '{naam}' is verplicht.")

    try:
        getal = float(waarde)
    except (TypeError, ValueError) as fout:
        raise ValueError(f"Instelling '{naam}' moet een geheel getal zijn.") from fout

    if not getal.is_integer() or getal < minimum:
        raise ValueError(
            f"Instelling '{naam}' moet een geheel getal van minimaal {minimum} zijn."
        )
    return int(getal)


def valideer_episodecontactaantallen(
    episodes: list[EpisodeInstelling],
    totaal_deelcontacten: int,
) -> None:
    """Controleer de optionele verdeling van deelcontacten over episodes."""
    ingevulde_contactaantallen = [
        episode.aantal_deelcontacten
        for episode in episodes
        if episode.aantal_deelcontacten is not None
    ]
    if ingevulde_contactaantallen and len(ingevulde_contactaantallen) != len(episodes):
        raise ValueError(
            "Vul 'aantal_deelcontacten' op het tabblad 'Episodes' óf voor alle "
            "episodes in, óf laat de volledige kolom leeg voor automatische verdeling."
        )
    if ingevulde_contactaantallen:
        totaal_per_episode = sum(ingevulde_contactaantallen)
        if totaal_per_episode != totaal_deelcontacten:
            raise ValueError(
                "De aantallen deelcontacten op het tabblad 'Episodes' tellen op tot "
                f"{totaal_per_episode}, maar op het tabblad 'Patient' staat "
                f"{totaal_deelcontacten}. Maak deze totalen gelijk."
            )


def normaliseer_modus(waarde: object, naam: str) -> str:
    """Normaliseer een gebruikersvriendelijke keuzewaarde uit Excel."""
    invoer = lees_tekst(waarde).lower()
    omzetting = {
        "automatisch": "Automatisch",
        "zelf invoeren": "Zelf invoeren",
        "zelf": "Zelf invoeren",
        "geen": "Geen",
    }
    if invoer not in omzetting:
        raise ValueError(
            f"Instelling '{naam}' moet Automatisch, Zelf invoeren of Geen zijn."
        )
    return omzetting[invoer]


def normaliseer_ruisniveau(waarde: object, naam: str) -> str:
    """Normaliseer een afzonderlijk ruisniveau uit het instellingenbestand."""
    invoer = lees_tekst(waarde).lower()
    omzetting = {
        "geen": "Geen",
        "laag": "Laag",
        "gemiddeld": "Gemiddeld",
        # 'Matig' uit oudere instellingenbestanden wordt als Gemiddeld gelezen.
        "matig": "Gemiddeld",
    }
    if invoer not in omzetting:
        raise ValueError(
            f"Instelling '{naam}' moet Geen, Laag of Gemiddeld zijn."
        )
    return omzetting[invoer]


def lees_patientinstellingen(instellingenpad: Path) -> PatientInstellingen:
    """Lees patiëntkenmerken en episodes uit patient_settings.xlsx."""
    if not instellingenpad.exists():
        raise FileNotFoundError(
            f"Instellingenbestand niet gevonden: {instellingenpad}. "
            f"Plaats {SETTINGS_BESTANDSNAAM} in dezelfde map als dit Python-bestand."
        )

    try:
        patient_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Patient",
            header=2,
        )
        episodes_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Episodes",
        )
    except ValueError as fout:
        raise ValueError(
            "Het instellingenbestand moet de tabbladen 'Patient' en 'Episodes' bevatten."
        ) from fout

    # Deze twee tabbladen zijn alleen verplicht wanneer de bijbehorende modus
    # op 'Zelf invoeren' staat. Anders mogen ze leeg zijn of ontbreken.
    try:
        allergie_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Allergieen",
        )
    except ValueError:
        allergie_dataframe = pd.DataFrame()

    try:
        microbiologie_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Microbiologie",
        )
    except ValueError:
        microbiologie_dataframe = pd.DataFrame()

    try:
        laboratorium_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Laboratorium",
        )
    except ValueError:
        laboratorium_dataframe = pd.DataFrame()

    vereiste_patientkolommen = {"Instelling", "Waarde"}
    if not vereiste_patientkolommen.issubset(patient_dataframe.columns):
        raise ValueError(
            "Het tabblad 'Patient' moet de kolommen 'Instelling' en 'Waarde' bevatten."
        )

    waarden = {}
    for _, rij in patient_dataframe.iterrows():
        sleutel = lees_tekst(rij.get("Instelling"))
        if sleutel:
            waarden[sleutel] = rij.get("Waarde")

    verplichte_sleutels = {
        "patient_id",
        "geslacht",
        "leeftijd",
        "aantal_jaren",
        "aantal_deelcontacten",
        "ruis_spelfouten",
        "ruis_afkortingen",
        "ruis_telegramstijl",
        "ruis_interpunctie",
        "ruis_herhaling",
        "allergie_modus",
        "microbiologie_modus",
        "laboratorium_modus",
    }
    ontbrekend = sorted(verplichte_sleutels - waarden.keys())
    if ontbrekend:
        raise ValueError(
            "Deze verplichte instellingen ontbreken op het tabblad 'Patient': "
            + ", ".join(ontbrekend)
        )

    totaal_deelcontacten = lees_geheel_getal(
        waarden["aantal_deelcontacten"],
        "aantal_deelcontacten",
        minimum=1,
    )

    vereiste_episodekolommen = {
        "volgorde",
        "episode",
        "patient_specifieke_omschrijving",
    }
    if not vereiste_episodekolommen.issubset(episodes_dataframe.columns):
        raise ValueError(
            "Het tabblad 'Episodes' moet de kolommen 'volgorde', 'episode' en "
            "'patient_specifieke_omschrijving' bevatten."
        )

    episodes = []
    for _, rij in episodes_dataframe.iterrows():
        episodenaam = lees_tekst(rij.get("episode"))
        if not episodenaam:
            continue
        episodes.append(
            EpisodeInstelling(
                volgorde=lees_geheel_getal(
                    rij.get("volgorde"),
                    "volgorde op het tabblad Episodes",
                    minimum=1,
                ),
                episode=episodenaam,
                patient_specifieke_omschrijving=lees_tekst(
                    rij.get("patient_specifieke_omschrijving")
                ),
                aantal_deelcontacten=(
                    None
                    if waarde_is_leeg(rij.get("aantal_deelcontacten"))
                    else lees_geheel_getal(
                        rij.get("aantal_deelcontacten"),
                        "aantal_deelcontacten op het tabblad Episodes",
                        minimum=1,
                    )
                ),
            )
        )

    if not episodes:
        raise ValueError("Vul minimaal één episode in op het tabblad 'Episodes'.")

    volgordes = [episode.volgorde for episode in episodes]
    if len(volgordes) != len(set(volgordes)):
        raise ValueError(
            "Iedere episode moet op het tabblad 'Episodes' een unieke volgorde hebben."
        )
    episodes.sort(key=lambda episode: episode.volgorde)

    valideer_episodecontactaantallen(episodes, totaal_deelcontacten)

    patient_id = lees_tekst(waarden["patient_id"])
    geslacht = lees_tekst(waarden["geslacht"])
    if not patient_id:
        raise ValueError("Instelling 'patient_id' mag niet leeg zijn.")
    if not geslacht:
        raise ValueError("Instelling 'geslacht' mag niet leeg zijn.")

    allergie_modus = normaliseer_modus(
        waarden["allergie_modus"],
        "allergie_modus",
    )
    microbiologie_modus = normaliseer_modus(
        waarden["microbiologie_modus"],
        "microbiologie_modus",
    )
    laboratorium_modus = normaliseer_modus(
        waarden["laboratorium_modus"],
        "laboratorium_modus",
    )

    handmatige_allergieen = []
    if allergie_modus == "Zelf invoeren":
        vereiste_kolom = {"allergeen"}
        if not vereiste_kolom.issubset(allergie_dataframe.columns):
            raise ValueError(
                "Bij allergie_modus 'Zelf invoeren' moet het tabblad "
                "'Allergieen' minimaal de kolom 'allergeen' bevatten."
            )

        for _, rij in allergie_dataframe.iterrows():
            allergeen = lees_tekst(rij.get("allergeen"))
            if not allergeen:
                continue
            handmatige_allergieen.append(
                AllergieInstelling(
                    allergeen=allergeen,
                    type_allergie=lees_tekst(rij.get("type_allergie"))
                    or "Model bepaalt",
                    reactie=lees_tekst(rij.get("reactie")) or "Model bepaalt",
                    ernst=lees_tekst(rij.get("ernst")) or "Model bepaalt",
                    status=lees_tekst(rij.get("status")) or "Actief",
                )
            )

        if not handmatige_allergieen:
            raise ValueError(
                "allergie_modus staat op 'Zelf invoeren', maar op het tabblad "
                "'Allergieen' is geen allergeen ingevuld."
            )

    handmatige_microbiologie = []
    if microbiologie_modus == "Zelf invoeren":
        mogelijke_kolommen = {
            "episode",
            "materiaal",
            "onderzoek",
            "gewenste_uitslag",
            "verwekker",
        }
        if microbiologie_dataframe.empty or not mogelijke_kolommen.intersection(
            microbiologie_dataframe.columns
        ):
            raise ValueError(
                "Bij microbiologie_modus 'Zelf invoeren' moet het tabblad "
                "'Microbiologie' de daarvoor bestemde kolommen bevatten."
            )

        for _, rij in microbiologie_dataframe.iterrows():
            waarden_rij = {
                kolom: lees_tekst(rij.get(kolom)) for kolom in mogelijke_kolommen
            }
            if not any(waarden_rij.values()):
                continue
            handmatige_microbiologie.append(
                MicrobiologieInstelling(
                    episode=waarden_rij["episode"] or "Model bepaalt",
                    materiaal=waarden_rij["materiaal"] or "Model bepaalt",
                    onderzoek=waarden_rij["onderzoek"] or "Model bepaalt",
                    gewenste_uitslag=waarden_rij["gewenste_uitslag"]
                    or "Model bepaalt",
                    verwekker=waarden_rij["verwekker"] or "Model bepaalt",
                )
            )

        if not handmatige_microbiologie:
            raise ValueError(
                "microbiologie_modus staat op 'Zelf invoeren', maar het tabblad "
                "'Microbiologie' bevat geen ingevulde regel."
            )

    handmatig_laboratorium = []
    if laboratorium_modus == "Zelf invoeren":
        if "bepaling" not in laboratorium_dataframe.columns:
            raise ValueError(
                "Bij laboratorium_modus 'Zelf invoeren' moet het tabblad "
                "'Laboratorium' minimaal de kolom 'bepaling' bevatten."
            )

        for _, rij in laboratorium_dataframe.iterrows():
            bepaling = lees_tekst(rij.get("bepaling"))
            if not bepaling:
                continue
            handmatig_laboratorium.append(
                LaboratoriumInstelling(
                    episode=lees_tekst(rij.get("episode")) or "Model bepaalt",
                    aanvraagreden=lees_tekst(rij.get("aanvraagreden"))
                    or "Model bepaalt",
                    bepaling=bepaling,
                    gewenste_waarde=lees_tekst(rij.get("gewenste_waarde"))
                    or "Model bepaalt",
                    eenheid=lees_tekst(rij.get("eenheid")) or "Model bepaalt",
                )
            )

        if not handmatig_laboratorium:
            raise ValueError(
                "laboratorium_modus staat op 'Zelf invoeren', maar op het tabblad "
                "'Laboratorium' is geen bepaling ingevuld."
            )

    return PatientInstellingen(
        patient_id=patient_id,
        geslacht=geslacht,
        leeftijd=lees_geheel_getal(
            waarden["leeftijd"],
            "leeftijd",
            minimum=0,
        ),
        aantal_jaren=lees_geheel_getal(
            waarden["aantal_jaren"],
            "aantal_jaren",
            minimum=1,
        ),
        aantal_deelcontacten=totaal_deelcontacten,
        ruis_spelfouten=normaliseer_ruisniveau(
            waarden["ruis_spelfouten"],
            "ruis_spelfouten",
        ),
        ruis_afkortingen=normaliseer_ruisniveau(
            waarden["ruis_afkortingen"],
            "ruis_afkortingen",
        ),
        ruis_telegramstijl=normaliseer_ruisniveau(
            waarden["ruis_telegramstijl"],
            "ruis_telegramstijl",
        ),
        ruis_interpunctie=normaliseer_ruisniveau(
            waarden["ruis_interpunctie"],
            "ruis_interpunctie",
        ),
        ruis_herhaling=normaliseer_ruisniveau(
            waarden["ruis_herhaling"],
            "ruis_herhaling",
        ),
        allergie_modus=allergie_modus,
        microbiologie_modus=microbiologie_modus,
        laboratorium_modus=laboratorium_modus,
        rookstatus=lees_tekst(waarden.get("rookstatus")) or "Automatisch",
        alcoholgebruik=lees_tekst(waarden.get("alcoholgebruik")) or "Automatisch",
        familieanamnese=(
            lees_tekst(waarden.get("familieanamnese")) or "Automatisch"
        ),
        woonsituatie=lees_tekst(waarden.get("woonsituatie")) or "Automatisch",
        beroep_functionele_context=(
            lees_tekst(waarden.get("beroep_functionele_context"))
            or "Automatisch"
        ),
        episodes=episodes,
        handmatige_allergieen=handmatige_allergieen,
        handmatige_microbiologie=handmatige_microbiologie,
        handmatig_laboratorium=handmatig_laboratorium,
    )


def bepaal_dossierperiode(aantal_jaren: int) -> tuple[str, str]:
    """Bereken automatisch een periode die eindigt op de uitvoerdatum."""
    einddatum = date.today()
    try:
        startdatum = einddatum.replace(year=einddatum.year - aantal_jaren)
    except ValueError:
        # Alleen relevant als het script op 29 februari wordt uitgevoerd.
        startdatum = einddatum.replace(
            year=einddatum.year - aantal_jaren,
            day=28,
        )
    return startdatum.isoformat(), einddatum.isoformat()


def bepaal_fictieve_geboortedatum(
    patient_id: str,
    leeftijd: int,
    einddatum: str,
) -> str:
    """Maak reproduceerbaar een fictieve geboortedatum die exact bij de leeftijd past."""
    einde = date.fromisoformat(einddatum)
    digest = hashlib.sha256(patient_id.encode("utf-8")).digest()
    maand = digest[0] % 12 + 1
    maximaal_aantal_dagen = calendar.monthrange(2000, maand)[1]
    dag = digest[1] % maximaal_aantal_dagen + 1

    geboortejaar = einde.year - leeftijd
    if (maand, dag) > (einde.month, einde.day):
        geboortejaar -= 1
    if maand == 2 and dag == 29 and not calendar.isleap(geboortejaar):
        dag = 28

    return date(geboortejaar, maand, dag).isoformat()


def bereken_leeftijd_op_datum(geboortedatum: str, peildatum: str) -> int:
    """Bereken de leeftijd op een specifieke datum zonder modelinterpretatie."""
    geboorte = date.fromisoformat(geboortedatum)
    peil = date.fromisoformat(peildatum)
    return peil.year - geboorte.year - (
        (peil.month, peil.day) < (geboorte.month, geboorte.day)
    )


def maak_leeftijdskader(
    instellingen: PatientInstellingen,
    startdatum: str,
    einddatum: str,
) -> str:
    """Leg de juiste leeftijdsontwikkeling voor één tijdsblok expliciet vast."""
    dossierstart, dossiereinde = bepaal_dossierperiode(
        instellingen.aantal_jaren
    )
    geboortedatum = bepaal_fictieve_geboortedatum(
        instellingen.patient_id,
        instellingen.leeftijd,
        dossiereinde,
    )
    leeftijd_blokstart = bereken_leeftijd_op_datum(
        geboortedatum,
        startdatum,
    )
    leeftijd_blokeinde = bereken_leeftijd_op_datum(
        geboortedatum,
        einddatum,
    )
    leeftijd_dossiereinde = bereken_leeftijd_op_datum(
        geboortedatum,
        dossiereinde,
    )
    if leeftijd_dossiereinde != instellingen.leeftijd:
        raise RuntimeError(
            "De deterministisch berekende geboortedatum past niet bij de "
            "ingestelde leeftijd aan het einde van het dossier."
        )

    return f"""
- fictieve geboortedatum: {geboortedatum}
- volledige dossierperiode: {dossierstart} tot en met {dossiereinde}
- ingestelde leeftijd op de einddatum van het volledige dossier: {instellingen.leeftijd} jaar
- huidig tijdsblok: {startdatum} tot en met {einddatum}
- leeftijd op de eerste dag van dit blok: {leeftijd_blokstart} jaar
- leeftijd op de laatste dag van dit blok: {leeftijd_blokeinde} jaar

De instelling 'leeftijd' is dus uitsluitend de leeftijd op {dossiereinde} en
is geen vaste leeftijd in eerdere jaren. Bereken bij een gebeurtenis zo nodig
de leeftijd uit de geboortedatum en de datum van die gebeurtenis. Een lagere
leeftijd in een eerder tijdsblok is correct en mag niet als inconsistentie
worden aangemerkt.
""".strip()


def maak_achtergrondregel(label: str, waarde: str) -> str:
    """Maak een promptregel voor een automatisch of handmatig achtergrondgegeven."""
    if waarde.strip().lower() == "automatisch":
        return (
            f"- {label}: bepaal automatisch een realistische waarde die past bij "
            "de patiënt en het klinische verloop"
        )
    return f"- {label}: gebruik verplicht de volgende specificatie: {waarde}"


def maak_patient_prompt(instellingen: PatientInstellingen) -> str:
    """Bouw de variabele patiëntenprompt uit het Excel-instellingenbestand."""
    startdatum, einddatum = bepaal_dossierperiode(instellingen.aantal_jaren)
    geboortedatum = bepaal_fictieve_geboortedatum(
        instellingen.patient_id,
        instellingen.leeftijd,
        einddatum,
    )
    episode_regels = []
    for episode in instellingen.episodes:
        regel = f"{episode.volgorde}. {episode.episode}"
        if episode.patient_specifieke_omschrijving:
            regel += (
                " | patiëntspecifieke omschrijving: "
                f"{episode.patient_specifieke_omschrijving}"
            )
        if episode.aantal_deelcontacten is not None:
            regel += (
                " | exact aantal deelcontacten: "
                f"{episode.aantal_deelcontacten}"
            )
        episode_regels.append(regel)

    contactaantallen_ingevuld = all(
        episode.aantal_deelcontacten is not None
        for episode in instellingen.episodes
    )
    if contactaantallen_ingevuld:
        contactverdeling_instructie = (
            "Gebruik voor iedere episode exact het opgegeven aantal deelcontacten. "
            f"Deze aantallen tellen samen op tot {instellingen.aantal_deelcontacten}. "
            "Koppel ieder deelcontact in het veld 'episode' aan precies één primaire "
            "episode uit de opgegeven lijst. Een contact waarin meerdere problemen "
            "worden besproken telt slechts eenmaal, bij de primaire episode. Maak in "
            f"totaal exact {instellingen.aantal_deelcontacten} deelcontacten."
        )
    else:
        contactverdeling_instructie = (
            "Er zijn geen aantallen per episode opgegeven. Verdeel daarom het totale "
            "aantal deelcontacten zelf medisch logisch over de episodes. Maak in "
            f"totaal exact {instellingen.aantal_deelcontacten} deelcontacten."
        )

    ruisinstructies = {
        "spelfouten": {
            "Geen": "Gebruik geen opzettelijke spelfouten of typefouten.",
            "Laag": (
                "Gebruik incidenteel een kleine, realistische type- of spelfout, "
                "terwijl ieder woord goed te reconstrueren blijft."
            ),
            "Gemiddeld": (
                "Gebruik verspreid door het dossier regelmatig kleine, realistische "
                "type- of spelfouten, zonder medische termen onherkenbaar te maken."
            ),
        },
        "afkortingen": {
            "Geen": "Schrijf termen in de vrije tekst zo veel mogelijk voluit.",
            "Laag": (
                "Gebruik incidenteel een gangbare Nederlandse medische afkorting."
            ),
            "Gemiddeld": (
                "Gebruik geregeld gangbare Nederlandse huisartsafkortingen, maar "
                "zorg dat de betekenis uit de context duidelijk blijft."
            ),
        },
        "telegramstijl": {
            "Geen": "Gebruik grammaticaal volledige zinnen.",
            "Laag": (
                "Gebruik incidenteel korte zinsdelen of weggelaten lidwoorden zoals "
                "in een realistische huisartsnotitie."
            ),
            "Gemiddeld": (
                "Gebruik geregeld beknopte telegramstijl en korte zinsdelen, zonder "
                "klinisch relevante informatie weg te laten."
            ),
        },
        "interpunctie en hoofdletters": {
            "Geen": "Gebruik consequente interpunctie en hoofdletters.",
            "Laag": (
                "Laat incidenteel een punt of hoofdletter weg in vrije tekst."
            ),
            "Gemiddeld": (
                "Gebruik geregeld licht inconsistente interpunctie of hoofdletters, "
                "maar behoud de leesbaarheid."
            ),
        },
        "herhaling": {
            "Geen": "Vermijd onnodige herhaling van eerder vastgelegde informatie.",
            "Laag": (
                "Herhaal incidenteel een korte relevante voorgeschiedenis of een "
                "eerder beleidspunt in een later deelcontact."
            ),
            "Gemiddeld": (
                "Laat regelmatig beperkte, realistische herhaling van relevante "
                "voorgeschiedenis of beleid voorkomen, zonder volledige notities te "
                "dupliceren."
            ),
        },
    }

    gekozen_ruis = [
        ("Spelfouten", instellingen.ruis_spelfouten, "spelfouten"),
        ("Afkortingen", instellingen.ruis_afkortingen, "afkortingen"),
        ("Telegramstijl", instellingen.ruis_telegramstijl, "telegramstijl"),
        (
            "Interpunctie en hoofdletters",
            instellingen.ruis_interpunctie,
            "interpunctie en hoofdletters",
        ),
        ("Herhaling", instellingen.ruis_herhaling, "herhaling"),
    ]
    ruis_regels = []
    for label, niveau, sleutel in gekozen_ruis:
        ruis_regels.append(
            f"- {label} — {niveau}: {ruisinstructies[sleutel][niveau]}"
        )

    achtergrond_regels = [
        maak_achtergrondregel("rookstatus", instellingen.rookstatus),
        maak_achtergrondregel("alcoholgebruik", instellingen.alcoholgebruik),
        maak_achtergrondregel("familieanamnese", instellingen.familieanamnese),
        maak_achtergrondregel("woonsituatie", instellingen.woonsituatie),
        maak_achtergrondregel(
            "beroep, dagbesteding en functionele context",
            instellingen.beroep_functionele_context,
        ),
    ]

    if instellingen.allergie_modus == "Automatisch":
        allergie_instructie = (
            "Bepaal zelf of een klinisch relevante allergie bij de casus past. "
            "Het allergieoverzicht mag leeg zijn. Als je een allergie opneemt, "
            "laat deze dan consistent terugkomen in het voorschrijfbeleid en de "
            "medicatielijst."
        )
    elif instellingen.allergie_modus == "Geen":
        allergie_instructie = (
            "Neem geen allergieën op. Geef voor allergieen een lege lijst terug."
        )
    else:
        allergieregels = []
        for nummer, allergie in enumerate(
            instellingen.handmatige_allergieen,
            start=1,
        ):
            allergieregels.append(
                f"{nummer}. allergeen={allergie.allergeen}; "
                f"type={allergie.type_allergie}; reactie={allergie.reactie}; "
                f"ernst={allergie.ernst}; status={allergie.status}"
            )
        allergie_instructie = (
            "Neem de onderstaande allergieën verplicht op en voeg geen andere "
            "allergieën toe. Kies alleen voor velden met 'Model bepaalt' zelf "
            "een passende waarde. Kies een passende registratiedatum binnen de "
            "dossierperiode en stem medicatie en beleid hierop af.\n"
            + "\n".join(allergieregels)
        )

    if instellingen.microbiologie_modus == "Automatisch":
        microbiologie_instructie = (
            "Genereer alleen microbiologische uitslagen als de episodes of het "
            "klinische beloop daar aanleiding toe geven. De microbiologielijst "
            "mag leeg zijn."
        )
    elif instellingen.microbiologie_modus == "Geen":
        microbiologie_instructie = (
            "Neem geen microbiologische uitslagen op. Geef voor microbiologie "
            "een lege lijst terug."
        )
    else:
        microbiologieregels = []
        for nummer, bepaling in enumerate(
            instellingen.handmatige_microbiologie,
            start=1,
        ):
            microbiologieregels.append(
                f"{nummer}. episode={bepaling.episode}; materiaal={bepaling.materiaal}; "
                f"onderzoek={bepaling.onderzoek}; "
                f"gewenste uitslag={bepaling.gewenste_uitslag}; "
                f"verwekker={bepaling.verwekker}"
            )
        microbiologie_instructie = (
            "Verwerk de onderstaande microbiologische bepalingen verplicht en "
            "voeg geen andere bepalingen toe. Kies alleen voor velden met "
            "'Model bepaalt' zelf een passende waarde. Kies passende datums, "
            "hoeveelheid/groei en gevoeligheid/resistentie en laat alles "
            "aansluiten op journaal en behandelbeleid.\n"
            + "\n".join(microbiologieregels)
        )

    if instellingen.laboratorium_modus == "Automatisch":
        laboratorium_instructie = (
            "Genereer laboratoriumuitslagen wanneer deze passen bij diagnostiek, "
            "medicatiebewaking of controles van de opgegeven episodes. Maak van "
            "ieder prikmoment een herkenbare aanvraag met een aanvraag-ID zoals "
            "L001. Gebruik voor alle bepalingen uit hetzelfde prikmoment dezelfde "
            "datum en aanvraagreden. Genereer per aanvraag een realistisch pakket "
            "van meerdere samenhangende bepalingen in plaats van uitsluitend één "
            "episode-specifieke waarde. Neem ook normale nevenbepalingen op die bij "
            "hetzelfde pakket plausibel zijn aangevraagd. Combineer controles voor "
            "meerdere aandoeningen zo nodig in één prikmoment. Laat bij chronische "
            "aandoeningen passende meetmomenten terugkomen, zonder onnodige "
            "bepalingen of onrealistisch veel uitslagen toe te voegen. Bij alleen "
            "hypertensie is een volledig bloedbeeld niet standaard verplicht; "
            "kies het pakket op basis van cardiovasculair risico, medicatie en "
            "eventuele bijkomende klachten. De laboratoriumlijst mag leeg zijn als "
            "geen enkele episode daar aanleiding toe geeft."
        )
    elif instellingen.laboratorium_modus == "Geen":
        laboratorium_instructie = (
            "Neem geen laboratoriumuitslagen op. Geef voor laboratorium een lege "
            "lijst terug."
        )
    else:
        laboratoriumregels = []
        for nummer, bepaling in enumerate(
            instellingen.handmatig_laboratorium,
            start=1,
        ):
            laboratoriumregels.append(
                f"{nummer}. episode={bepaling.episode}; "
                f"aanvraagreden={bepaling.aanvraagreden}; "
                f"bepaling={bepaling.bepaling}; "
                f"gewenste waarde={bepaling.gewenste_waarde}; "
                f"eenheid={bepaling.eenheid}"
            )
        laboratorium_instructie = (
            "Verwerk de onderstaande laboratoriumbepalingen verplicht en voeg "
            "geen andere bepalingen toe. Kies alleen voor velden met 'Model "
            "bepaalt' zelf een passende waarde. Groepeer bepalingen die medisch bij "
            "hetzelfde prikmoment passen onder één aanvraag-ID, datum en "
            "aanvraagreden. Kies passende datums en referentiewaarden en laat "
            "aanvragen, uitslagen en eventueel beleid aansluiten op het journaal.\n"
            + "\n".join(laboratoriumregels)
        )

    return f"""
Maak een synthetisch huisartsendossier voor:

- fictieve patiënt-ID: {instellingen.patient_id}
- geslacht: {instellingen.geslacht}
- fictieve geboortedatum: {geboortedatum}
- leeftijd aan het einde van de dossierperiode: {instellingen.leeftijd} jaar
- duur van het dossier: {instellingen.aantal_jaren} jaar
- automatisch berekende periode: {startdatum} tot en met {einddatum}
- gewenst aantal deelcontacten: {instellingen.aantal_deelcontacten}

Episodes:
{chr(10).join(episode_regels)}

Persoons- en achtergrondgegevens:
{chr(10).join(achtergrond_regels)}

Neem de patiënt-ID, geboortedatum en het formele geslacht exact over. De
opgegeven leeftijd geldt alleen op de einddatum van het volledige dossier.
Bereken de leeftijd op eerdere contact- en peildatums steeds uit de fictieve
geboortedatum; gebruik de eindleeftijd nooit als vaste leeftijd in alle jaren.
Genereer voor iedere opgegeven episode precies één regel in de episodelijst.
Behoud daarin exact de opgegeven volgorde en episodenaam. Kies een passende
ICPC-code, begin- en eventuele einddatum, status, attentiewaarde, samenvatting
en beleid die volledig aansluiten op het journaal.

De kolom 'volgorde' uit het instellingenbestand bepaalt de chronologische
volgorde waarin de episodes voor het eerst optreden of voor het eerst in het
dossier worden geregistreerd. Respecteer deze volgorde. Een patiëntspecifieke
omschrijving is een bindende specificatie van de betreffende episode en moet
consistent terugkomen in relevante SOEP-regels, bevindingen, medicatie,
correspondentie en uitslagen.

Verdeling van deelcontacten:
{contactverdeling_instructie}

Genereer een realistisch longitudinaal huisartsendossier over de volledige
automatisch berekende periode. Verdeel het gewenste aantal deelcontacten
chronologisch en medisch logisch. Het aantal contacten hoeft niet gelijk over
alle jaren verdeeld te zijn; bij een klein testaantal mogen sommige jaren
weinig contacten bevatten. Laat ontwikkelingen uit eerdere contacten logisch
terugkomen in latere contacten. Gebruik per deelcontact precies één S-, O-, E-
en P-regel.

BELANGRIJK: koppel ieder deelcontact aan precies één primaire episode. Vul in
het veld 'episode' letterlijk en ongewijzigd precies één episodenaam uit de
hierboven opgegeven episodelijst in. Combineer in dit veld nooit meerdere
episodenamen met een komma, puntkomma, schuine streep of het woord 'en'. Als
meerdere gezondheidsproblemen tijdens één werkelijk contact worden besproken,
registreer die als afzonderlijke deelcontacten, ieder gekoppeld aan één primaire
episode. Houd daarbij het totale gewenste aantal deelcontacten gelijk.

Bepaal zelf het realistische beloop van iedere opgegeven episode. Een episode
mag bijvoorbeeld chronisch actief, eenmalig of later afgesloten zijn. Laat
chronische episodes terugkomen in passende controles en voeg enkele passende
tussentijdse contacten toe.

Maak een medicatielijst die volledig aansluit op het journaal. Geef bij actieve
medicatie een lege einddatum en bij gestopte of eenmalige medicatie een passende
einddatum. Laat iedere start, wijziging of stop ook terugkomen in het journaal.

Genereer alleen verwijstrajecten als deze klinisch logisch volgen uit de
opgegeven episodes en het verloop van de klachten. Forceer geen vast aantal.
Maak bij iedere verwijzing één verwijsbrief en één latere specialistenbrief met
hetzelfde traject-ID. Gebruik bij een acute ziekenhuisopname zonder voorafgaande
huisartsverwijzing één latere ontslagbrief en verzin daarvoor geen
verwijsbrief. De lijst met correspondentie mag leeg zijn als geen verwijzing of
opname nodig is.

Allergieën — gekozen modus: {instellingen.allergie_modus}
{allergie_instructie}

Microbiologie — gekozen modus: {instellingen.microbiologie_modus}
{microbiologie_instructie}

Laboratorium — gekozen modus: {instellingen.laboratorium_modus}
{laboratorium_instructie}

Stijl en ruis — uitsluitend voor de S-, O-, E- en P-regels van het journaal:
{chr(10).join(ruis_regels)}

Pas de bovenstaande ruis uitsluitend toe op de vrije tekst in de S-, O-, E- en
P-regels van deelcontacten. Pas nooit ruis toe op verwijsbrieven,
specialistenbrieven of ontslagbrieven, ongeacht de ingestelde ruisniveaus.
Schrijf alle brieven foutloos, professioneel en in volledige zinnen met
consequente spelling, grammatica, hoofdletters en interpunctie. Verander door
ruis nooit patiënt-ID's, datums, geneesmiddelnamen, doseringen, meetwaarden,
eenheden, uitslagen, diagnoses of traject-ID's. Ruis mag geen medische
onjuistheden, tegenstrijdigheden of klinische onduidelijkheid veroorzaken.

Zorg dat patiëntgegevens, episodelijst, journaal, medicatie, correspondentie,
allergieën, microbiologie en laboratorium onderling consistent zijn.
""".strip()


def veilige_bestandsnaam(waarde: str) -> str:
    """Maak van de fictieve patiënt-ID een veilige bestandsnaam."""
    veilig = re.sub(r"[^A-Za-z0-9_-]+", "_", waarde.strip())
    return veilig.strip("_") or "patient"


class PatientAchtergrond(BaseModel):
    rookstatus: str = Field(
        description="Rookstatus, bijvoorbeeld nooit gerookt, voormalig roker of roker"
    )
    alcoholgebruik: str = Field(
        description="Beknopte omschrijving van alcoholgebruik in eenheden per week"
    )
    familieanamnese: str = Field(
        description="Klinisch relevante familieanamnese of expliciet geen bijzonderheden"
    )
    woonsituatie: str = Field(
        description="Woonsituatie en relevante sociale ondersteuning"
    )
    beroep_dagbesteding: str = Field(
        description="Beroep, pensioenstatus of relevante dagelijkse bezigheden"
    )
    functionele_context: str = Field(
        description="Beknopte functionele context, bijvoorbeeld mobiliteit en zelfstandigheid"
    )


class EpisodeOverzichtRegel(BaseModel):
    volgorde: int = Field(description="Volgorde uit de opgegeven episodelijst")
    episode: str = Field(description="Exacte episodenaam uit de patiëntenprompt")
    icpc_code: str = Field(description="Passende ICPC-code, bijvoorbeeld T90.02")
    startdatum: str = Field(description="Begindatum in JJJJ-MM-DD")
    einddatum: str = Field(
        description="Einddatum in JJJJ-MM-DD; leeg laten als de episode actief is"
    )
    status: Literal["Actief", "Afgesloten"]
    attentiewaarde: Literal["Ja", "Nee"]
    samenvatting_beloop: str = Field(
        description="Beknopte samenvatting van het relevante klinische beloop"
    )
    beleid: str = Field(description="Actueel of afsluitend beleid voor deze episode")


class Deelcontact(BaseModel):
    contact_id: str = Field(description="Uniek contactnummer, bijvoorbeeld C001")
    datum: str = Field(description="Datum in JJJJ-MM-DD")
    zorgverlener: Literal["Huisarts", "POH-S", "Doktersassistente"]
    contactvorm: Literal[
        "Praktijkconsult", "Telefonisch", "Huisbezoek", "Administratief"
    ]
    episode: str
    s: str = Field(description="Subjectieve SOEP-regel")
    o: str = Field(description="Objectieve SOEP-regel")
    e: str = Field(description="Evaluatie met passende diagnose en eventueel ICPC-code")
    p: str = Field(description="Plan en beleid")


class MedicatieRegel(BaseModel):
    geneesmiddel: str = Field(description="Naam van het geneesmiddel")
    sterkte: str = Field(description="Sterkte, bijvoorbeeld 500 mg")
    dosering: str = Field(description="Dosering en frequentie, bijvoorbeeld 2dd1")
    indicatie: str = Field(description="Episode of reden waarvoor het middel wordt gebruikt")
    startdatum: str = Field(description="Startdatum in JJJJ-MM-DD")
    einddatum: str = Field(
        description="Einddatum in JJJJ-MM-DD; leeg laten als de medicatie actief is"
    )
    status: Literal["Actief", "Gestopt", "Eenmalig"]


class CorrespondentieRegel(BaseModel):
    traject_id: str = Field(
        description=(
            "Gedeeld trajectnummer voor een regulier briefkoppel of uniek nummer "
            "voor één acute ontslagbrief, bijvoorbeeld T001"
        )
    )
    datum: str = Field(description="Datum van de brief in JJJJ-MM-DD")
    type_brief: Literal["Verwijsbrief", "Specialistenbrief", "Ontslagbrief"]
    specialisme: str = Field(description="Betrokken specialisme, bijvoorbeeld Urologie")
    episode: str = Field(description="Episode waarop de correspondentie betrekking heeft")
    van: str = Field(description="Afzender, bijvoorbeeld Huisarts of Uroloog")
    aan: str = Field(description="Ontvanger, bijvoorbeeld Uroloog of Huisarts")
    onderwerp: str = Field(description="Beknopt onderwerp van de brief")
    inhoud: str = Field(
        description=(
            "Volledige fictieve brieftekst; niet alleen een samenvatting. "
            "Altijd professioneel en foutloos geschreven, zonder de ruis uit "
            "de SOEP-notities"
        )
    )


class AllergieRegel(BaseModel):
    allergeen: str = Field(description="Stof of geneesmiddel waarvoor de allergie bestaat")
    type_allergie: Literal["Geneesmiddel", "Voedsel", "Omgeving", "Overig"]
    reactie: str = Field(description="Klinische reactie, bijvoorbeeld huiduitslag")
    ernst: Literal["Mild", "Matig", "Ernstig", "Onbekend"]
    registratiedatum: str = Field(description="Registratiedatum in JJJJ-MM-DD")
    status: Literal["Actief", "Inactief"]


class MicrobiologieRegel(BaseModel):
    datum: str = Field(description="Datum van de microbiologische uitslag in JJJJ-MM-DD")
    episode: str = Field(description="Episode waarop de uitslag betrekking heeft")
    materiaal: str = Field(description="Onderzocht materiaal, bijvoorbeeld urine")
    onderzoek: str = Field(description="Type onderzoek, bijvoorbeeld urinekweek")
    uitslag: str = Field(description="Hoofduitslag, bijvoorbeeld positief of negatief")
    verwekker: str = Field(
        description="Aangetoonde verwekker; leeg laten wanneer geen verwekker is gevonden"
    )
    hoeveelheid: str = Field(
        description="Hoeveelheid of groei indien relevant; anders leeg laten"
    )
    gevoeligheid_resistentie: str = Field(
        description="Beknopt en klinisch relevant gevoeligheids- of resistentiepatroon"
    )
    conclusie: str = Field(description="Beknopte klinische conclusie van de uitslag")


class LaboratoriumRegel(BaseModel):
    aanvraag_id: str = Field(
        description=(
            "Gedeeld nummer voor alle bepalingen uit hetzelfde prikmoment, "
            "bijvoorbeeld L001"
        )
    )
    datum: str = Field(description="Datum van de laboratoriumuitslag in JJJJ-MM-DD")
    aanvraagreden: str = Field(
        description=(
            "Reden of type aanvraag, bijvoorbeeld jaarlijkse CVRM-controle of "
            "algemeen onderzoek wegens vermoeidheid"
        )
    )
    episode: str = Field(description="Episode waarop de uitslag betrekking heeft")
    bepaling: str = Field(description="Naam van de bepaling, bijvoorbeeld HbA1c")
    waarde: str = Field(description="Gemeten waarde zonder eenheid")
    eenheid: str = Field(description="Bijpassende eenheid, bijvoorbeeld mmol/mol")
    referentiewaarde: str = Field(
        description="Passende referentiewaarde of streefwaarde als tekst"
    )
    afwijking: Literal["Laag", "Normaal", "Hoog", "Niet van toepassing"]
    conclusie: str = Field(description="Beknopte klinische duiding van de uitslag")


class SynthetischDossier(BaseModel):
    patient_achtergrond: PatientAchtergrond
    episodelijst: list[EpisodeOverzichtRegel]
    contacten: list[Deelcontact]
    medicatie: list[MedicatieRegel]
    correspondentie: list[CorrespondentieRegel]
    allergieen: list[AllergieRegel]
    microbiologie: list[MicrobiologieRegel]
    laboratorium: list[LaboratoriumRegel]


# ---------------------------------------------------------------------------
# Gestructureerde tussenproducten van de iteratieve pipeline
# ---------------------------------------------------------------------------


class EpisodeOntwerp(BaseModel):
    """Vooraf ontworpen hoofdlijn van één episode."""

    volgorde: int
    episode: str
    icpc_code: str
    geplande_startdatum: str = Field(description="Datum in JJJJ-MM-DD")
    geplande_einddatum: str = Field(
        description="Datum in JJJJ-MM-DD; leeg als de episode actief eindigt"
    )
    status_einde_dossier: Literal["Actief", "Afgesloten"]
    attentiewaarde: Literal["Ja", "Nee"]
    kernbeloop: str
    verwacht_beleid: str


class EpisodeContactVerdeling(BaseModel):
    """Aantal geplande deelcontacten voor één episode binnen één tijdsblok."""

    episode: str
    aantal_deelcontacten: int = Field(ge=0)


class BlokKader(BaseModel):
    """Globale verdeling van contacten en gebeurtenissen over één tijdsblok."""

    bloknummer: int
    startdatum: str
    einddatum: str
    aantal_deelcontacten: int = Field(ge=1)
    episodeverdeling: list[EpisodeContactVerdeling]
    kernontwikkelingen: list[str]


class DossierBlauwdruk(BaseModel):
    """Globale patiënt- en tijdlijnplanning vóór het schrijven van SOEP-tekst."""

    patient_achtergrond: PatientAchtergrond
    episodeontwerp: list[EpisodeOntwerp]
    blokken: list[BlokKader]


AanvullendOnderdeel = Literal[
    "Medicatie",
    "Laboratorium",
    "Microbiologie",
    "Verwijsbrief",
    "Specialistenbrief",
    "Ontslagbrief",
    "Allergie",
    "Geen",
]


class GeplandContact(BaseModel):
    """Compact plan van één deelcontact, nog zonder uitgewerkte SOEP-regels."""

    contact_id: str
    datum: str
    zorgverlener: Literal["Huisarts", "POH-S", "Doktersassistente"]
    contactvorm: Literal[
        "Praktijkconsult", "Telefonisch", "Huisbezoek", "Administratief"
    ]
    episode: str
    klinisch_doel: str
    kerngebeurtenis: str
    aanvullende_onderdelen: list[AanvullendOnderdeel]


class PeriodePlan(BaseModel):
    """Volledig contactplan voor één tijdsblok."""

    bloknummer: int
    startdatum: str
    einddatum: str
    aantal_deelcontacten: int
    contacten: list[GeplandContact]
    samenvatting_gepland_beloop: str


class ActueleLabwaarde(BaseModel):
    bepaling: str
    waarde: str
    eenheid: str
    datum: str


class EpisodeStatusTussenstand(BaseModel):
    """Cumulatieve status van een episode na een tijdsblok."""

    volgorde: int
    episode: str
    icpc_code: str
    startdatum: str = Field(description="Leeg als de episode nog niet is gestart")
    einddatum: str = Field(description="Leeg als de episode niet is afgesloten")
    status: Literal["Nog niet gestart", "Actief", "Afgesloten"]
    attentiewaarde: Literal["Ja", "Nee"]
    samenvatting_beloop: str
    beleid: str


class DossierStatus(BaseModel):
    """Compacte cumulatieve toestand die naar het volgende blok wordt doorgegeven."""

    peildatum: str
    geboortedatum: str = Field(
        description="Exacte fictieve geboortedatum van de patiënt in JJJJ-MM-DD"
    )
    leeftijd_op_peildatum: int = Field(
        ge=0,
        description=(
            "Deterministisch uit geboortedatum en peildatum berekende leeftijd"
        ),
    )
    episode_statussen: list[EpisodeStatusTussenstand]
    medicatiehistorie: list[MedicatieRegel]
    allergieen: list[AllergieRegel]
    laatste_relevante_labwaarden: list[ActueleLabwaarde]
    openstaande_verwijstrajecten: list[str]
    openstaande_vervolgacties: list[str]
    functionele_en_sociale_context: str
    samenvatting_tot_dusver: str


class DossierPeriode(BaseModel):
    """Uitgeschreven dossieronderdelen die in één tijdsblok zijn ontstaan."""

    bloknummer: int
    startdatum: str
    einddatum: str
    contacten: list[Deelcontact]
    medicatie: list[MedicatieRegel]
    correspondentie: list[CorrespondentieRegel]
    allergieen: list[AllergieRegel]
    microbiologie: list[MicrobiologieRegel]
    laboratorium: list[LaboratoriumRegel]
    samenvatting_periode: str
    eindstatus: DossierStatus


class PlanBeoordeling(BaseModel):
    goedgekeurd: bool
    kritieke_fouten: list[str]
    waarschuwingen: list[str]
    herstelinstructies: list[str]


class PeriodeBeoordeling(BaseModel):
    goedgekeurd: bool
    kritieke_fouten: list[str]
    waarschuwingen: list[str]
    herstelinstructies: list[str]


class Eindbeoordeling(BaseModel):
    goedgekeurd: bool
    kritieke_fouten: list[str]
    waarschuwingen: list[str]
    samenvatting: str


def maak_dynamisch_dossiermodel(
    instellingen: PatientInstellingen,
) -> type[SynthetischDossier]:
    """Beperk alle episodevelden tot de exacte namen uit het settingsbestand.

    De vaste Pydantic-modellen beschrijven de algemene uitvoerstructuur. Dit
    dynamische model voegt voor deze patiënt een enum toe met uitsluitend de
    toegestane episodenamen. Daardoor kan het taalmodel in een episodeveld geen
    samengestelde of anders geformuleerde episodenaam meer teruggeven.
    """
    episodenamen = tuple(episode.episode for episode in instellingen.episodes)
    episode_keuze = Literal[episodenamen]

    dynamisch_deelcontact = create_model(
        "DeelcontactMetVasteEpisode",
        __base__=Deelcontact,
        episode=(
            episode_keuze,
            Field(description="Exact één episodenaam uit het instellingenbestand"),
        ),
    )
    dynamische_episodeoverzichtregel = create_model(
        "EpisodeOverzichtRegelMetVasteEpisode",
        __base__=EpisodeOverzichtRegel,
        episode=(
            episode_keuze,
            Field(description="Exacte episodenaam uit het instellingenbestand"),
        ),
    )
    dynamische_correspondentieregel = create_model(
        "CorrespondentieRegelMetVasteEpisode",
        __base__=CorrespondentieRegel,
        episode=(
            episode_keuze,
            Field(description="Exacte gekoppelde episode uit het instellingenbestand"),
        ),
    )
    dynamische_microbiologieregel = create_model(
        "MicrobiologieRegelMetVasteEpisode",
        __base__=MicrobiologieRegel,
        episode=(
            episode_keuze,
            Field(description="Exacte gekoppelde episode uit het instellingenbestand"),
        ),
    )
    dynamische_laboratoriumregel = create_model(
        "LaboratoriumRegelMetVasteEpisode",
        __base__=LaboratoriumRegel,
        episode=(
            episode_keuze,
            Field(description="Exacte gekoppelde episode uit het instellingenbestand"),
        ),
    )

    return create_model(
        "SynthetischDossierMetVasteEpisodes",
        __base__=SynthetischDossier,
        episodelijst=(list[dynamische_episodeoverzichtregel], Field(...)),
        contacten=(list[dynamisch_deelcontact], Field(...)),
        correspondentie=(list[dynamische_correspondentieregel], Field(...)),
        microbiologie=(list[dynamische_microbiologieregel], Field(...)),
        laboratorium=(list[dynamische_laboratoriumregel], Field(...)),
    )


def maak_dynamische_iteratiemodellen(
    instellingen: PatientInstellingen,
) -> tuple[type[BaseModel], type[BaseModel], type[BaseModel]]:
    """Maak patiëntspecifieke schema's voor blauwdruk, plan en dossierblok.

    Alle episodevelden krijgen dezelfde strikte enum. Dit voorkomt dat een agent
    een episodenaam combineert, afkort of anders formuleert.
    """
    episodenamen = tuple(episode.episode for episode in instellingen.episodes)
    episode_keuze = Literal[episodenamen]

    dynamisch_episodeontwerp = create_model(
        "EpisodeOntwerpMetVasteEpisode",
        __base__=EpisodeOntwerp,
        episode=(episode_keuze, Field(description="Exacte episodenaam uit Excel")),
    )
    dynamische_episodeverdeling = create_model(
        "EpisodeContactVerdelingMetVasteEpisode",
        __base__=EpisodeContactVerdeling,
        episode=(episode_keuze, Field(description="Exacte episodenaam uit Excel")),
    )
    dynamisch_blokkader = create_model(
        "BlokKaderMetVasteEpisodes",
        __base__=BlokKader,
        episodeverdeling=(list[dynamische_episodeverdeling], Field(...)),
    )
    dynamische_blauwdruk = create_model(
        "DossierBlauwdrukMetVasteEpisodes",
        __base__=DossierBlauwdruk,
        episodeontwerp=(list[dynamisch_episodeontwerp], Field(...)),
        blokken=(list[dynamisch_blokkader], Field(...)),
    )

    dynamisch_gepland_contact = create_model(
        "GeplandContactMetVasteEpisode",
        __base__=GeplandContact,
        episode=(episode_keuze, Field(description="Exact één episode uit Excel")),
    )
    dynamisch_periodeplan = create_model(
        "PeriodePlanMetVasteEpisodes",
        __base__=PeriodePlan,
        contacten=(list[dynamisch_gepland_contact], Field(...)),
    )

    dynamisch_deelcontact = create_model(
        "IteratiefDeelcontactMetVasteEpisode",
        __base__=Deelcontact,
        episode=(episode_keuze, Field(description="Exact één episode uit Excel")),
    )
    dynamische_correspondentie = create_model(
        "IteratieveCorrespondentieMetVasteEpisode",
        __base__=CorrespondentieRegel,
        episode=(episode_keuze, Field(description="Exacte episode uit Excel")),
    )
    dynamische_microbiologie = create_model(
        "IteratieveMicrobiologieMetVasteEpisode",
        __base__=MicrobiologieRegel,
        episode=(episode_keuze, Field(description="Exacte episode uit Excel")),
    )
    dynamisch_laboratorium = create_model(
        "IteratiefLaboratoriumMetVasteEpisode",
        __base__=LaboratoriumRegel,
        episode=(episode_keuze, Field(description="Exacte episode uit Excel")),
    )
    dynamische_episode_status = create_model(
        "EpisodeStatusTussenstandMetVasteEpisode",
        __base__=EpisodeStatusTussenstand,
        episode=(episode_keuze, Field(description="Exacte episode uit Excel")),
    )
    dynamische_status = create_model(
        "DossierStatusMetVasteEpisodes",
        __base__=DossierStatus,
        episode_statussen=(list[dynamische_episode_status], Field(...)),
    )
    dynamische_dossierperiode = create_model(
        "DossierPeriodeMetVasteEpisodes",
        __base__=DossierPeriode,
        contacten=(list[dynamisch_deelcontact], Field(...)),
        correspondentie=(list[dynamische_correspondentie], Field(...)),
        microbiologie=(list[dynamische_microbiologie], Field(...)),
        laboratorium=(list[dynamisch_laboratorium], Field(...)),
        eindstatus=(dynamische_status, Field(...)),
    )

    return (
        dynamische_blauwdruk,
        dynamisch_periodeplan,
        dynamische_dossierperiode,
    )


def datum_naar_nederlands(datum: str) -> str:
    """Zet JJJJ-MM-DD om naar DD-MM-JJJJ voor weergave in Excel."""
    if not datum:
        return ""

    try:
        return datetime.strptime(datum, "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        # Laat de oorspronkelijke waarde staan als het model onverwacht een
        # andere notatie heeft gebruikt.
        return datum


def normaliseer_episodenaam(waarde: str) -> str:
    """Normaliseer alleen voor een robuuste vergelijking van episodenamen."""
    return re.sub(r"\s+", " ", waarde.strip()).casefold()


ACUTE_ZIEKENHUIS_SIGNALEN = (
    "ambulance",
    "seh",
    "spoedeisende hulp",
    "ziekenhuisopname",
    "ziekenhuis opname",
    "opgenomen",
    "opname",
    "112",
)


def bevat_acuut_ziekenhuissignaal(*teksten: str) -> bool:
    """Herken in plannings- of SOEP-tekst een acute ziekenhuisopname."""
    samengevoegd = " ".join(tekst for tekst in teksten if tekst).casefold()
    for ontkenning in (
        "geen ziekenhuisopname",
        "geen ziekenhuis opname",
        "geen opname",
        "niet opgenomen",
        "opname niet nodig",
        "geen indicatie voor opname",
    ):
        samengevoegd = samengevoegd.replace(ontkenning, "")
    return any(signaal in samengevoegd for signaal in ACUTE_ZIEKENHUIS_SIGNALEN)


def contact_past_bij_ontslagbrief(
    contact_episode: str,
    brief_episode: str,
    contactdatum: date,
    briefdatum: date,
    *contactteksten: str,
) -> bool:
    """Koppel een ontslagbrief aan dezelfde episode of dezelfde acute opname.

    Tijdens een opname kan een relevante nevendiagnose ontstaan. Daarom mag de
    ontslagbrief een andere episode hebben dan het eerdere acute contact, mits
    dat contact duidelijk een ziekenhuisopname beschrijft en maximaal 90 dagen
    vóór de brief ligt.
    """
    verschil = (briefdatum - contactdatum).days
    if verschil < 1 or verschil > 90:
        return False
    if normaliseer_episodenaam(contact_episode) == normaliseer_episodenaam(
        brief_episode
    ):
        return True
    return bevat_acuut_ziekenhuissignaal(*contactteksten)


def valideer_gegenereerd_dossier(
    dossier: SynthetischDossier,
    instellingen: PatientInstellingen,
) -> None:
    """Controleer essentiële aantallen, episodenamen en episodechronologie."""
    fouten = []

    if len(dossier.contacten) != instellingen.aantal_deelcontacten:
        fouten.append(
            f"verwacht {instellingen.aantal_deelcontacten} deelcontacten, "
            f"maar ontving {len(dossier.contacten)}"
        )

    verwachte_episodes = [episode.episode for episode in instellingen.episodes]
    ontvangen_episodes = [episode.episode for episode in dossier.episodelijst]
    if len(ontvangen_episodes) != len(verwachte_episodes):
        fouten.append(
            f"verwacht {len(verwachte_episodes)} regels in de episodelijst, "
            f"maar ontving {len(ontvangen_episodes)}"
        )
    else:
        for index, (verwacht, ontvangen) in enumerate(
            zip(verwachte_episodes, ontvangen_episodes),
            start=1,
        ):
            if normaliseer_episodenaam(verwacht) != normaliseer_episodenaam(ontvangen):
                fouten.append(
                    f"episode {index} heet '{ontvangen}' in plaats van '{verwacht}'"
                )

    toegestane_episodes = {
        normaliseer_episodenaam(episode) for episode in verwachte_episodes
    }
    onbekende_contactepisodes = sorted(
        {
            contact.episode
            for contact in dossier.contacten
            if normaliseer_episodenaam(contact.episode) not in toegestane_episodes
        }
    )
    if onbekende_contactepisodes:
        fouten.append(
            "onbekende episodenamen in het journaal: "
            + ", ".join(onbekende_contactepisodes)
        )

    _, dossier_einddatum = bepaal_dossierperiode(instellingen.aantal_jaren)
    uiterste_datum = date.fromisoformat(dossier_einddatum)
    for episode in dossier.episodelijst:
        try:
            start = date.fromisoformat(episode.startdatum)
        except ValueError:
            fouten.append(
                f"ongeldige startdatum bij episode '{episode.episode}': "
                f"{episode.startdatum}"
            )
            continue

        if start > uiterste_datum:
            fouten.append(
                f"startdatum van episode '{episode.episode}' ligt na de dossierperiode"
            )

        if episode.status == "Actief" and episode.einddatum:
            fouten.append(
                f"actieve episode '{episode.episode}' heeft toch een einddatum"
            )
        if episode.status == "Afgesloten" and not episode.einddatum:
            fouten.append(
                f"afgesloten episode '{episode.episode}' mist een einddatum"
            )
        if episode.einddatum:
            try:
                einde = date.fromisoformat(episode.einddatum)
            except ValueError:
                fouten.append(
                    f"ongeldige einddatum bij episode '{episode.episode}': "
                    f"{episode.einddatum}"
                )
                continue
            if einde < start:
                fouten.append(
                    f"einddatum van episode '{episode.episode}' ligt vóór de startdatum"
                )
            if einde > uiterste_datum:
                fouten.append(
                    f"einddatum van episode '{episode.episode}' ligt na de dossierperiode"
                )

    if fouten:
        raise RuntimeError(
            "Het model leverde een dossier op dat niet aan de basiscontroles voldoet:\n- "
            + "\n- ".join(fouten)
        )


def dossier_naar_excel(
    dossier: SynthetischDossier,
    instellingen: PatientInstellingen,
    uitvoerpad: Path,
) -> None:
    """Exporteer patiëntgegevens, episodelijst en dossieronderdelen naar Excel."""
    startdatum, einddatum = bepaal_dossierperiode(instellingen.aantal_jaren)
    geboortedatum = bepaal_fictieve_geboortedatum(
        instellingen.patient_id,
        instellingen.leeftijd,
        einddatum,
    )

    patient_dataframe = pd.DataFrame(
        [
            {"Onderdeel": "Patiënt-ID", "Waarde": instellingen.patient_id},
            {
                "Onderdeel": "Geboortedatum",
                "Waarde": datum_naar_nederlands(geboortedatum),
            },
            {
                "Onderdeel": "Leeftijd einde dossierperiode",
                "Waarde": f"{instellingen.leeftijd} jaar",
            },
            {"Onderdeel": "Formeel geslacht", "Waarde": instellingen.geslacht},
            {
                "Onderdeel": "Dossierperiode",
                "Waarde": (
                    f"{datum_naar_nederlands(startdatum)} t/m "
                    f"{datum_naar_nederlands(einddatum)}"
                ),
            },
            {
                "Onderdeel": "Rookstatus",
                "Waarde": dossier.patient_achtergrond.rookstatus,
            },
            {
                "Onderdeel": "Alcoholgebruik",
                "Waarde": dossier.patient_achtergrond.alcoholgebruik,
            },
            {
                "Onderdeel": "Familieanamnese",
                "Waarde": dossier.patient_achtergrond.familieanamnese,
            },
            {
                "Onderdeel": "Woonsituatie",
                "Waarde": dossier.patient_achtergrond.woonsituatie,
            },
            {
                "Onderdeel": "Beroep/dagbesteding",
                "Waarde": dossier.patient_achtergrond.beroep_dagbesteding,
            },
            {
                "Onderdeel": "Functionele context",
                "Waarde": dossier.patient_achtergrond.functionele_context,
            },
        ],
        columns=["Onderdeel", "Waarde"],
    )

    episode_rijen = []
    for episode in sorted(dossier.episodelijst, key=lambda regel: regel.volgorde):
        episode_rijen.append(
            {
                "Volgorde": episode.volgorde,
                "Episode": episode.episode,
                "ICPC-code": episode.icpc_code,
                "Startdatum": datum_naar_nederlands(episode.startdatum),
                "Einddatum": datum_naar_nederlands(episode.einddatum),
                "Status": episode.status,
                "Attentiewaarde": episode.attentiewaarde,
                "Samenvatting beloop": episode.samenvatting_beloop,
                "Beleid": episode.beleid,
            }
        )

    episodelijst_dataframe = pd.DataFrame(
        episode_rijen,
        columns=[
            "Volgorde",
            "Episode",
            "ICPC-code",
            "Startdatum",
            "Einddatum",
            "Status",
            "Attentiewaarde",
            "Samenvatting beloop",
            "Beleid",
        ],
    )

    rijen = []

    for contact in dossier.contacten:
        metadata = (
            f"[{contact.contact_id} | {datum_naar_nederlands(contact.datum)} | "
            f"{contact.zorgverlener} | "
            f"{contact.contactvorm} | Episode: {contact.episode}]"
        )
        for soep_code, tekst in (
            ("S", contact.s),
            ("O", contact.o),
            ("E", contact.e),
            ("P", contact.p),
        ):
            rijen.append({"SOEP": soep_code, "Inhoud": f"{metadata} {tekst}"})

    dataframe = pd.DataFrame(rijen, columns=["SOEP", "Inhoud"])

    medicatie_rijen = []
    for medicijn in dossier.medicatie:
        medicatie_rijen.append(
            {
                "Geneesmiddel": medicijn.geneesmiddel,
                "Sterkte": medicijn.sterkte,
                "Dosering": medicijn.dosering,
                "Indicatie": medicijn.indicatie,
                "Startdatum": datum_naar_nederlands(medicijn.startdatum),
                "Einddatum": datum_naar_nederlands(medicijn.einddatum),
                "Status": medicijn.status,
            }
        )

    medicatie_dataframe = pd.DataFrame(
        medicatie_rijen,
        columns=[
            "Geneesmiddel",
            "Sterkte",
            "Dosering",
            "Indicatie",
            "Startdatum",
            "Einddatum",
            "Status",
        ],
    )

    correspondentie_rijen = []
    # Sorteer eerst op de interne JJJJ-MM-DD-notatie. Daarna zetten we de datum
    # pas om naar DD-MM-JJJJ voor de zichtbare Excel-uitvoer.
    gesorteerde_correspondentie = sorted(
        dossier.correspondentie,
        key=lambda brief: (brief.datum, brief.traject_id),
    )
    for brief in gesorteerde_correspondentie:
        correspondentie_rijen.append(
            {
                "Traject-ID": brief.traject_id,
                "Datum": datum_naar_nederlands(brief.datum),
                "Type": brief.type_brief,
                "Specialisme": brief.specialisme,
                "Episode": brief.episode,
                "Van": brief.van,
                "Aan": brief.aan,
                "Onderwerp": brief.onderwerp,
                "Brieftekst": brief.inhoud,
            }
        )

    correspondentie_dataframe = pd.DataFrame(
        correspondentie_rijen,
        columns=[
            "Traject-ID",
            "Datum",
            "Type",
            "Specialisme",
            "Episode",
            "Van",
            "Aan",
            "Onderwerp",
            "Brieftekst",
        ],
    )

    allergie_rijen = []
    for allergie in dossier.allergieen:
        allergie_rijen.append(
            {
                "Allergeen": allergie.allergeen,
                "Type": allergie.type_allergie,
                "Reactie": allergie.reactie,
                "Ernst": allergie.ernst,
                "Registratiedatum": datum_naar_nederlands(allergie.registratiedatum),
                "Status": allergie.status,
            }
        )

    allergie_dataframe = pd.DataFrame(
        allergie_rijen,
        columns=[
            "Allergeen",
            "Type",
            "Reactie",
            "Ernst",
            "Registratiedatum",
            "Status",
        ],
    )

    microbiologie_rijen = []
    for uitslag in sorted(dossier.microbiologie, key=lambda regel: regel.datum):
        microbiologie_rijen.append(
            {
                "Datum": datum_naar_nederlands(uitslag.datum),
                "Episode": uitslag.episode,
                "Materiaal": uitslag.materiaal,
                "Onderzoek": uitslag.onderzoek,
                "Uitslag": uitslag.uitslag,
                "Verwekker": uitslag.verwekker,
                "Hoeveelheid/groei": uitslag.hoeveelheid,
                "Gevoeligheid/resistentie": uitslag.gevoeligheid_resistentie,
                "Conclusie": uitslag.conclusie,
            }
        )

    microbiologie_dataframe = pd.DataFrame(
        microbiologie_rijen,
        columns=[
            "Datum",
            "Episode",
            "Materiaal",
            "Onderzoek",
            "Uitslag",
            "Verwekker",
            "Hoeveelheid/groei",
            "Gevoeligheid/resistentie",
            "Conclusie",
        ],
    )

    laboratorium_rijen = []
    for uitslag in sorted(
        dossier.laboratorium,
        key=lambda regel: (regel.datum, regel.aanvraag_id, regel.bepaling),
    ):
        laboratorium_rijen.append(
            {
                "Aanvraag-ID": uitslag.aanvraag_id,
                "Datum": datum_naar_nederlands(uitslag.datum),
                "Aanvraagreden": uitslag.aanvraagreden,
                "Episode": uitslag.episode,
                "Bepaling": uitslag.bepaling,
                "Waarde": uitslag.waarde,
                "Eenheid": uitslag.eenheid,
                "Referentiewaarde": uitslag.referentiewaarde,
                "Afwijking": uitslag.afwijking,
                "Conclusie": uitslag.conclusie,
            }
        )

    laboratorium_dataframe = pd.DataFrame(
        laboratorium_rijen,
        columns=[
            "Aanvraag-ID",
            "Datum",
            "Aanvraagreden",
            "Episode",
            "Bepaling",
            "Waarde",
            "Eenheid",
            "Referentiewaarde",
            "Conclusie",
            "Afwijking",
        ],
    )

    # Plaats de correspondentie enkele rijen onder de medicatielijst.
    # Deze variabele is een Excel-rijnummer (Excel telt vanaf 1).
    correspondentie_titelrij = len(medicatie_dataframe) + 5
    correspondentie_eindrij = correspondentie_titelrij + 1 + len(
        correspondentie_dataframe
    )
    allergie_titelrij = correspondentie_eindrij + 3
    allergie_eindrij = allergie_titelrij + 1 + len(allergie_dataframe)
    microbiologie_titelrij = allergie_eindrij + 3
    microbiologie_eindrij = microbiologie_titelrij + 1 + len(
        microbiologie_dataframe
    )
    laboratorium_titelrij = microbiologie_eindrij + 3
    laboratorium_eindrij = laboratorium_titelrij + 1 + len(
        laboratorium_dataframe
    )

    with pd.ExcelWriter(uitvoerpad, engine="openpyxl") as writer:
        patient_dataframe.to_excel(
            writer,
            sheet_name="Patiëntgegevens",
            index=False,
            startrow=2,
        )
        episodelijst_dataframe.to_excel(
            writer,
            sheet_name="Episodelijst",
            index=False,
            startrow=2,
        )
        dataframe.to_excel(writer, sheet_name="Journaal", index=False)
        medicatie_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=1,
            startcol=3,
        )
        correspondentie_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=correspondentie_titelrij,
            startcol=3,
        )
        allergie_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=allergie_titelrij,
            startcol=3,
        )
        microbiologie_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=microbiologie_titelrij,
            startcol=3,
        )
        laboratorium_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=laboratorium_titelrij,
            startcol=3,
        )

        patient_werkblad = writer.book["Patiëntgegevens"]
        patient_werkblad.merge_cells("A1:B1")
        patient_werkblad["A1"] = "PATIËNT- EN ACHTERGRONDGEGEVENS"
        patient_werkblad.freeze_panes = "A4"
        patient_werkblad.auto_filter.ref = f"A3:B{len(patient_dataframe) + 3}"
        patient_werkblad.column_dimensions["A"].width = 32
        patient_werkblad.column_dimensions["B"].width = 85

        episode_werkblad = writer.book["Episodelijst"]
        episode_werkblad.merge_cells("A1:I1")
        episode_werkblad["A1"] = "EPISODE-/PROBLEEMLIJST"
        episode_werkblad.freeze_panes = "A4"
        episode_werkblad.auto_filter.ref = (
            f"A3:I{len(episodelijst_dataframe) + 3}"
        )
        episode_breedtes = {
            "A": 10,
            "B": 32,
            "C": 13,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 16,
            "H": 50,
            "I": 55,
        }
        for kolom, breedte in episode_breedtes.items():
            episode_werkblad.column_dimensions[kolom].width = breedte

        werkblad = writer.book["Journaal"]
        werkblad.freeze_panes = "A2"
        werkblad.auto_filter.ref = f"A1:B{len(dataframe) + 1}"
        werkblad.column_dimensions["A"].width = 10
        werkblad.column_dimensions["B"].width = 115

        werkblad.merge_cells("D1:J1")
        werkblad["D1"] = "MEDICATIELIJST"
        werkblad.column_dimensions["D"].width = 24
        werkblad.column_dimensions["E"].width = 14
        werkblad.column_dimensions["F"].width = 22
        werkblad.column_dimensions["G"].width = 32
        werkblad.column_dimensions["H"].width = 14
        werkblad.column_dimensions["I"].width = 14
        werkblad.column_dimensions["J"].width = 12
        werkblad.column_dimensions["K"].width = 30
        werkblad.column_dimensions["L"].width = 100
        werkblad.column_dimensions["M"].width = 14

        werkblad.merge_cells(
            start_row=correspondentie_titelrij,
            start_column=4,
            end_row=correspondentie_titelrij,
            end_column=12,
        )
        correspondentie_titelcel = werkblad.cell(
            row=correspondentie_titelrij,
            column=4,
        )
        correspondentie_titelcel.value = (
            "CORRESPONDENTIE – VERWIJS- EN SPECIALISTENBRIEVEN"
        )

        werkblad.merge_cells(
            start_row=allergie_titelrij,
            start_column=4,
            end_row=allergie_titelrij,
            end_column=9,
        )
        allergie_titelcel = werkblad.cell(row=allergie_titelrij, column=4)
        allergie_titelcel.value = "ALLERGIEËN"

        werkblad.merge_cells(
            start_row=microbiologie_titelrij,
            start_column=4,
            end_row=microbiologie_titelrij,
            end_column=12,
        )
        microbiologie_titelcel = werkblad.cell(
            row=microbiologie_titelrij,
            column=4,
        )
        microbiologie_titelcel.value = "MICROBIOLOGIE"

        werkblad.merge_cells(
            start_row=laboratorium_titelrij,
            start_column=4,
            end_row=laboratorium_titelrij,
            end_column=13,
        )
        laboratorium_titelcel = werkblad.cell(
            row=laboratorium_titelrij,
            column=4,
        )
        laboratorium_titelcel.value = "LABORATORIUMAANVRAGEN EN -UITSLAGEN"

        from openpyxl.styles import Alignment, Font, PatternFill

        patient_werkblad["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        patient_werkblad["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        patient_werkblad["A1"].alignment = Alignment(horizontal="center")
        for cel in patient_werkblad[3][0:2]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="5B9BD5")
            cel.alignment = Alignment(horizontal="center", vertical="center")
        for rij in patient_werkblad.iter_rows(
            min_row=4,
            max_row=len(patient_dataframe) + 3,
            min_col=1,
            max_col=2,
        ):
            rij[0].font = Font(bold=True, color="1F1F1F")
            rij[0].fill = PatternFill("solid", fgColor="DDEBF7")
            rij[1].fill = PatternFill("solid", fgColor="FFF2CC")
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        episode_werkblad["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        episode_werkblad["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        episode_werkblad["A1"].alignment = Alignment(horizontal="center")
        for cel in episode_werkblad[3][0:9]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="5B9BD5")
            cel.alignment = Alignment(horizontal="center", vertical="center")
        for rij in episode_werkblad.iter_rows(
            min_row=4,
            max_row=len(episodelijst_dataframe) + 3,
            min_col=1,
            max_col=9,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")
            rij[0].alignment = Alignment(horizontal="center", vertical="top")
            rij[2].alignment = Alignment(horizontal="center", vertical="top")
            rij[5].alignment = Alignment(horizontal="center", vertical="top")
            rij[6].alignment = Alignment(horizontal="center", vertical="top")

        for cel in werkblad[1][0:2]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="1F4E78")
            cel.alignment = Alignment(vertical="center")

        werkblad["D1"].font = Font(bold=True, color="FFFFFF")
        werkblad["D1"].fill = PatternFill("solid", fgColor="548235")
        werkblad["D1"].alignment = Alignment(horizontal="center")

        for cel in werkblad[2][3:10]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="70AD47")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        correspondentie_titelcel.font = Font(bold=True, color="FFFFFF")
        correspondentie_titelcel.fill = PatternFill("solid", fgColor="7030A0")
        correspondentie_titelcel.alignment = Alignment(horizontal="center")

        for cel in werkblad[correspondentie_titelrij + 1][3:12]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="9E66B4")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        allergie_titelcel.font = Font(bold=True, color="FFFFFF")
        allergie_titelcel.fill = PatternFill("solid", fgColor="C65911")
        allergie_titelcel.alignment = Alignment(horizontal="center")

        for cel in werkblad[allergie_titelrij + 1][3:9]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="ED7D31")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        microbiologie_titelcel.font = Font(bold=True, color="FFFFFF")
        microbiologie_titelcel.fill = PatternFill("solid", fgColor="008C95")
        microbiologie_titelcel.alignment = Alignment(horizontal="center")

        for cel in werkblad[microbiologie_titelrij + 1][3:12]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="00A6B2")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        laboratorium_titelcel.font = Font(bold=True, color="FFFFFF")
        laboratorium_titelcel.fill = PatternFill("solid", fgColor="2F5597")
        laboratorium_titelcel.alignment = Alignment(horizontal="center")

        for cel in werkblad[laboratorium_titelrij + 1][3:13]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="4472C4")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        for rij in werkblad.iter_rows(min_row=2, max_col=2):
            rij[0].alignment = Alignment(horizontal="center", vertical="top")
            rij[1].alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=3,
            max_row=len(medicatie_dataframe) + 2,
            min_col=4,
            max_col=10,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=correspondentie_titelrij + 2,
            max_row=correspondentie_eindrij,
            min_col=4,
            max_col=12,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=allergie_titelrij + 2,
            max_row=allergie_eindrij,
            min_col=4,
            max_col=9,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=microbiologie_titelrij + 2,
            max_row=microbiologie_eindrij,
            min_col=4,
            max_col=12,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=laboratorium_titelrij + 2,
            max_row=laboratorium_eindrij,
            min_col=4,
            max_col=13,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")


def model_naar_json(waarde: BaseModel | dict | list) -> str:
    """Serialiseer ook lijsten en dicts met geneste Pydantic-modellen."""

    def json_standaard(item: object) -> object:
        if isinstance(item, BaseModel):
            return item.model_dump(mode="json")
        if isinstance(item, (date, datetime)):
            return item.isoformat()
        if isinstance(item, Path):
            return str(item)
        raise TypeError(
            f"Object van type {item.__class__.__name__} is niet JSON-serialiseerbaar"
        )

    return json.dumps(
        waarde,
        ensure_ascii=False,
        indent=2,
        default=json_standaard,
    )


def schrijf_checkpoint(pad: Path, waarde: BaseModel | dict | list) -> None:
    """Bewaar een reproduceerbaar tussenproduct van de iteratieve run."""
    pad.parent.mkdir(parents=True, exist_ok=True)
    pad.write_text(model_naar_json(waarde), encoding="utf-8")


def verschuif_datum_met_jaren(bron: date, aantal_jaren: int) -> date:
    """Verschuif een datum, met een veilige afhandeling van 29 februari."""
    try:
        return bron.replace(year=bron.year + aantal_jaren)
    except ValueError:
        return bron.replace(year=bron.year + aantal_jaren, day=28)


def maak_tijdsblokken(
    instellingen: PatientInstellingen,
) -> list[tuple[int, str, str]]:
    """Verdeel de dossierperiode in aansluitende blokken van maximaal vijf jaar."""
    starttekst, eindtekst = bepaal_dossierperiode(instellingen.aantal_jaren)
    start = date.fromisoformat(starttekst)
    uiterste_einde = date.fromisoformat(eindtekst)
    blokken = []
    cursor = start
    aantal_blokken = (
        instellingen.aantal_jaren + BLOKGROOTTE_JAREN - 1
    ) // BLOKGROOTTE_JAREN

    for bloknummer in range(1, aantal_blokken + 1):
        if bloknummer == aantal_blokken:
            # De bestaande dossierperiode gebruikt dezelfde kalenderdatum aan
            # begin en einde. Daardoor is de inclusieve periode één dag langer
            # dan precies N x 365 dagen. Laat die slotdag in het laatste blok
            # vallen in plaats van een afzonderlijk één-dagsblok te maken.
            blokeinde = uiterste_einde
        else:
            kandidaat_einde = (
                verschuif_datum_met_jaren(cursor, BLOKGROOTTE_JAREN)
                - timedelta(days=1)
            )
            blokeinde = min(kandidaat_einde, uiterste_einde)
        blokken.append((bloknummer, cursor.isoformat(), blokeinde.isoformat()))
        cursor = blokeinde + timedelta(days=1)

    return blokken


def veilige_iso_datum(waarde: str, label: str, fouten: list[str]) -> date | None:
    """Lees een ISO-datum en voeg bij fouten een begrijpelijke melding toe."""
    try:
        return date.fromisoformat(waarde)
    except (TypeError, ValueError):
        fouten.append(f"{label} heeft geen geldige datum in JJJJ-MM-DD: '{waarde}'")
        return None


def verdeel_geheel_getal_proportioneel(
    totaal: int,
    gewichten: list[int],
    minimum_per_item: int,
) -> list[int]:
    """Verdeel een geheel totaal reproduceerbaar over meerdere items."""
    aantal_items = len(gewichten)
    minimumtotaal = aantal_items * minimum_per_item
    if totaal < minimumtotaal:
        raise ValueError(
            f"Totaal {totaal} is te klein voor {aantal_items} items met "
            f"minimum {minimum_per_item}."
        )
    if not gewichten:
        return []

    restant = totaal - minimumtotaal
    positieve_gewichten = [max(0, int(gewicht)) for gewicht in gewichten]
    if sum(positieve_gewichten) == 0:
        positieve_gewichten = [1] * aantal_items

    gewichttotaal = sum(positieve_gewichten)
    exacte_aandelen = [
        restant * gewicht / gewichttotaal for gewicht in positieve_gewichten
    ]
    gehele_aandelen = [int(aandeel) for aandeel in exacte_aandelen]
    uitkomst = [minimum_per_item + aandeel for aandeel in gehele_aandelen]

    nog_te_verdelen = totaal - sum(uitkomst)
    restvolgorde = sorted(
        range(aantal_items),
        key=lambda index: (
            exacte_aandelen[index] - gehele_aandelen[index],
            positieve_gewichten[index],
            -index,
        ),
        reverse=True,
    )
    for index in restvolgorde[:nog_te_verdelen]:
        uitkomst[index] += 1
    return uitkomst


def episode_overlapt_blok(
    ontwerp: EpisodeOntwerp,
    blok_start: str,
    blok_einde: str,
) -> bool:
    """Geef aan of een episode volgens de blauwdruk in een blok aanwezig is."""
    try:
        episode_start = date.fromisoformat(ontwerp.geplande_startdatum)
        start = date.fromisoformat(blok_start)
        einde = date.fromisoformat(blok_einde)
        episode_einde = (
            date.fromisoformat(ontwerp.geplande_einddatum)
            if ontwerp.geplande_einddatum
            else einde
        )
    except (TypeError, ValueError):
        # Ongeldige datums worden later door valideer_blauwdruk gemeld.
        return True
    return episode_start <= einde and episode_einde >= start


def maak_sluitende_contactmatrix(
    ruwe_matrix: list[list[int]],
    bloktotalen: list[int],
    episodetotalen: list[int],
    overlapmatrix: list[list[bool]],
) -> list[list[int]]:
    """Maak blok- en episodetotalen exact zonder het model te laten rekenen."""
    matrix = [
        [max(0, int(waarde)) for waarde in rij]
        for rij in ruwe_matrix
    ]
    aantal_blokken = len(bloktotalen)
    aantal_episodes = len(episodetotalen)

    def kolomsommen() -> list[int]:
        return [
            sum(matrix[blok][episode] for blok in range(aantal_blokken))
            for episode in range(aantal_episodes)
        ]

    # Verschuif eerst contacten tussen episodes binnen hetzelfde tijdsblok.
    # Hierdoor blijven de bloktotalen voorlopig gelijk.
    while True:
        huidig = kolomsommen()
        surplus = [
            huidig[index] - episodetotalen[index]
            for index in range(aantal_episodes)
        ]
        bronnen = [index for index, waarde in enumerate(surplus) if waarde > 0]
        doelen = [index for index, waarde in enumerate(surplus) if waarde < 0]
        if not bronnen or not doelen:
            break
        bron = max(bronnen, key=lambda index: surplus[index])
        doel = max(doelen, key=lambda index: -surplus[index])
        mogelijke_blokken = [
            blok for blok in range(aantal_blokken) if matrix[blok][bron] > 0
        ]
        blok = max(
            mogelijke_blokken,
            key=lambda index: (
                overlapmatrix[index][doel],
                ruwe_matrix[index][doel] > 0,
                ruwe_matrix[index][doel] - matrix[index][doel],
                matrix[index][bron],
                index,
            ),
        )
        matrix[blok][bron] -= 1
        matrix[blok][doel] += 1

    # Verwijder of voeg het eventuele verschil in het totale aantal toe.
    while True:
        huidig = kolomsommen()
        bronnen = [
            index
            for index in range(aantal_episodes)
            if huidig[index] > episodetotalen[index]
        ]
        if not bronnen:
            break
        bron = max(
            bronnen,
            key=lambda index: huidig[index] - episodetotalen[index],
        )
        mogelijke_blokken = [
            blok for blok in range(aantal_blokken) if matrix[blok][bron] > 0
        ]
        blok = max(
            mogelijke_blokken,
            key=lambda index: (
                sum(matrix[index]) - bloktotalen[index],
                matrix[index][bron],
                -ruwe_matrix[index][bron],
                index,
            ),
        )
        matrix[blok][bron] -= 1

    while True:
        huidig = kolomsommen()
        doelen = [
            index
            for index in range(aantal_episodes)
            if huidig[index] < episodetotalen[index]
        ]
        if not doelen:
            break
        doel = max(
            doelen,
            key=lambda index: episodetotalen[index] - huidig[index],
        )
        blok = max(
            range(aantal_blokken),
            key=lambda index: (
                bloktotalen[index] - sum(matrix[index]),
                overlapmatrix[index][doel],
                ruwe_matrix[index][doel] > 0,
                ruwe_matrix[index][doel] - matrix[index][doel],
                index,
            ),
        )
        matrix[blok][doel] += 1

    # Verschuif ten slotte contacten binnen dezelfde episode tussen blokken.
    # De episodetotalen blijven hierdoor exact behouden.
    while True:
        rijsommen = [sum(rij) for rij in matrix]
        bronblokken = [
            index
            for index in range(aantal_blokken)
            if rijsommen[index] > bloktotalen[index]
        ]
        doelblokken = [
            index
            for index in range(aantal_blokken)
            if rijsommen[index] < bloktotalen[index]
        ]
        if not bronblokken and not doelblokken:
            break
        if not bronblokken or not doelblokken:
            raise RuntimeError("Contactmatrix kon niet sluitend worden herverdeeld.")

        bronblok = max(
            bronblokken,
            key=lambda index: rijsommen[index] - bloktotalen[index],
        )
        doelblok = max(
            doelblokken,
            key=lambda index: bloktotalen[index] - rijsommen[index],
        )
        mogelijke_episodes = [
            episode
            for episode in range(aantal_episodes)
            if matrix[bronblok][episode] > 0
        ]
        episode = max(
            mogelijke_episodes,
            key=lambda index: (
                overlapmatrix[doelblok][index],
                ruwe_matrix[doelblok][index] > 0,
                ruwe_matrix[doelblok][index] - matrix[doelblok][index],
                matrix[bronblok][index],
                -index,
            ),
        )
        matrix[bronblok][episode] -= 1
        matrix[doelblok][episode] += 1

    return matrix


def herstel_mechanische_blauwdrukwaarden(
    blauwdruk: DossierBlauwdruk,
    instellingen: PatientInstellingen,
    tijdsblokken: list[tuple[int, str, str]],
) -> DossierBlauwdruk:
    """Herstel Excel-volgorde, blokgrenzen en contacttotalen met Python."""
    verwachte_namen = [episode.episode for episode in instellingen.episodes]
    verwacht_op_normnaam = {
        normaliseer_episodenaam(episode.episode): episode
        for episode in instellingen.episodes
    }
    ontwerpen_op_normnaam = {}
    for ontwerp in blauwdruk.episodeontwerp:
        normnaam = normaliseer_episodenaam(ontwerp.episode)
        ontwerpen_op_normnaam.setdefault(normnaam, ontwerp)

    # Gaten in de Excel-volgorde zijn geldig. Kopieer daarom de nummers uit
    # Excel in plaats van de door het model gemaakte doorlopende nummering.
    if all(normnaam in ontwerpen_op_normnaam for normnaam in verwacht_op_normnaam):
        nieuwe_ontwerpen = []
        for instelling in instellingen.episodes:
            ontwerp = ontwerpen_op_normnaam[
                normaliseer_episodenaam(instelling.episode)
            ]
            nieuwe_ontwerpen.append(
                ontwerp.model_copy(
                    update={
                        "volgorde": instelling.volgorde,
                        "episode": instelling.episode,
                    }
                )
            )
        blauwdruk = blauwdruk.model_copy(
            update={"episodeontwerp": nieuwe_ontwerpen}
        )

    if len(blauwdruk.blokken) != len(tijdsblokken):
        return blauwdruk

    ontwerpen = {
        normaliseer_episodenaam(ontwerp.episode): ontwerp
        for ontwerp in blauwdruk.episodeontwerp
    }
    ruwe_matrix = []
    for blok in blauwdruk.blokken:
        telling = Counter()
        for regel in blok.episodeverdeling:
            telling[normaliseer_episodenaam(regel.episode)] += max(
                0,
                regel.aantal_deelcontacten,
            )
        ruwe_matrix.append(
            [
                telling[normaliseer_episodenaam(episodenaam)]
                for episodenaam in verwachte_namen
            ]
        )

    ruwe_bloktotalen = [blok.aantal_deelcontacten for blok in blauwdruk.blokken]
    if sum(ruwe_bloktotalen) == instellingen.aantal_deelcontacten:
        bloktotalen = ruwe_bloktotalen
    else:
        bloktotalen = verdeel_geheel_getal_proportioneel(
            instellingen.aantal_deelcontacten,
            ruwe_bloktotalen,
            minimum_per_item=1,
        )

    if all(
        episode.aantal_deelcontacten is not None
        for episode in instellingen.episodes
    ):
        episodetotalen = [
            int(episode.aantal_deelcontacten)
            for episode in instellingen.episodes
        ]
    else:
        ruwe_episodetotalen = [
            sum(ruwe_matrix[blok][episode] for blok in range(len(ruwe_matrix)))
            for episode in range(len(verwachte_namen))
        ]
        episodetotalen = verdeel_geheel_getal_proportioneel(
            instellingen.aantal_deelcontacten,
            ruwe_episodetotalen,
            minimum_per_item=1,
        )

    overlapmatrix = []
    for _, startdatum, einddatum in tijdsblokken:
        overlapmatrix.append(
            [
                episode_overlapt_blok(
                    ontwerpen[normaliseer_episodenaam(episodenaam)],
                    startdatum,
                    einddatum,
                )
                if normaliseer_episodenaam(episodenaam) in ontwerpen
                else True
                for episodenaam in verwachte_namen
            ]
        )

    contactmatrix = maak_sluitende_contactmatrix(
        ruwe_matrix,
        bloktotalen,
        episodetotalen,
        overlapmatrix,
    )

    # Een episode die in een bepaald blok wordt afgesloten, moet in dat blok
    # minstens één contact hebben. Corrigeer dit met een 2x2-ruil zodat zowel
    # de bloktotalen als de episode-totalen exact gelijk blijven.
    for episode_index, episodenaam in enumerate(verwachte_namen):
        ontwerp = ontwerpen.get(normaliseer_episodenaam(episodenaam))
        if ontwerp is None or not ontwerp.geplande_einddatum:
            continue
        try:
            einddatum = date.fromisoformat(ontwerp.geplande_einddatum)
        except ValueError:
            continue
        doelblok = next(
            (
                blok_index
                for blok_index, (_, starttekst, eindtekst) in enumerate(tijdsblokken)
                if date.fromisoformat(starttekst)
                <= einddatum
                <= date.fromisoformat(eindtekst)
            ),
            None,
        )
        if doelblok is None or contactmatrix[doelblok][episode_index] > 0:
            continue

        bronblokken = [
            blok_index
            for blok_index in range(len(contactmatrix))
            if contactmatrix[blok_index][episode_index] > 0
            and blok_index != doelblok
        ]
        for bronblok in bronblokken:
            donoropties = [
                donor_index
                for donor_index in range(len(verwachte_namen))
                if donor_index != episode_index
                and contactmatrix[doelblok][donor_index] > 0
            ]
            donoropties.sort(
                key=lambda donor_index: (
                    overlapmatrix[bronblok][donor_index],
                    contactmatrix[doelblok][donor_index],
                ),
                reverse=True,
            )
            if not donoropties:
                continue
            donor_index = donoropties[0]
            contactmatrix[doelblok][donor_index] -= 1
            contactmatrix[doelblok][episode_index] += 1
            contactmatrix[bronblok][episode_index] -= 1
            contactmatrix[bronblok][donor_index] += 1
            break

    regelklasse = next(
        (
            type(regel)
            for blok in blauwdruk.blokken
            for regel in blok.episodeverdeling
        ),
        EpisodeContactVerdeling,
    )
    nieuwe_blokken = []
    for index, (blok, verwacht_blok) in enumerate(
        zip(blauwdruk.blokken, tijdsblokken)
    ):
        verwacht_nummer, verwacht_start, verwacht_einde = verwacht_blok
        bestaande_regels = {
            normaliseer_episodenaam(regel.episode): regel
            for regel in blok.episodeverdeling
        }
        nieuwe_verdeling = []
        for episode_index, episodenaam in enumerate(verwachte_namen):
            normnaam = normaliseer_episodenaam(episodenaam)
            bestaande_regel = bestaande_regels.get(normnaam)
            waarden = {
                "episode": episodenaam,
                "aantal_deelcontacten": contactmatrix[index][episode_index],
            }
            if bestaande_regel is not None:
                nieuwe_verdeling.append(
                    bestaande_regel.model_copy(update=waarden)
                )
            else:
                nieuwe_verdeling.append(regelklasse(**waarden))

        nieuwe_blokken.append(
            blok.model_copy(
                update={
                    "bloknummer": verwacht_nummer,
                    "startdatum": verwacht_start,
                    "einddatum": verwacht_einde,
                    "aantal_deelcontacten": bloktotalen[index],
                    "episodeverdeling": nieuwe_verdeling,
                }
            )
        )

    return blauwdruk.model_copy(update={"blokken": nieuwe_blokken})


def valideer_blauwdruk(
    blauwdruk: DossierBlauwdruk,
    instellingen: PatientInstellingen,
    tijdsblokken: list[tuple[int, str, str]],
) -> list[str]:
    """Controleer de globale planning voordat een SOEP-regel wordt geschreven."""
    fouten = []
    verwachte_episodes = [episode.episode for episode in instellingen.episodes]

    if len(blauwdruk.episodeontwerp) != len(verwachte_episodes):
        fouten.append(
            f"verwacht {len(verwachte_episodes)} episodeontwerpen, maar ontving "
            f"{len(blauwdruk.episodeontwerp)}"
        )
    else:
        for index, (verwacht, ontwerp) in enumerate(
            zip(instellingen.episodes, blauwdruk.episodeontwerp),
            start=1,
        ):
            if ontwerp.volgorde != verwacht.volgorde:
                fouten.append(
                    f"episodeontwerp {index} heeft volgorde {ontwerp.volgorde} in "
                    f"plaats van {verwacht.volgorde}"
                )
            if normaliseer_episodenaam(ontwerp.episode) != normaliseer_episodenaam(
                verwacht.episode
            ):
                fouten.append(
                    f"episodeontwerp {index} heet '{ontwerp.episode}' in plaats van "
                    f"'{verwacht.episode}'"
                )

    vorige_start = None
    _, dossier_eindtekst = bepaal_dossierperiode(instellingen.aantal_jaren)
    dossier_einde = date.fromisoformat(dossier_eindtekst)
    for ontwerp in blauwdruk.episodeontwerp:
        start = veilige_iso_datum(
            ontwerp.geplande_startdatum,
            f"startdatum van '{ontwerp.episode}'",
            fouten,
        )
        if start is None:
            continue
        if start > dossier_einde:
            fouten.append(f"'{ontwerp.episode}' start na de dossierperiode")
        if vorige_start is not None and start < vorige_start:
            fouten.append(
                "de geplande startdatums respecteren de chronologische "
                f"episodevolgorde niet bij '{ontwerp.episode}'"
            )
        vorige_start = start

        if ontwerp.status_einde_dossier == "Actief" and ontwerp.geplande_einddatum:
            fouten.append(
                f"actief eindigende episode '{ontwerp.episode}' heeft een einddatum"
            )
        if (
            ontwerp.status_einde_dossier == "Afgesloten"
            and not ontwerp.geplande_einddatum
        ):
            fouten.append(
                f"afgesloten episode '{ontwerp.episode}' mist een einddatum"
            )
        if ontwerp.geplande_einddatum:
            einde = veilige_iso_datum(
                ontwerp.geplande_einddatum,
                f"einddatum van '{ontwerp.episode}'",
                fouten,
            )
            if einde is not None and (einde < start or einde > dossier_einde):
                fouten.append(
                    f"einddatum van '{ontwerp.episode}' ligt buiten het geldige beloop"
                )

    if len(blauwdruk.blokken) != len(tijdsblokken):
        fouten.append(
            f"verwacht {len(tijdsblokken)} tijdsblokken, maar ontving "
            f"{len(blauwdruk.blokken)}"
        )

    totaaltelling = Counter()
    for index, verwacht_blok in enumerate(tijdsblokken):
        if index >= len(blauwdruk.blokken):
            break
        blok = blauwdruk.blokken[index]
        verwacht_nummer, verwacht_start, verwacht_einde = verwacht_blok
        if (
            blok.bloknummer != verwacht_nummer
            or blok.startdatum != verwacht_start
            or blok.einddatum != verwacht_einde
        ):
            fouten.append(
                f"blok {index + 1} moet exact {verwacht_start} t/m {verwacht_einde} "
                f"zijn met bloknummer {verwacht_nummer}"
            )

        ontvangen_namen = [regel.episode for regel in blok.episodeverdeling]
        if ontvangen_namen != verwachte_episodes:
            fouten.append(
                f"episodeverdeling van blok {blok.bloknummer} moet iedere episode "
                "precies één keer en in Excel-volgorde bevatten"
            )
        verdelingstotaal = sum(
            regel.aantal_deelcontacten for regel in blok.episodeverdeling
        )
        if verdelingstotaal != blok.aantal_deelcontacten:
            fouten.append(
                f"episodeverdeling van blok {blok.bloknummer} telt op tot "
                f"{verdelingstotaal}, niet tot {blok.aantal_deelcontacten}"
            )
        for regel in blok.episodeverdeling:
            totaaltelling[regel.episode] += regel.aantal_deelcontacten

    gepland_totaal = sum(blok.aantal_deelcontacten for blok in blauwdruk.blokken)
    if gepland_totaal != instellingen.aantal_deelcontacten:
        fouten.append(
            f"de blauwdruk plant {gepland_totaal} contacten in plaats van "
            f"{instellingen.aantal_deelcontacten}"
        )

    verdeling_per_blok = {
        blok.bloknummer: {
            regel.episode: regel.aantal_deelcontacten
            for regel in blok.episodeverdeling
        }
        for blok in blauwdruk.blokken
    }
    for ontwerp in blauwdruk.episodeontwerp:
        if not ontwerp.geplande_einddatum:
            continue
        einddatum = veilige_iso_datum(
            ontwerp.geplande_einddatum,
            f"einddatum van '{ontwerp.episode}'",
            fouten,
        )
        if einddatum is None:
            continue
        eindblok = next(
            (
                bloknummer
                for bloknummer, starttekst, eindtekst in tijdsblokken
                if date.fromisoformat(starttekst)
                <= einddatum
                <= date.fromisoformat(eindtekst)
            ),
            None,
        )
        if eindblok is not None and (
            verdeling_per_blok.get(eindblok, {}).get(ontwerp.episode, 0) < 1
        ):
            fouten.append(
                f"afgesloten episode '{ontwerp.episode}' heeft geen contact in "
                f"het blok van de geplande einddatum {ontwerp.geplande_einddatum}"
            )

    if all(
        episode.aantal_deelcontacten is not None
        for episode in instellingen.episodes
    ):
        for episode in instellingen.episodes:
            ontvangen = totaaltelling[episode.episode]
            if ontvangen != episode.aantal_deelcontacten:
                fouten.append(
                    f"'{episode.episode}' krijgt {ontvangen} contacten in plaats van "
                    f"het ingestelde aantal {episode.aantal_deelcontacten}"
                )
    else:
        zonder_contact = [
            episode for episode in verwachte_episodes if totaaltelling[episode] == 0
        ]
        if zonder_contact:
            fouten.append(
                "iedere ingestelde episode moet minimaal één contact krijgen; nul bij: "
                + ", ".join(zonder_contact)
            )

    return fouten


def herstel_mechanische_periodeplanwaarden(
    plan: PeriodePlan,
    blokkader: BlokKader,
    blauwdruk: DossierBlauwdruk,
) -> PeriodePlan:
    """Herstel uitsluitend de episodeaantallen van een verder compleet plan.

    Het model bepaalt de klinische planning, maar exacte totalen zijn een
    boekhoudkundige randvoorwaarde. Als het totaal aantal contacten al klopt,
    wordt ieder surpluscontact daarom deterministisch toegewezen aan een
    episode met een tekort. Brief- en afsluitcontacten worden zo veel mogelijk
    behouden.
    """
    if len(plan.contacten) != blokkader.aantal_deelcontacten:
        return plan

    verwacht = Counter(
        {
            regel.episode: regel.aantal_deelcontacten
            for regel in blokkader.episodeverdeling
        }
    )
    ontvangen = Counter(contact.episode for contact in plan.contacten)
    if ontvangen == verwacht:
        return plan
    if sum(ontvangen.values()) != sum(verwacht.values()):
        return plan

    tekorten = []
    for episode, verwacht_aantal in verwacht.items():
        tekorten.extend(
            [episode] * max(0, verwacht_aantal - ontvangen.get(episode, 0))
        )
    overschotten = Counter(
        {
            episode: max(0, ontvangen.get(episode, 0) - verwacht_aantal)
            for episode, verwacht_aantal in verwacht.items()
        }
    )
    if len(tekorten) != sum(overschotten.values()):
        return plan

    ontwerpen = {
        ontwerp.episode: ontwerp for ontwerp in blauwdruk.episodeontwerp
    }
    contacten = list(plan.contacten)
    briefonderdelen = {"Verwijsbrief", "Specialistenbrief", "Ontslagbrief"}

    def contactdatum(contact: GeplandContact) -> date | None:
        try:
            return date.fromisoformat(contact.datum)
        except ValueError:
            return None

    def is_afsluitanker(contact: GeplandContact) -> bool:
        ontwerp = ontwerpen.get(contact.episode)
        if ontwerp is None or not ontwerp.geplande_einddatum:
            return False
        try:
            einddatum = date.fromisoformat(ontwerp.geplande_einddatum)
        except ValueError:
            return False
        datum = contactdatum(contact)
        return (
            datum is not None
            and einddatum - timedelta(days=120) <= datum <= einddatum
        )

    def past_binnen_episode(
        contact: GeplandContact,
        episodenaam: str,
    ) -> bool:
        ontwerp = ontwerpen.get(episodenaam)
        datum = contactdatum(contact)
        if ontwerp is None or datum is None:
            return True
        try:
            startdatum = date.fromisoformat(ontwerp.geplande_startdatum)
        except ValueError:
            return True
        if datum < startdatum:
            return False
        if not ontwerp.geplande_einddatum:
            return True
        try:
            einddatum = date.fromisoformat(ontwerp.geplande_einddatum)
        except ValueError:
            return True
        return datum <= einddatum

    for ontbrekende_episode in tekorten:
        kandidaten = [
            index
            for index, contact in enumerate(contacten)
            if overschotten.get(contact.episode, 0) > 0
            and past_binnen_episode(contact, ontbrekende_episode)
        ]
        if not kandidaten:
            return plan

        ontbrekend_ontwerp = ontwerpen.get(ontbrekende_episode)
        if (
            ontbrekend_ontwerp is not None
            and ontbrekend_ontwerp.geplande_einddatum
            and not any(
                contact.episode == ontbrekende_episode
                and is_afsluitanker(contact)
                for contact in contacten
            )
        ):
            try:
                ontbrekende_einddatum = date.fromisoformat(
                    ontbrekend_ontwerp.geplande_einddatum
                )
            except ValueError:
                ontbrekende_einddatum = None
            ankerkandidaten = [
                index
                for index in kandidaten
                if contactdatum(contacten[index]) is not None
                and ontbrekende_einddatum is not None
                and ontbrekende_einddatum - timedelta(days=120)
                <= contactdatum(contacten[index])
                <= ontbrekende_einddatum
            ]
            if not ankerkandidaten:
                return plan
            kandidaten = ankerkandidaten
        kandidaten.sort(
            key=lambda index: (
                bool(
                    briefonderdelen
                    & set(contacten[index].aanvullende_onderdelen)
                ),
                is_afsluitanker(contacten[index]),
                contacten[index].contactvorm == "Administratief",
                len(contacten[index].aanvullende_onderdelen),
                index,
            )
        )
        gekozen_index = kandidaten[0]
        oud_contact = contacten[gekozen_index]
        ontwerp = ontwerpen.get(ontbrekende_episode)
        verwacht_beleid = (
            ontwerp.verwacht_beleid
            if ontwerp is not None
            else "passende episodegerichte controle en vervolg"
        )
        contacten[gekozen_index] = oud_contact.model_copy(
            update={
                "episode": ontbrekende_episode,
                "klinisch_doel": (
                    f"Passende controle of vervolgstap voor {ontbrekende_episode}"
                ),
                "kerngebeurtenis": (
                    f"Episodegericht vervolg conform de globale blauwdruk: "
                    f"{verwacht_beleid}"
                ),
                "aanvullende_onderdelen": ["Geen"],
            }
        )
        overschotten[oud_contact.episode] -= 1

    return plan.model_copy(
        update={
            "bloknummer": blokkader.bloknummer,
            "startdatum": blokkader.startdatum,
            "einddatum": blokkader.einddatum,
            "aantal_deelcontacten": blokkader.aantal_deelcontacten,
            "contacten": contacten,
        }
    )


def herstel_correspondentie_in_periodeplan(
    plan: PeriodePlan,
    blokkader: BlokKader,
    blauwdruk: DossierBlauwdruk,
    vorige_status: DossierStatus | None,
    openstaande_trajectdetails: list[dict[str, str]] | None = None,
) -> PeriodePlan:
    """Maak briefrelaties deterministisch geldig zonder contacten toe te voegen.

    Structured output dwingt de velden af, maar niet de betekenisvolle relatie
    tussen twee contacten. Daarom wordt een losse specialistenbrief hier aan een
    eerder contact van dezelfde episode gekoppeld. Als het om een acute opname
    gaat, wordt de brief een ontslagbrief. Alleen wanneer geen van beide
    klinisch verdedigbaar is, vervalt het losse briefonderdeel.
    """
    contacten = list(plan.contacten)
    openstaand_uit_vorig_blok = bool(
        vorige_status and vorige_status.openstaande_verwijstrajecten
    )
    laatste_blok = blokkader.bloknummer == max(
        blok.bloknummer for blok in blauwdruk.blokken
    )

    def contactdatum(contact: GeplandContact) -> date | None:
        try:
            return date.fromisoformat(contact.datum)
        except ValueError:
            return None

    def met_onderdelen(
        contact: GeplandContact,
        onderdelen: list[AanvullendOnderdeel] | set[AanvullendOnderdeel],
    ) -> GeplandContact:
        uniek = list(dict.fromkeys(onderdelen))
        if "Ontslagbrief" in uniek:
            uniek = [
                onderdeel
                for onderdeel in uniek
                if onderdeel not in {"Verwijsbrief", "Specialistenbrief"}
            ]
        if len(uniek) > 1 and "Geen" in uniek:
            uniek.remove("Geen")
        if not uniek:
            uniek = ["Geen"]
        return contact.model_copy(update={"aanvullende_onderdelen": uniek})

    # Maak combinaties zoals ['Geen', 'Laboratorium'] of een ontslagbrief plus
    # een regulier briefonderdeel eerst mechanisch eenduidig.
    contacten = [
        met_onderdelen(contact, contact.aanvullende_onderdelen)
        for contact in contacten
    ]

    # Herstel losse specialistenbrieven chronologisch. Een verwijzing wordt aan
    # het meest recente eerdere contact van dezelfde episode toegevoegd.
    for index, contact in enumerate(contacten):
        if "Specialistenbrief" not in contact.aanvullende_onderdelen:
            continue
        huidige_datum = contactdatum(contact)
        eerdere_verwijzingen = [
            eerder
            for eerder in contacten[:index]
            if normaliseer_episodenaam(eerder.episode)
            == normaliseer_episodenaam(contact.episode)
            and "Verwijsbrief" in eerder.aanvullende_onderdelen
            and contactdatum(eerder) is not None
            and huidige_datum is not None
            and contactdatum(eerder) < huidige_datum
        ]
        if eerdere_verwijzingen or openstaand_uit_vorig_blok:
            continue

        verwijs_kandidaten = [
            kandidaat_index
            for kandidaat_index, eerder in enumerate(contacten[:index])
            if normaliseer_episodenaam(eerder.episode)
            == normaliseer_episodenaam(contact.episode)
            and contactdatum(eerder) is not None
            and huidige_datum is not None
            and contactdatum(eerder) < huidige_datum
            and not (
                {"Ontslagbrief", "Specialistenbrief"}
                & set(eerder.aanvullende_onderdelen)
            )
        ]
        if verwijs_kandidaten:
            kandidaat_index = verwijs_kandidaten[-1]
            eerder = contacten[kandidaat_index]
            nieuwe_onderdelen = list(eerder.aanvullende_onderdelen)
            nieuwe_onderdelen.append("Verwijsbrief")
            contacten[kandidaat_index] = met_onderdelen(
                eerder,
                nieuwe_onderdelen,
            )
            continue

        # Een ziekenhuisbrief over een tijdens dezelfde acute opname ontdekte
        # diagnose is een ontslagbrief, niet een losse specialistenbrief.
        gekoppeld_aan_acute_opname = any(
            contactdatum(eerder) is not None
            and huidige_datum is not None
            and contact_past_bij_ontslagbrief(
                eerder.episode,
                contact.episode,
                contactdatum(eerder),
                huidige_datum,
                eerder.klinisch_doel,
                eerder.kerngebeurtenis,
            )
            for eerder in contacten[:index]
        )
        nieuwe_onderdelen = [
            onderdeel
            for onderdeel in contact.aanvullende_onderdelen
            if onderdeel != "Specialistenbrief"
        ]
        if gekoppeld_aan_acute_opname:
            nieuwe_onderdelen.append("Ontslagbrief")
        contacten[index] = met_onderdelen(contact, nieuwe_onderdelen)

    # In het laatste blok mag geen reguliere verwijzing open blijven. Koppel
    # zo mogelijk een later contact van dezelfde episode; verwijder de losse
    # verwijzing alleen wanneer zo'n retourmoment niet bestaat.
    if laatste_blok:
        for index, contact in enumerate(contacten):
            if "Verwijsbrief" not in contact.aanvullende_onderdelen:
                continue
            huidige_datum = contactdatum(contact)
            heeft_retourbrief = any(
                normaliseer_episodenaam(later.episode)
                == normaliseer_episodenaam(contact.episode)
                and "Specialistenbrief" in later.aanvullende_onderdelen
                and contactdatum(later) is not None
                and huidige_datum is not None
                and contactdatum(later) > huidige_datum
                for later in contacten[index + 1 :]
            )
            if heeft_retourbrief:
                continue
            retour_kandidaten = [
                later_index
                for later_index in range(index + 1, len(contacten))
                if normaliseer_episodenaam(contacten[later_index].episode)
                == normaliseer_episodenaam(contact.episode)
                and contactdatum(contacten[later_index]) is not None
                and huidige_datum is not None
                and contactdatum(contacten[later_index]) > huidige_datum
                and not (
                    {"Verwijsbrief", "Ontslagbrief"}
                    & set(contacten[later_index].aanvullende_onderdelen)
                )
            ]
            if retour_kandidaten:
                later_index = retour_kandidaten[0]
                later = contacten[later_index]
                nieuwe_onderdelen = list(later.aanvullende_onderdelen)
                nieuwe_onderdelen.append("Specialistenbrief")
                contacten[later_index] = met_onderdelen(
                    later,
                    nieuwe_onderdelen,
                )
            else:
                nieuwe_onderdelen = [
                    onderdeel
                    for onderdeel in contact.aanvullende_onderdelen
                    if onderdeel != "Verwijsbrief"
                ]
                contacten[index] = met_onderdelen(contact, nieuwe_onderdelen)

        # Sluit ook verwijzingen uit eerdere blokken. De modelstatus bevatte
        # voorheen alleen losse ID's, waardoor bijvoorbeeld een verwijzing uit
        # 2008 tot het einde van het dossier open kon blijven. Python koppelt
        # ieder resterend traject nu aan een geschikt bestaand contact in het
        # laatste blok, zonder een extra deelcontact toe te voegen.
        details = openstaande_trajectdetails or []
        geplande_retouren = Counter(
            normaliseer_episodenaam(contact.episode)
            for contact in contacten
            if "Specialistenbrief" in contact.aanvullende_onderdelen
        )
        nieuwe_verwijzingen = Counter(
            normaliseer_episodenaam(contact.episode)
            for contact in contacten
            if "Verwijsbrief" in contact.aanvullende_onderdelen
        )
        beschikbare_retouren_voor_eerdere_blokken = Counter(
            {
                episode: max(
                    0,
                    aantal - nieuwe_verwijzingen[episode],
                )
                for episode, aantal in geplande_retouren.items()
            }
        )
        gebruikte_contactindexen: set[int] = set()
        for detail in details:
            episode = detail.get("episode", "")
            episode_norm = normaliseer_episodenaam(episode)
            if beschikbare_retouren_voor_eerdere_blokken[episode_norm] > 0:
                beschikbare_retouren_voor_eerdere_blokken[episode_norm] -= 1
                continue

            verwijsdatum = veilige_iso_datum(
                detail.get("verwijsdatum", ""),
                "verwijsdatum van openstaand traject",
                [],
            )
            kandidaten = []
            for kandidaat_index, kandidaat in enumerate(contacten):
                kandidaatdatum = contactdatum(kandidaat)
                if kandidaat_index in gebruikte_contactindexen:
                    continue
                if kandidaatdatum is None:
                    continue
                if verwijsdatum is not None and kandidaatdatum <= verwijsdatum:
                    continue
                if {
                    "Verwijsbrief",
                    "Specialistenbrief",
                    "Ontslagbrief",
                } & set(kandidaat.aanvullende_onderdelen):
                    continue
                kandidaten.append(kandidaat_index)

            if not kandidaten:
                continue
            kandidaten.sort(
                key=lambda kandidaat_index: (
                    normaliseer_episodenaam(
                        contacten[kandidaat_index].episode
                    )
                    != episode_norm,
                    contacten[kandidaat_index].contactvorm != "Administratief",
                    contactdatum(contacten[kandidaat_index]) or date.max,
                    kandidaat_index,
                )
            )
            gekozen_index = kandidaten[0]
            gekozen = contacten[gekozen_index]
            nieuwe_onderdelen = list(gekozen.aanvullende_onderdelen)
            nieuwe_onderdelen.append("Specialistenbrief")
            traject_id = detail.get("traject_id", "onbekend traject")
            specialisme = detail.get("specialisme", "specialist")
            contacten[gekozen_index] = met_onderdelen(
                gekozen.model_copy(
                    update={
                        "klinisch_doel": (
                            gekozen.klinisch_doel
                            + f"; ontvangst retourbrief {specialisme} voor "
                            f"{episode} ({traject_id})"
                        ),
                        "kerngebeurtenis": (
                            gekozen.kerngebeurtenis
                            + f"; openstaand verwijstraject {traject_id} wordt "
                            "afgerond met een specialistenbrief"
                        ),
                    }
                ),
                nieuwe_onderdelen,
            )
            gebruikte_contactindexen.add(gekozen_index)

    return plan.model_copy(update={"contacten": contacten})


def synchroniseer_episodeontwerp_met_periodeplan(
    plan: PeriodePlan,
    blokkader: BlokKader,
    blauwdruk: DossierBlauwdruk,
) -> PeriodePlan:
    """Laat modelbedachte episodestartdatums aansluiten op het contactplan.

    De startdatums in de blauwdruk zijn geen gebruikersinvoer. Als een episode
    volgens de blauwdruk in dit blok begint, wordt de datum daarom gelijkgezet
    aan het eerste werkelijk geplande contact voor die episode. Een eerste
    contact dat ten onrechte als herstel of afsluiting is beschreven, wordt
    bovendien als eerste presentatie geformuleerd. Dit voorkomt dat de
    reviewer een fictieve modeldatum als onaantastbare randvoorwaarde gebruikt.
    """
    blokstart = date.fromisoformat(blokkader.startdatum)
    blokeinde = date.fromisoformat(blokkader.einddatum)
    contacten = list(plan.contacten)
    ontwerpen = list(blauwdruk.episodeontwerp)
    herstelwoorden = (
        "herstel",
        "hersteld",
        "nacontrole",
        "afsluiting",
        "afsluiten",
        "klachtenvrij",
    )

    for ontwerp_index, ontwerp in enumerate(ontwerpen):
        try:
            geplande_start = date.fromisoformat(ontwerp.geplande_startdatum)
        except ValueError:
            continue
        if not blokstart <= geplande_start <= blokeinde:
            continue

        episodecontacten = [
            (index, contact)
            for index, contact in enumerate(contacten)
            if normaliseer_episodenaam(contact.episode)
            == normaliseer_episodenaam(ontwerp.episode)
        ]
        if not episodecontacten:
            continue
        episodecontacten.sort(key=lambda item: item[1].datum)
        eerste_index, eerste_contact = episodecontacten[0]
        try:
            eerste_datum = date.fromisoformat(eerste_contact.datum)
        except ValueError:
            continue

        if eerste_datum != geplande_start:
            ontwerpen[ontwerp_index] = ontwerp.model_copy(
                update={"geplande_startdatum": eerste_datum.isoformat()}
            )

        eerste_tekst = (
            f"{eerste_contact.klinisch_doel} {eerste_contact.kerngebeurtenis}"
        ).casefold()
        if any(woord in eerste_tekst for woord in herstelwoorden):
            contacten[eerste_index] = eerste_contact.model_copy(
                update={
                    "klinisch_doel": (
                        f"Eerste presentatie en beoordeling van {ontwerp.episode}"
                    ),
                    "kerngebeurtenis": (
                        f"Start van episode {ontwerp.episode}: {ontwerp.kernbeloop}. "
                        f"Eerste beleid: {ontwerp.verwacht_beleid}"
                    ),
                }
            )

    blauwdruk.episodeontwerp = ontwerpen
    return plan.model_copy(update={"contacten": contacten})


def valideer_periodeplan(
    plan: PeriodePlan,
    blokkader: BlokKader,
    eerste_contactnummer: int,
    blauwdruk: DossierBlauwdruk,
    vorige_status: DossierStatus | None,
) -> list[str]:
    """Controleer aantallen, IDs, datums, correspondentie en episodeafsluiting."""
    fouten = []
    if (
        plan.bloknummer != blokkader.bloknummer
        or plan.startdatum != blokkader.startdatum
        or plan.einddatum != blokkader.einddatum
    ):
        fouten.append("bloknummer of periode van het periodeplan wijkt af")
    if plan.aantal_deelcontacten != blokkader.aantal_deelcontacten:
        fouten.append("aantal_deelcontacten in het periodeplan wijkt af")
    if len(plan.contacten) != blokkader.aantal_deelcontacten:
        fouten.append(
            f"verwacht {blokkader.aantal_deelcontacten} geplande contacten, maar "
            f"ontving {len(plan.contacten)}"
        )

    verwachte_ids = [
        f"C{nummer:03d}"
        for nummer in range(
            eerste_contactnummer,
            eerste_contactnummer + blokkader.aantal_deelcontacten,
        )
    ]
    ontvangen_ids = [contact.contact_id for contact in plan.contacten]
    if ontvangen_ids != verwachte_ids:
        fouten.append(
            f"contact-ID's moeten exact oplopen van {verwachte_ids[0]} tot "
            f"{verwachte_ids[-1]}"
        )

    blokstart = date.fromisoformat(blokkader.startdatum)
    blokeinde = date.fromisoformat(blokkader.einddatum)
    vorige_datum = None
    contactdatums: dict[str, date] = {}
    for contact in plan.contacten:
        contactdatum = veilige_iso_datum(
            contact.datum,
            f"datum van gepland contact {contact.contact_id}",
            fouten,
        )
        if contactdatum is None:
            continue
        if not blokstart <= contactdatum <= blokeinde:
            fouten.append(f"{contact.contact_id} valt buiten het tijdsblok")
        if vorige_datum is not None and contactdatum < vorige_datum:
            fouten.append("geplande contacten staan niet chronologisch")
        vorige_datum = contactdatum
        contactdatums[contact.contact_id] = contactdatum

    verwachte_verdeling = {
        regel.episode: regel.aantal_deelcontacten
        for regel in blokkader.episodeverdeling
    }
    ontvangen_verdeling = Counter(contact.episode for contact in plan.contacten)
    if ontvangen_verdeling != Counter(verwachte_verdeling):
        fouten.append(
            "episodeaantallen in het periodeplan wijken af van het blokkader: "
            f"verwacht {verwachte_verdeling}, ontving {dict(ontvangen_verdeling)}"
        )

    laatste_blok = blokkader.bloknummer == max(
        blok.bloknummer for blok in blauwdruk.blokken
    )
    openstaand_uit_vorig_blok = bool(
        vorige_status
        and vorige_status.openstaande_verwijstrajecten
    )
    for index, contact in enumerate(plan.contacten):
        onderdelen = set(contact.aanvullende_onderdelen)
        if "Geen" in onderdelen and len(onderdelen) > 1:
            fouten.append(
                f"{contact.contact_id} combineert 'Geen' met andere onderdelen"
            )
        if "Ontslagbrief" in onderdelen and (
            "Verwijsbrief" in onderdelen or "Specialistenbrief" in onderdelen
        ):
            fouten.append(
                f"{contact.contact_id} combineert een acute ontslagbrief met een "
                "regulier verwijstraject"
            )

        huidige_datum = contactdatums.get(contact.contact_id)
        eerdere_contacten = plan.contacten[:index]
        if "Ontslagbrief" in onderdelen:
            heeft_eerder_acute_contact = any(
                contactdatums.get(eerder.contact_id) is not None
                and huidige_datum is not None
                and contact_past_bij_ontslagbrief(
                    eerder.episode,
                    contact.episode,
                    contactdatums[eerder.contact_id],
                    huidige_datum,
                    eerder.klinisch_doel,
                    eerder.kerngebeurtenis,
                )
                for eerder in eerdere_contacten
            )
            if not heeft_eerder_acute_contact:
                fouten.append(
                    f"{contact.contact_id} plant een ontslagbrief zonder een eerder "
                    "contact voor dezelfde episode of een aantoonbaar gekoppelde "
                    "acute ziekenhuisopname in de voorafgaande 90 dagen "
                    f"(briefepisode: '{contact.episode}')"
                )

        if "Specialistenbrief" in onderdelen:
            heeft_eerdere_verwijzing = any(
                eerder.episode == contact.episode
                and "Verwijsbrief" in eerder.aanvullende_onderdelen
                and contactdatums.get(eerder.contact_id) is not None
                and huidige_datum is not None
                and contactdatums[eerder.contact_id] < huidige_datum
                for eerder in eerdere_contacten
            )
            if not heeft_eerdere_verwijzing and not openstaand_uit_vorig_blok:
                fouten.append(
                    f"{contact.contact_id} plant een specialistenbrief zonder "
                    "eerdere reguliere verwijzing"
                )

        if "Verwijsbrief" in onderdelen and laatste_blok:
            heeft_latere_specialistenbrief = any(
                later.episode == contact.episode
                and "Specialistenbrief" in later.aanvullende_onderdelen
                and contactdatums.get(later.contact_id) is not None
                and huidige_datum is not None
                and contactdatums[later.contact_id] > huidige_datum
                for later in plan.contacten[index + 1 :]
            )
            if not heeft_latere_specialistenbrief:
                fouten.append(
                    f"{contact.contact_id} plant in het laatste blok een "
                    "verwijsbrief zonder latere specialistenbrief"
                )

    for ontwerp in blauwdruk.episodeontwerp:
        if not ontwerp.geplande_einddatum:
            continue
        geplande_einddatum = veilige_iso_datum(
            ontwerp.geplande_einddatum,
            f"geplande einddatum van '{ontwerp.episode}'",
            fouten,
        )
        if (
            geplande_einddatum is None
            or not blokstart <= geplande_einddatum <= blokeinde
        ):
            continue
        vensterstart = max(
            blokstart,
            geplande_einddatum - timedelta(days=120),
        )
        afsluitende_contacten = [
            contact
            for contact in plan.contacten
            if contact.episode == ontwerp.episode
            and contactdatums.get(contact.contact_id) is not None
            and vensterstart
            <= contactdatums[contact.contact_id]
            <= geplande_einddatum
        ]
        if not afsluitende_contacten:
            fouten.append(
                f"episode '{ontwerp.episode}' eindigt op "
                f"{ontwerp.geplande_einddatum}, maar heeft in de 120 dagen "
                "ervoor geen gepland contact waarin herstel, resultaat of "
                "afsluiting kan worden vastgelegd"
            )

    return fouten


def valideer_dossierperiode(
    periode: DossierPeriode,
    plan: PeriodePlan,
    instellingen: PatientInstellingen,
    vorige_status: DossierStatus | None,
    laatste_blok: bool,
) -> list[str]:
    """Controleer het geschreven tijdsblok en de cumulatieve eindstatus."""
    fouten = []
    if (
        periode.bloknummer != plan.bloknummer
        or periode.startdatum != plan.startdatum
        or periode.einddatum != plan.einddatum
    ):
        fouten.append("identificatie of datums van het geschreven blok wijken af")
    if len(periode.contacten) != len(plan.contacten):
        fouten.append(
            f"geschreven blok bevat {len(periode.contacten)} contacten, plan bevat "
            f"{len(plan.contacten)}"
        )

    for gepland, geschreven in zip(plan.contacten, periode.contacten):
        verwachte_kern = (
            gepland.contact_id,
            gepland.datum,
            gepland.zorgverlener,
            gepland.contactvorm,
            gepland.episode,
        )
        ontvangen_kern = (
            geschreven.contact_id,
            geschreven.datum,
            geschreven.zorgverlener,
            geschreven.contactvorm,
            geschreven.episode,
        )
        if ontvangen_kern != verwachte_kern:
            fouten.append(
                f"geschreven contact {geschreven.contact_id} wijkt in ID, datum, "
                "zorgverlener, contactvorm of episode af van het plan"
            )

    blokstart = date.fromisoformat(plan.startdatum)
    blokeinde = date.fromisoformat(plan.einddatum)

    def controleer_blokdatum(waarde: str, label: str) -> None:
        gelezen = veilige_iso_datum(waarde, label, fouten)
        if gelezen is not None and not blokstart <= gelezen <= blokeinde:
            fouten.append(f"{label} valt buiten het huidige tijdsblok")

    for contact in periode.contacten:
        controleer_blokdatum(contact.datum, f"contact {contact.contact_id}")
    for brief in periode.correspondentie:
        controleer_blokdatum(brief.datum, f"brief {brief.traject_id}")
    for allergie in periode.allergieen:
        controleer_blokdatum(
            allergie.registratiedatum,
            f"allergieregistratie {allergie.allergeen}",
        )
    for uitslag in periode.microbiologie:
        controleer_blokdatum(uitslag.datum, f"microbiologie {uitslag.onderzoek}")
    for uitslag in periode.laboratorium:
        controleer_blokdatum(
            uitslag.datum,
            f"laboratoriumaanvraag {uitslag.aanvraag_id}",
        )

    geplande_briefsoorten = Counter(
        onderdeel
        for contact in plan.contacten
        for onderdeel in contact.aanvullende_onderdelen
        if onderdeel in {"Verwijsbrief", "Specialistenbrief", "Ontslagbrief"}
    )
    ontvangen_briefsoorten = Counter(
        brief.type_brief for brief in periode.correspondentie
    )
    if ontvangen_briefsoorten != geplande_briefsoorten:
        fouten.append(
            "briefsoorten in het geschreven blok wijken af van het plan: "
            f"verwacht {dict(geplande_briefsoorten)}, ontving "
            f"{dict(ontvangen_briefsoorten)}"
        )

    for brief in periode.correspondentie:
        if brief.type_brief != "Ontslagbrief":
            continue
        try:
            briefdatum = date.fromisoformat(brief.datum)
        except ValueError:
            continue
        passende_opnamecontacten = []
        for contact in periode.contacten:
            try:
                contactdatum = date.fromisoformat(contact.datum)
            except ValueError:
                continue
            if contact_past_bij_ontslagbrief(
                contact.episode,
                brief.episode,
                contactdatum,
                briefdatum,
                contact.s,
                contact.o,
                contact.e,
                contact.p,
            ):
                passende_opnamecontacten.append(contact)
        if not passende_opnamecontacten:
            fouten.append(
                f"ontslagbrief {brief.traject_id} heeft geen eerder contact voor "
                "dezelfde episode of een aantoonbaar gekoppelde acute "
                "ziekenhuisopname in de voorafgaande 90 dagen"
            )

    for medicijn in periode.medicatie:
        start = veilige_iso_datum(
            medicijn.startdatum,
            f"startdatum medicatie {medicijn.geneesmiddel}",
            fouten,
        )
        if start is not None and start > blokeinde:
            fouten.append(f"{medicijn.geneesmiddel} start na het huidige blok")
        if medicijn.status == "Actief" and medicijn.einddatum:
            fouten.append(f"actieve medicatie {medicijn.geneesmiddel} heeft einddatum")
        if medicijn.status != "Actief" and not medicijn.einddatum:
            fouten.append(
                f"{medicijn.status.lower()}e medicatie {medicijn.geneesmiddel} "
                "mist einddatum"
            )
        if medicijn.einddatum:
            einde = veilige_iso_datum(
                medicijn.einddatum,
                f"einddatum medicatie {medicijn.geneesmiddel}",
                fouten,
            )
            if einde is not None and not blokstart <= einde <= blokeinde:
                fouten.append(
                    f"einddatum van {medicijn.geneesmiddel} valt buiten dit blok"
                )

    status = periode.eindstatus
    if status.peildatum != plan.einddatum:
        fouten.append("peildatum van de eindstatus moet gelijk zijn aan het blokeinde")

    _, dossiereinde = bepaal_dossierperiode(instellingen.aantal_jaren)
    verwachte_geboortedatum = bepaal_fictieve_geboortedatum(
        instellingen.patient_id,
        instellingen.leeftijd,
        dossiereinde,
    )
    verwachte_leeftijd = bereken_leeftijd_op_datum(
        verwachte_geboortedatum,
        plan.einddatum,
    )
    if status.geboortedatum != verwachte_geboortedatum:
        fouten.append(
            "geboortedatum in de eindstatus wijkt af van de deterministisch "
            "vastgelegde fictieve geboortedatum"
        )
    if status.leeftijd_op_peildatum != verwachte_leeftijd:
        fouten.append(
            "leeftijd_op_peildatum in de eindstatus moet "
            f"{verwachte_leeftijd} jaar zijn op {plan.einddatum}"
        )

    verwachte_episodes = [episode.episode for episode in instellingen.episodes]
    ontvangen_statusnamen = [regel.episode for regel in status.episode_statussen]
    if ontvangen_statusnamen != verwachte_episodes:
        fouten.append(
            "de eindstatus moet alle episodes precies één keer en in Excel-volgorde "
            "bevatten"
        )

    for regel in status.episode_statussen:
        if regel.status == "Nog niet gestart":
            if regel.startdatum or regel.einddatum:
                fouten.append(
                    f"nog niet gestarte episode '{regel.episode}' heeft al een datum"
                )
            if laatste_blok:
                fouten.append(
                    f"episode '{regel.episode}' is na het laatste blok niet gestart"
                )
        else:
            start = veilige_iso_datum(
                regel.startdatum,
                f"startdatum in status van '{regel.episode}'",
                fouten,
            )
            if start is not None and start > blokeinde:
                fouten.append(f"status van '{regel.episode}' start in de toekomst")
            if regel.status == "Actief" and regel.einddatum:
                fouten.append(f"actieve episode '{regel.episode}' heeft een einddatum")
            if regel.status == "Afgesloten" and not regel.einddatum:
                fouten.append(f"afgesloten episode '{regel.episode}' mist einddatum")

    if vorige_status is not None:
        vorige_per_episode = {
            regel.episode: regel for regel in vorige_status.episode_statussen
        }
        for huidig in status.episode_statussen:
            vorig = vorige_per_episode.get(huidig.episode)
            if vorig is None:
                continue
            if vorig.status == "Afgesloten" and huidig.status != "Afgesloten":
                fouten.append(
                    f"afgesloten episode '{huidig.episode}' wordt opnieuw actief"
                )
            if vorig.startdatum and huidig.startdatum != vorig.startdatum:
                fouten.append(
                    f"startdatum van '{huidig.episode}' veranderde tussen blokken"
                )

        oude_medicatiesleutels = {
            (regel.geneesmiddel.casefold(), regel.startdatum)
            for regel in vorige_status.medicatiehistorie
        }
        nieuwe_medicatiesleutels = {
            (regel.geneesmiddel.casefold(), regel.startdatum)
            for regel in status.medicatiehistorie
        }
        ontbrekende_medicatie = oude_medicatiesleutels - nieuwe_medicatiesleutels
        if ontbrekende_medicatie:
            fouten.append(
                "cumulatieve medicatiehistorie verloor eerdere regels: "
                + ", ".join(naam for naam, _ in sorted(ontbrekende_medicatie))
            )

        oude_allergiesleutels = {
            (regel.allergeen.casefold(), regel.registratiedatum)
            for regel in vorige_status.allergieen
        }
        nieuwe_allergiesleutels = {
            (regel.allergeen.casefold(), regel.registratiedatum)
            for regel in status.allergieen
        }
        if not oude_allergiesleutels.issubset(nieuwe_allergiesleutels):
            fouten.append("cumulatieve allergielijst verloor een eerdere registratie")

    if instellingen.allergie_modus == "Geen" and (
        periode.allergieen or status.allergieen
    ):
        fouten.append("allergie_modus is Geen, maar er zijn allergieën gegenereerd")
    if instellingen.microbiologie_modus == "Geen" and periode.microbiologie:
        fouten.append(
            "microbiologie_modus is Geen, maar er zijn uitslagen gegenereerd"
        )
    if instellingen.laboratorium_modus == "Geen" and periode.laboratorium:
        fouten.append(
            "laboratorium_modus is Geen, maar er zijn uitslagen gegenereerd"
        )

    return fouten


def normaliseer_correspondentietrajecten(
    perioden: list[DossierPeriode],
) -> list[DossierPeriode]:
    """Ken dossierbreed unieke en onderling passende traject-ID's toe.

    Een taalmodel kan ondanks de prompt in een later tijdsblok opnieuw ``T001``
    gebruiken. Een traject-ID is daarom geen vrije modelbeslissing: Python
    nummert alle trajecten chronologisch opnieuw. Een reguliere verwijsbrief en
    de bijbehorende latere specialistenbrief behouden samen één ID; iedere
    ontslagbrief krijgt een eigen ID. Open verwijzingen blijven beschikbaar om
    in een later blok door een specialistenbrief te worden gesloten.
    """
    volgende_nummer = 1
    # (nieuwe ID, oorspronkelijke ID, episode, specialisme, verwijsdatum)
    openstaande_verwijzingen: list[tuple[str, str, str, str, date]] = []
    resultaat: list[DossierPeriode] = []

    def nieuw_traject_id() -> str:
        nonlocal volgende_nummer
        waarde = f"T{volgende_nummer:03d}"
        volgende_nummer += 1
        return waarde

    def briefdatum(brief: CorrespondentieRegel) -> date:
        # De datum is vóór deze normalisatie al gevalideerd. De fallback houdt
        # deze herstelstap desondanks totaal en voorspelbaar.
        try:
            return date.fromisoformat(brief.datum)
        except ValueError:
            return date.min

    for periode in perioden:
        aangepaste_brieven: dict[int, CorrespondentieRegel] = {}
        typevolgorde = {
            "Verwijsbrief": 0,
            "Specialistenbrief": 1,
            "Ontslagbrief": 2,
        }
        geordende_brieven = sorted(
            enumerate(periode.correspondentie),
            key=lambda item: (
                briefdatum(item[1]),
                typevolgorde[item[1].type_brief],
                item[0],
            ),
        )

        for oorspronkelijke_index, brief in geordende_brieven:
            oorspronkelijk_id = brief.traject_id.strip().casefold()
            episode = normaliseer_episodenaam(brief.episode)
            specialisme = brief.specialisme.strip().casefold()
            datum_brief = briefdatum(brief)

            if brief.type_brief == "Verwijsbrief":
                toegewezen_id = nieuw_traject_id()
                openstaande_verwijzingen.append(
                    (
                        toegewezen_id,
                        oorspronkelijk_id,
                        episode,
                        specialisme,
                        datum_brief,
                    )
                )
            elif brief.type_brief == "Ontslagbrief":
                # Een acute ontslagbrief is altijd een zelfstandig traject.
                toegewezen_id = nieuw_traject_id()
            else:
                kandidaten = [
                    (index, verwijzing)
                    for index, verwijzing in enumerate(openstaande_verwijzingen)
                    if verwijzing[4] < datum_brief
                    and (
                        verwijzing[1] == oorspronkelijk_id
                        or verwijzing[2] == episode
                        or verwijzing[3] == specialisme
                    )
                ]

                if kandidaten:
                    def overeenkomstscore(
                        kandidaat: tuple[
                            int,
                            tuple[str, str, str, str, date],
                        ],
                    ) -> tuple[int, int]:
                        _, verwijzing = kandidaat
                        zelfde_id = verwijzing[1] == oorspronkelijk_id
                        zelfde_episode = verwijzing[2] == episode
                        zelfde_specialisme = verwijzing[3] == specialisme
                        if zelfde_id and zelfde_episode and zelfde_specialisme:
                            score = 0
                        elif zelfde_episode and zelfde_specialisme:
                            score = 1
                        elif zelfde_id and zelfde_episode:
                            score = 2
                        elif zelfde_episode:
                            score = 3
                        elif zelfde_id and zelfde_specialisme:
                            score = 4
                        elif zelfde_specialisme:
                            score = 5
                        else:
                            score = 6
                        # Bij gelijke overeenkomst hoort de meest recente
                        # passende verwijzing bij de retourbrief.
                        return score, -verwijzing[4].toordinal()

                    gekozen_index, gekozen = min(
                        kandidaten,
                        key=overeenkomstscore,
                    )
                    toegewezen_id = gekozen[0]
                    openstaande_verwijzingen.pop(gekozen_index)
                else:
                    # Laat een echte losse specialistenbrief zichtbaar als een
                    # zelfstandig ongeldig traject. De eindcontrole geeft dan
                    # terecht een gerichte fout in plaats van een verkeerde
                    # verwijzing te verzinnen.
                    toegewezen_id = nieuw_traject_id()

            aangepaste_brieven[oorspronkelijke_index] = brief.model_copy(
                update={"traject_id": toegewezen_id}
            )

        correspondentie = [
            aangepaste_brieven[index]
            for index in range(len(periode.correspondentie))
        ]
        openstaande_ids = [
            verwijzing[0] for verwijzing in openstaande_verwijzingen
        ]
        eindstatus = periode.eindstatus.model_copy(
            update={"openstaande_verwijstrajecten": openstaande_ids}
        )
        resultaat.append(
            periode.model_copy(
                update={
                    "correspondentie": correspondentie,
                    "eindstatus": eindstatus,
                }
            )
        )

    return resultaat


def bepaal_openstaande_trajectdetails(
    perioden: list[DossierPeriode],
) -> list[dict[str, str]]:
    """Beschrijf reguliere verwijzingen waarvoor nog geen retourbrief bestaat."""
    trajecten: dict[str, list[CorrespondentieRegel]] = {}
    for periode in perioden:
        for brief in periode.correspondentie:
            trajecten.setdefault(brief.traject_id, []).append(brief)

    details = []
    for traject_id, brieven in trajecten.items():
        verwijzingen = [
            brief for brief in brieven if brief.type_brief == "Verwijsbrief"
        ]
        retourbrieven = [
            brief for brief in brieven if brief.type_brief == "Specialistenbrief"
        ]
        if len(verwijzingen) != 1 or retourbrieven:
            continue
        verwijzing = verwijzingen[0]
        details.append(
            {
                "traject_id": traject_id,
                "verwijsdatum": verwijzing.datum,
                "episode": verwijzing.episode,
                "specialisme": verwijzing.specialisme,
                "onderwerp": verwijzing.onderwerp,
            }
        )
    return sorted(
        details,
        key=lambda detail: (detail["verwijsdatum"], detail["traject_id"]),
    )


def voeg_dossierperioden_samen(
    blauwdruk: DossierBlauwdruk,
    perioden: list[DossierPeriode],
) -> SynthetischDossier:
    """Maak van goedgekeurde tijdsblokken één einddossier."""
    if not perioden:
        raise RuntimeError("Er zijn geen dossierperioden om samen te voegen.")

    # Defensieve eindnormalisatie: ook als een aanroeper de normalisatie tussen
    # blokken overslaat, kunnen traject-ID's nooit door modelhergebruik botsen.
    perioden = normaliseer_correspondentietrajecten(perioden)

    eindstatus = perioden[-1].eindstatus
    episodelijst = []
    for regel in eindstatus.episode_statussen:
        if regel.status == "Nog niet gestart":
            raise RuntimeError(
                f"Episode '{regel.episode}' is in het einddossier niet gestart."
            )
        episodelijst.append(
            EpisodeOverzichtRegel(
                volgorde=regel.volgorde,
                episode=regel.episode,
                icpc_code=regel.icpc_code,
                startdatum=regel.startdatum,
                einddatum=regel.einddatum,
                status=regel.status,
                attentiewaarde=regel.attentiewaarde,
                samenvatting_beloop=regel.samenvatting_beloop,
                beleid=regel.beleid,
            )
        )

    return SynthetischDossier(
        patient_achtergrond=blauwdruk.patient_achtergrond,
        episodelijst=episodelijst,
        contacten=sorted(
            [contact for periode in perioden for contact in periode.contacten],
            key=lambda contact: (contact.datum, contact.contact_id),
        ),
        # De laatste status bevat de cumulatieve, bijgewerkte historie en voorkomt
        # dubbele actieve regels uit opeenvolgende blokken.
        medicatie=eindstatus.medicatiehistorie,
        correspondentie=sorted(
            [brief for periode in perioden for brief in periode.correspondentie],
            key=lambda brief: (brief.datum, brief.traject_id, brief.type_brief),
        ),
        allergieen=eindstatus.allergieen,
        microbiologie=sorted(
            [uitslag for periode in perioden for uitslag in periode.microbiologie],
            key=lambda uitslag: uitslag.datum,
        ),
        laboratorium=sorted(
            [uitslag for periode in perioden for uitslag in periode.laboratorium],
            key=lambda uitslag: (
                uitslag.datum,
                uitslag.aanvraag_id,
                uitslag.bepaling,
            ),
        ),
    )


def valideer_iteratief_einddossier(
    dossier: SynthetischDossier,
    instellingen: PatientInstellingen,
    open_verwijzingen_als_waarschuwing: bool = False,
) -> list[str]:
    """Voer aanvullende deterministische controles uit over alle blokgrenzen."""
    valideer_gegenereerd_dossier(dossier, instellingen)
    fouten = []
    waarschuwingen = []
    starttekst, eindtekst = bepaal_dossierperiode(instellingen.aantal_jaren)
    dossierstart = date.fromisoformat(starttekst)
    dossiereinde = date.fromisoformat(eindtekst)

    verwachte_ids = [
        f"C{nummer:03d}"
        for nummer in range(1, instellingen.aantal_deelcontacten + 1)
    ]
    ids_op_nummer = sorted(
        [contact.contact_id for contact in dossier.contacten],
        key=lambda waarde: int(waarde[1:]) if waarde[1:].isdigit() else 10**9,
    )
    if ids_op_nummer != verwachte_ids:
        fouten.append("de contact-ID's vormen niet exact de reeks C001 t/m het totaal")

    vorige_datum = None
    for contact in dossier.contacten:
        contactdatum = veilige_iso_datum(
            contact.datum,
            f"datum van {contact.contact_id}",
            fouten,
        )
        if contactdatum is None:
            continue
        if not dossierstart <= contactdatum <= dossiereinde:
            fouten.append(f"{contact.contact_id} ligt buiten de dossierperiode")
        if vorige_datum is not None and contactdatum < vorige_datum:
            fouten.append("het samengevoegde journaal staat niet chronologisch")
        vorige_datum = contactdatum

    if all(
        episode.aantal_deelcontacten is not None
        for episode in instellingen.episodes
    ):
        telling = Counter(contact.episode for contact in dossier.contacten)
        for episode in instellingen.episodes:
            if telling[episode.episode] != episode.aantal_deelcontacten:
                fouten.append(
                    f"'{episode.episode}' heeft {telling[episode.episode]} contacten "
                    f"in plaats van {episode.aantal_deelcontacten}"
                )

    trajecten: dict[str, list[CorrespondentieRegel]] = {}
    for brief in dossier.correspondentie:
        trajecten.setdefault(brief.traject_id, []).append(brief)
    for traject_id, brieven in trajecten.items():
        typen = Counter(brief.type_brief for brief in brieven)
        if len(brieven) == 1 and typen == Counter({"Ontslagbrief": 1}):
            ontslagbrief = brieven[0]
            try:
                ontslagdatum = date.fromisoformat(ontslagbrief.datum)
            except ValueError:
                continue
            passende_opnamecontacten = []
            for contact in dossier.contacten:
                try:
                    contactdatum = date.fromisoformat(contact.datum)
                except ValueError:
                    continue
                if contact_past_bij_ontslagbrief(
                    contact.episode,
                    ontslagbrief.episode,
                    contactdatum,
                    ontslagdatum,
                    contact.s,
                    contact.o,
                    contact.e,
                    contact.p,
                ):
                    passende_opnamecontacten.append(contact)
            if not passende_opnamecontacten:
                fouten.append(
                    f"ontslagbrief van {traject_id} heeft geen eerder contact voor "
                    "dezelfde episode of een aantoonbaar gekoppelde acute "
                    "ziekenhuisopname in de voorafgaande 90 dagen"
                )
            continue
        if len(brieven) == 1 and typen == Counter({"Verwijsbrief": 1}):
            verwijzing = brieven[0]
            melding = (
                f"traject {traject_id} bevat een verwijsbrief op "
                f"{verwijzing.datum} voor '{verwijzing.episode}', maar geen "
                "ontvangen specialistenbrief"
            )
            if open_verwijzingen_als_waarschuwing:
                waarschuwingen.append(melding)
                continue
            fouten.append(melding)
            continue
        if len(brieven) != 2 or typen != Counter(
            {"Verwijsbrief": 1, "Specialistenbrief": 1}
        ):
            briefdetails = ", ".join(
                f"{brief.type_brief} op {brief.datum} ({brief.episode})"
                for brief in sorted(brieven, key=lambda regel: regel.datum)
            )
            fouten.append(
                f"traject {traject_id} moet óf één acute ontslagbrief bevatten, "
                "óf precies één verwijsbrief en één specialistenbrief; ontvangen: "
                f"{briefdetails or 'geen brieven'}"
            )
            continue
        verwijzing = next(
            brief for brief in brieven if brief.type_brief == "Verwijsbrief"
        )
        specialist = next(
            brief for brief in brieven if brief.type_brief == "Specialistenbrief"
        )
        if specialist.datum <= verwijzing.datum:
            fouten.append(
                f"specialistenbrief van {traject_id} is niet later dan de verwijzing"
            )

    for medicijn in dossier.medicatie:
        if medicijn.status == "Actief" and medicijn.einddatum:
            fouten.append(f"actieve medicatie {medicijn.geneesmiddel} heeft einddatum")
        if medicijn.status != "Actief" and not medicijn.einddatum:
            fouten.append(
                f"{medicijn.status.lower()}e medicatie {medicijn.geneesmiddel} "
                "mist einddatum"
            )

    if instellingen.allergie_modus == "Geen" and dossier.allergieen:
        fouten.append("allergie_modus is Geen, maar de eindlijst is niet leeg")
    if instellingen.microbiologie_modus == "Geen" and dossier.microbiologie:
        fouten.append("microbiologie_modus is Geen, maar de eindlijst is niet leeg")
    if instellingen.laboratorium_modus == "Geen" and dossier.laboratorium:
        fouten.append("laboratorium_modus is Geen, maar de eindlijst is niet leeg")

    if instellingen.allergie_modus == "Zelf invoeren":
        ontvangen = {regel.allergeen.casefold() for regel in dossier.allergieen}
        for verwacht in instellingen.handmatige_allergieen:
            if verwacht.allergeen.casefold() not in ontvangen:
                fouten.append(
                    f"handmatig opgegeven allergie '{verwacht.allergeen}' ontbreekt"
                )
    if instellingen.laboratorium_modus == "Zelf invoeren":
        ontvangen = {regel.bepaling.casefold() for regel in dossier.laboratorium}
        for verwacht in instellingen.handmatig_laboratorium:
            if verwacht.bepaling.casefold() not in ontvangen:
                fouten.append(
                    f"handmatig opgegeven labbepaling '{verwacht.bepaling}' ontbreekt"
                )

    if fouten:
        raise RuntimeError(
            "Het samengevoegde iteratieve dossier voldoet niet aan de "
            "eindcontroles:\n- "
            + "\n- ".join(fouten)
        )
    return waarschuwingen


def laad_laatste_herstelbare_run(
    instellingen: PatientInstellingen,
    instellingenpad: Path,
) -> tuple[Path, DossierBlauwdruk, list[DossierPeriode]]:
    """Laad de nieuwste volledige, maar nog niet geëxporteerde iteratieve run."""
    basismap = Path(__file__).with_name("iteratieve_runs")
    if not basismap.exists():
        raise FileNotFoundError(
            "De map 'iteratieve_runs' bestaat niet; er is geen run om te herstellen."
        )

    verwachte_hash = hashlib.sha256(instellingenpad.read_bytes()).hexdigest()
    kandidaten = sorted(
        (pad for pad in basismap.iterdir() if pad.is_dir()),
        key=lambda pad: pad.stat().st_mtime,
        reverse=True,
    )
    redenen = []

    for runmap in kandidaten:
        try:
            metadata = json.loads(
                (runmap / "00_metadata.json").read_text(encoding="utf-8")
            )
            if metadata.get("patient_id") != instellingen.patient_id:
                continue
            if metadata.get("settings_sha256") != verwachte_hash:
                continue

            blauwdrukpad = runmap / "00_blauwdruk_goedgekeurd.json"
            blauwdruk = DossierBlauwdruk.model_validate_json(
                blauwdrukpad.read_text(encoding="utf-8")
            )
            perioden = []
            for blokkader in blauwdruk.blokken:
                nummer = blokkader.bloknummer
                mogelijke_paden = [
                    runmap
                    / f"{nummer:02d}_dossierperiode_trajecten_genormaliseerd.json",
                    runmap / f"{nummer:02d}_dossierperiode_goedgekeurd.json",
                    runmap / f"{nummer:02d}_dossierperiode_structureel_geldig.json",
                ]
                periodepad = next(
                    (pad for pad in mogelijke_paden if pad.exists()),
                    None,
                )
                if periodepad is None:
                    raise FileNotFoundError(
                        f"tijdsblok {nummer} heeft geen voltooid dossiercheckpoint"
                    )
                perioden.append(
                    DossierPeriode.model_validate_json(
                        periodepad.read_text(encoding="utf-8")
                    )
                )

            if len(perioden) != len(blauwdruk.blokken):
                raise RuntimeError("niet alle tijdsblokken konden worden geladen")
            return runmap, blauwdruk, perioden
        except (FileNotFoundError, ValueError, RuntimeError) as fout:
            redenen.append(f"{runmap.name}: {fout}")

    detail = f" Laatste controlepunten: {'; '.join(redenen[:3])}" if redenen else ""
    raise FileNotFoundError(
        "Geen volledige eerdere run gevonden voor de huidige patiënt en exact "
        f"dezelfde settings.{detail}"
    )


def herstel_laatste_run() -> None:
    """Herstel en exporteer een voltooide run na een fout bij eindvalidatie."""
    instellingenpad = Path(__file__).with_name(SETTINGS_BESTANDSNAAM)
    instellingen = lees_patientinstellingen(instellingenpad)
    runmap, blauwdruk, perioden = laad_laatste_herstelbare_run(
        instellingen,
        instellingenpad,
    )
    perioden = normaliseer_correspondentietrajecten(perioden)
    for periode in perioden:
        schrijf_checkpoint(
            runmap
            / (
                f"{periode.bloknummer:02d}_dossierperiode_"
                "trajecten_genormaliseerd.json"
            ),
            periode,
        )

    dossier = voeg_dossierperioden_samen(blauwdruk, perioden)
    herstelwaarschuwingen = valideer_iteratief_einddossier(
        dossier,
        instellingen,
        open_verwijzingen_als_waarschuwing=True,
    )

    tijdstip = datetime.now().strftime("%Y%m%d_%H%M%S")
    patient_id = veilige_bestandsnaam(instellingen.patient_id)
    uitvoerpad = Path(__file__).with_name(
        f"synthetisch_epd_{patient_id}_{tijdstip}_iteratief_hersteld.xlsx"
    )
    dossier_naar_excel(dossier, instellingen, uitvoerpad)
    schrijf_checkpoint(runmap / "99_einddossier_hersteld.json", dossier)
    schrijf_checkpoint(
        runmap / "99_herstelwaarschuwingen.json",
        {"waarschuwingen": herstelwaarschuwingen},
    )
    schrijf_checkpoint(
        runmap / "99_herstelmetadata.json",
        {
            "hersteld_met_pipelineversie": PIPELINE_VERSIE,
            "hersteld_op": datetime.now().isoformat(timespec="seconds"),
            "bron_runmap": str(runmap.resolve()),
            "reden": (
                "Dossierbrede normalisatie van correspondentie-ID's na een "
                "eerdere fout bij de eindvalidatie."
            ),
        },
    )
    print(f"Herstelde run: {runmap.resolve()}")
    print(f"Klaar: {uitvoerpad.resolve()}")
    print(
        "LET OP: alleen de bestaande, reeds gegenereerde blokken zijn hersteld "
        "en geëxporteerd; voer daarna nog steeds de handmatige checklist uit."
    )
    if herstelwaarschuwingen:
        print(
            "LET OP: de bestaande run bevat openstaande verwijstrajecten zonder "
            "retourbrief. Deze zijn niet kunstmatig ingevuld en staan als "
            "waarschuwing in 99_herstelwaarschuwingen.json."
        )


def maak_compacte_eindsamenvatting(
    dossier: SynthetischDossier,
    perioden: list[DossierPeriode],
    instellingen: PatientInstellingen,
) -> dict:
    """Maak een compacte invoer voor de laatste reviewer-agent."""
    dossierstart, dossiereinde = bepaal_dossierperiode(
        instellingen.aantal_jaren
    )
    geboortedatum = bepaal_fictieve_geboortedatum(
        instellingen.patient_id,
        instellingen.leeftijd,
        dossiereinde,
    )
    return {
        "patient_id": instellingen.patient_id,
        "geboortedatum": geboortedatum,
        "leeftijd_einde_dossier": instellingen.leeftijd,
        "dossierperiode": [dossierstart, dossiereinde],
        "aantallen": {
            "deelcontacten": len(dossier.contacten),
            "episodes": len(dossier.episodelijst),
            "medicatieregels": len(dossier.medicatie),
            "brieven": len(dossier.correspondentie),
            "allergieen": len(dossier.allergieen),
            "microbiologie": len(dossier.microbiologie),
            "laboratoriumregels": len(dossier.laboratorium),
        },
        "contacten_per_episode": dict(
            Counter(contact.episode for contact in dossier.contacten)
        ),
        "contacten_per_zorgverlener": dict(
            Counter(contact.zorgverlener for contact in dossier.contacten)
        ),
        "episodelijst": [
            regel.model_dump(mode="json") for regel in dossier.episodelijst
        ],
        "medicatie": [regel.model_dump(mode="json") for regel in dossier.medicatie],
        "correspondentie_metadata": [
            {
                "traject_id": brief.traject_id,
                "datum": brief.datum,
                "type": brief.type_brief,
                "specialisme": brief.specialisme,
                "episode": brief.episode,
            }
            for brief in dossier.correspondentie
        ],
        "periodesamenvattingen": [
            {
                "bloknummer": periode.bloknummer,
                "periode": [periode.startdatum, periode.einddatum],
                "samenvatting": periode.samenvatting_periode,
                "eindstatus": periode.eindstatus.model_dump(mode="json"),
            }
            for periode in perioden
        ],
    }


def formatteer_feedback(fouten: list[str]) -> str:
    if not fouten:
        return ""
    return (
        "\n\nHERSTEL DEZE PUNTEN IN EEN VOLLEDIG NIEUWE UITVOER:\n- "
        + "\n- ".join(fouten)
    )


async def genereer_blauwdruk(
    agent: Agent,
    instellingen: PatientInstellingen,
    tijdsblokken: list[tuple[int, str, str]],
    runmap: Path,
) -> DossierBlauwdruk:
    """Genereer en valideer de globale patiëntblauwdruk."""
    patient_context = maak_patient_prompt(instellingen)
    blokregels = "\n".join(
        f"- blok {nummer}: {start} tot en met {einde}"
        for nummer, start, einde in tijdsblokken
    )
    basisinvoer = f"""
Maak nog GEEN uitgewerkte SOEP-notities. Ontwerp eerst uitsluitend de globale
patiëntblauwdruk voor de onderstaande configuratie.

{patient_context}

Gebruik exact deze tijdsblokken:
{blokregels}

Verdeel alle {instellingen.aantal_deelcontacten} contacten over de blokken en
binnen ieder blok over alle episodes. Iedere episode moet in iedere
episodeverdeling precies één keer en in de Excel-volgorde staan; gebruik nul als
de episode in dat blok geen contact heeft. Respecteer exacte episodeaantallen als
die in Excel zijn ingevuld. Geef een afgesloten episode altijd minimaal één
contact in het tijdsblok waarin de geplande einddatum valt, zodat het latere
contactplan herstel, resultaat of afsluiting kan vastleggen. Plan het
longitudinale beloop, maar schrijf nog geen S-, O-, E- of P-regels.
""".strip()

    feedback = []
    for poging in range(1, MAX_HERSTELPOGINGEN + 2):
        resultaat = await Runner.run(
            agent,
            basisinvoer + formatteer_feedback(feedback),
        )
        blauwdruk = resultaat.final_output
        if blauwdruk is None:
            feedback = ["de agent leverde geen gestructureerde blauwdruk op"]
            continue
        schrijf_checkpoint(
            runmap / f"00_blauwdruk_poging_{poging}.json",
            blauwdruk,
        )
        blauwdruk = herstel_mechanische_blauwdrukwaarden(
            blauwdruk,
            instellingen,
            tijdsblokken,
        )
        schrijf_checkpoint(
            runmap / f"00_blauwdruk_poging_{poging}_genormaliseerd.json",
            blauwdruk,
        )
        feedback = valideer_blauwdruk(blauwdruk, instellingen, tijdsblokken)
        if not feedback:
            schrijf_checkpoint(runmap / "00_blauwdruk_goedgekeurd.json", blauwdruk)
            return blauwdruk

    raise RuntimeError(
        "De blauwdruk bleef ongeldig na herstelpogingen:\n- "
        + "\n- ".join(feedback)
    )


async def genereer_periodeplan(
    planner_agent: Agent,
    reviewer_agent: Agent,
    blokkader: BlokKader,
    blauwdruk: DossierBlauwdruk,
    instellingen: PatientInstellingen,
    vorige_status: DossierStatus | None,
    openstaande_trajectdetails: list[dict[str, str]],
    eerste_contactnummer: int,
    runmap: Path,
) -> tuple[PeriodePlan, PlanBeoordeling]:
    """Plan één tijdsblok en laat het plan vóór het schrijven beoordelen."""
    laatste_contactnummer = (
        eerste_contactnummer + blokkader.aantal_deelcontacten - 1
    )
    vorige_status_tekst = (
        model_naar_json(vorige_status)
        if vorige_status is not None
        else "Nog geen eerdere dossierstatus: dit is het eerste tijdsblok."
    )
    leeftijdskader = maak_leeftijdskader(
        instellingen,
        blokkader.startdatum,
        blokkader.einddatum,
    )
    basisinvoer = f"""
Maak uitsluitend het concrete contactplan voor tijdsblok {blokkader.bloknummer}.
Schrijf nog geen volledige SOEP-regels.

PATIËNTINSTELLINGEN:
{model_naar_json(instellingen)}

GLOBAAL EPISODEONTWERP:
{model_naar_json(blauwdruk.episodeontwerp)}

KADER VAN DIT BLOK:
{model_naar_json(blokkader)}

DETERMINISTISCH LEEFTIJDSKADER:
{leeftijdskader}

CUMULATIEVE STATUS AAN HET BEGIN:
{vorige_status_tekst}

OPENSTAANDE REGULIERE VERWIJSTRAJECTEN UIT EERDERE BLOKKEN:
{model_naar_json(openstaande_trajectdetails)}

Ieder openstaand traject moet uiterlijk in het laatste tijdsblok worden
afgerond met één latere specialistenbrief. Gebruik daarbij hetzelfde traject-ID,
dezelfde episode en hetzelfde specialisme als in dit overzicht. Voeg hiervoor
geen extra deelcontact toe, maar koppel de retourbrief aan een passend bestaand
contact binnen het toegewezen aantal.

Maak exact {blokkader.aantal_deelcontacten} contacten, chronologisch binnen
{blokkader.startdatum} tot en met {blokkader.einddatum}. Gebruik exact de
contact-ID's C{eerste_contactnummer:03d} tot en met
C{laatste_contactnummer:03d}. Laat de aantallen per episode exact aansluiten op
de episodeverdeling van het blokkader. Plan controles, medicatieveranderingen,
uitslagen en correspondentie op klinisch samenhangende momenten. Een
specialistenbrief moet later zijn dan de bijbehorende verwijzing, eventueel in
een volgend blok. Gebruik 'Ontslagbrief' voor correspondentie na een acute
SEH-presentatie of ziekenhuisopname zonder voorafgaande huisartsverwijzing.
Plan deze op een latere datum dan het acute contact en verzin geen
huisartsverwijzing. Gebruik 'Specialistenbrief' alleen als retourbrief na een
reguliere verwijzing. Plan in het laatste tijdsblok geen nieuwe verwijsbrief
zonder een latere specialistenbrief in hetzelfde blok.

Een diagnose die tijdens de acute opname wordt vastgesteld, mag de episode van
de ontslagbrief zijn terwijl het eerdere acute contact een andere primaire
episode heeft. Voorbeeld: een acuut CVA-contact kan worden gevolgd door een
ontslagbrief onder Atriumfibrilleren als dit tijdens dezelfde opname is ontdekt.
Zet dan in het eerdere contact expliciet ambulance, SEH of ziekenhuisopname en
plan de ontslagbrief binnen 90 dagen na dat contact.

Controleer het globale episodeontwerp vóór je de contacten vastlegt. Voor
iedere episode met een geplande einddatum in dit blok moet in de laatste 120
dagen vóór die einddatum minimaal één contact voor diezelfde episode staan.
Laat daarin herstel, een operatieresultaat, nacontrole of afsluiting zichtbaar
worden. Dit afsluitcontact valt binnen het reeds toegewezen episodeaantal en is
nooit een extra contact bovenop de episodeverdeling. Verschuif of combineer dus
een bestaand contact als dat nodig is. Een episode mag niet alleen in de
samenvatting of eindstatus worden afgesloten. Als specialistische behandeling
de afsluiting bepaalt, plan dan ook de passende latere specialisten- of
ontslagbrief binnen hetzelfde toegewezen aantal.
""".strip()

    feedback = []
    laatste_beoordeling = None
    laatste_structureel_geldige_plan = None
    laatste_structureel_geldige_beoordeling = None
    for poging in range(1, MAX_HERSTELPOGINGEN + 2):
        resultaat = await Runner.run(
            planner_agent,
            basisinvoer + formatteer_feedback(feedback),
        )
        plan = resultaat.final_output
        if plan is None:
            feedback = ["de planner leverde geen gestructureerd periodeplan op"]
            continue
        schrijf_checkpoint(
            runmap
            / f"{blokkader.bloknummer:02d}_periodeplan_poging_{poging}.json",
            plan,
        )
        plan = herstel_mechanische_periodeplanwaarden(
            plan,
            blokkader,
            blauwdruk,
        )
        plan = herstel_correspondentie_in_periodeplan(
            plan,
            blokkader,
            blauwdruk,
            vorige_status,
            openstaande_trajectdetails,
        )
        plan = synchroniseer_episodeontwerp_met_periodeplan(
            plan,
            blokkader,
            blauwdruk,
        )
        schrijf_checkpoint(
            runmap
            / (
                f"{blokkader.bloknummer:02d}_periodeplan_poging_"
                f"{poging}_genormaliseerd.json"
            ),
            plan,
        )
        schrijf_checkpoint(
            runmap
            / (
                f"00_blauwdruk_bijgesteld_na_blok_{blokkader.bloknummer}_"
                f"poging_{poging}.json"
            ),
            blauwdruk,
        )

        feedback = valideer_periodeplan(
            plan,
            blokkader,
            eerste_contactnummer,
            blauwdruk,
            vorige_status,
        )
        if feedback:
            continue
        laatste_structureel_geldige_plan = plan

        review_invoer = f"""
Beoordeel uitsluitend of dit geplande tijdsblok intern logisch, chronologisch,
klinisch aannemelijk en consistent met de eerdere status is. Een waarschuwing
zonder concrete inconsistentie is geen kritieke fout. Zet 'goedgekeurd' alleen
op false als het plan vóór het schrijven echt moet worden hersteld.

Dit is nog een compact contactplan en geen uitgewerkt dossier. Eis daarom in
deze fase geen volledige vitale parameters, exacte onderzoeksdetails,
antibioticakeuze of complete SOEP-inhoud; die worden pas door de schrijver
uitgewerkt. Een ontbrekend detail zoals welk urineonderzoek wordt verricht is
hoogstens een waarschuwing als datum, episode en klinisch doel wel duidelijk
zijn. De startdatums in het globale episodeontwerp zijn door het model bedacht
en mogen door Python worden gesynchroniseerd met het eerste geplande contact.
Behandel zo'n bijgestelde datum niet als een fout.

Het onderstaande leeftijdskader is door Python berekend. Beoordeel een eerdere
leeftijd nooit tegen de eindleeftijd van het volledige dossier. De leeftijd op
de datum van een gebeurtenis moet uit geboortedatum en gebeurtenisdatum volgen.

Maak duidelijk onderscheid tussen:
- een regulier traject met één verwijsbrief en één latere specialistenbrief; en
- een acute SEH-presentatie of ziekenhuisopname zonder huisartsverwijzing, met
  één latere ontslagbrief en zonder verzonnen verwijsbrief.

Een ontslagbrief mag betrekking hebben op een nevendiagnose die tijdens diezelfde
acute opname is vastgesteld. De episode hoeft dan niet gelijk te zijn aan de
primaire episode van het eerdere acute contact, mits het opnameverband expliciet
is en de ontslagbrief binnen 90 dagen volgt. Keur bijvoorbeeld een CVA-contact
met een latere ontslagbrief voor tijdens de opname ontdekt atriumfibrilleren niet
om die reden af.

Controleer tevens of iedere episode die volgens het globale ontwerp in dit blok
eindigt, rond de einddatum een contact bevat waarin herstel, resultaat,
nacontrole of afsluiting aantoonbaar kan worden uitgewerkt. Een afsluiting die
alleen in de samenvatting staat, is een kritieke fout.

PATIËNTINSTELLINGEN:
{model_naar_json(instellingen)}

GLOBAAL EPISODEONTWERP:
{model_naar_json(blauwdruk.episodeontwerp)}

DETERMINISTISCH LEEFTIJDSKADER:
{leeftijdskader}

BEGINSTATUS:
{vorige_status_tekst}

OPENSTAANDE VERWIJSTRAJECTEN UIT EERDERE BLOKKEN:
{model_naar_json(openstaande_trajectdetails)}

TE BEOORDELEN PLAN:
{model_naar_json(plan)}
""".strip()
        review_resultaat = await Runner.run(reviewer_agent, review_invoer)
        beoordeling = review_resultaat.final_output
        if beoordeling is None:
            feedback = ["de reviewer leverde geen gestructureerde beoordeling op"]
            laatste_structureel_geldige_beoordeling = PlanBeoordeling(
                goedgekeurd=False,
                kritieke_fouten=[],
                waarschuwingen=feedback,
                herstelinstructies=[],
            )
            continue
        # Alleen expliciete kritieke fouten starten binnen het begrensde
        # herstelbudget een nieuwe poging. Na dat budget blijven reviewerpunten
        # zichtbaar, maar een structureel geldig plan veroorzaakt geen crash.
        if not beoordeling.kritieke_fouten:
            beoordeling.goedgekeurd = True
        laatste_beoordeling = beoordeling
        laatste_structureel_geldige_beoordeling = beoordeling
        schrijf_checkpoint(
            runmap
            / f"{blokkader.bloknummer:02d}_planbeoordeling_poging_{poging}.json",
            beoordeling,
        )
        if beoordeling.goedgekeurd and not beoordeling.kritieke_fouten:
            schrijf_checkpoint(
                runmap / f"{blokkader.bloknummer:02d}_periodeplan_goedgekeurd.json",
                plan,
            )
            return plan, beoordeling
        feedback = beoordeling.kritieke_fouten + beoordeling.herstelinstructies

    if laatste_structureel_geldige_plan is not None:
        beoordeling = laatste_structureel_geldige_beoordeling or PlanBeoordeling(
            goedgekeurd=False,
            kritieke_fouten=[],
            waarschuwingen=[
                "De planreview kon niet volledig worden afgerond; het plan "
                "voldeed wel aan alle deterministische Pythoncontroles."
            ],
            herstelinstructies=[],
        )
        resterende_punten = list(
            dict.fromkeys(
                beoordeling.kritieke_fouten
                + beoordeling.waarschuwingen
                + beoordeling.herstelinstructies
            )
        )
        plan = laatste_structureel_geldige_plan
        if resterende_punten:
            plan = plan.model_copy(
                update={
                    "samenvatting_gepland_beloop": (
                        plan.samenvatting_gepland_beloop
                        + "\nOpenstaande automatische reviewpunten voor de "
                        "uitwerking: "
                        + "; ".join(resterende_punten[:10])
                    )
                }
            )
        schrijf_checkpoint(
            runmap
            / f"{blokkader.bloknummer:02d}_periodeplan_structureel_geldig.json",
            plan,
        )
        schrijf_checkpoint(
            runmap
            / f"{blokkader.bloknummer:02d}_planbeoordeling_aandachtspunten.json",
            beoordeling,
        )
        print(
            "    Waarschuwing: de klinische planreview hield aandachtspunten, "
            "maar alle harde Pythoncontroles zijn geslaagd. De pipeline gaat "
            "door en bewaart de reviewpunten voor controle."
        )
        return plan, beoordeling

    details = feedback or (
        laatste_beoordeling.kritieke_fouten
        if laatste_beoordeling is not None
        else ["onbekende beoordelingsfout"]
    )
    raise RuntimeError(
        f"Periodeplan {blokkader.bloknummer} bleef ongeldig:\n- "
        + "\n- ".join(details)
    )


async def genereer_dossierperiode(
    schrijver_agent: Agent,
    reviewer_agent: Agent,
    plan: PeriodePlan,
    blauwdruk: DossierBlauwdruk,
    instellingen: PatientInstellingen,
    vorige_status: DossierStatus | None,
    openstaande_trajectdetails: list[dict[str, str]],
    laatste_blok: bool,
    gebruikte_traject_ids: set[str],
    gebruikte_lab_ids: set[str],
    runmap: Path,
) -> tuple[DossierPeriode, PeriodeBeoordeling]:
    """Werk een goedgekeurd plan uit en herstel uitsluitend het huidige blok."""
    vorige_status_tekst = (
        model_naar_json(vorige_status)
        if vorige_status is not None
        else "Nog geen eerdere dossierstatus: dit is het eerste tijdsblok."
    )
    openstaande_trajecten = (
        vorige_status.openstaande_verwijstrajecten
        if vorige_status is not None
        else []
    )
    patient_context = maak_patient_prompt(instellingen)
    leeftijdskader = maak_leeftijdskader(
        instellingen,
        plan.startdatum,
        plan.einddatum,
    )
    basisinvoer = f"""
Werk uitsluitend het goedgekeurde plan voor tijdsblok {plan.bloknummer} uit tot
volledige Nederlandse huisartsendossiergegevens.

ALGEMENE PATIËNTCONFIGURATIE:
{patient_context}

DETERMINISTISCH LEEFTIJDSKADER VOOR DIT BLOK:
{leeftijdskader}

GLOBAAL EPISODEONTWERP:
{model_naar_json(blauwdruk.episodeontwerp)}

GOEDGEKEURD CONTACTPLAN:
{model_naar_json(plan)}

CUMULATIEVE STATUS AAN HET BEGIN:
{vorige_status_tekst}

Gebruik voor ieder contact exact dezelfde contact-ID, datum, zorgverlener,
contactvorm en episode als in het plan. Schrijf daarna precies één passende S-,
O-, E- en P-regel.

BELANGRIJKE STIJLSCHEIDING: pas de ingestelde ruis alleen toe op deze S-, O-,
E- en P-regels. Schrijf iedere verwijsbrief, specialistenbrief en ontslagbrief
altijd zonder ruis: correcte spelling en interpunctie, professionele volledige
zinnen, geen telegramstijl en geen opzettelijke herhaling of typefouten.

De lijsten medicatie, correspondentie, allergieën, microbiologie en laboratorium
in DossierPeriode bevatten alleen nieuwe registraties of wijzigingen die in dit
blok plaatsvinden. De eindstatus is daarentegen cumulatief: neem daarin alle
episodegegevens, de volledige bijgewerkte medicatiehistorie en de volledige
allergielijst tot en met {plan.einddatum} op. Verwijder nooit historische
medicatieregels; werk status en einddatum bij wanneer een eerder actief middel
in dit blok stopt.

Eerder gebruikte traject-ID's: {sorted(gebruikte_traject_ids)}.
Nog openstaande trajecten die met een specialistenbrief mogen worden afgerond:
{openstaande_trajecten}.
Volledige details van deze openstaande trajecten:
{model_naar_json(openstaande_trajectdetails)}
Eerder gebruikte laboratoriumaanvraag-ID's: {sorted(gebruikte_lab_ids)}.
Hergebruik gesloten traject-ID's of oude laboratoriumaanvraag-ID's niet.

Als het contactplan een specialistenbrief bevat waarmee een traject uit het
overzicht wordt afgerond, neem dan exact het bestaande traject-ID, de episode
en het specialisme uit dat overzicht over. Maak daarvoor geen nieuw traject-ID.

Werk een gepland onderdeel 'Ontslagbrief' uit als één brief van ziekenhuis of
specialist aan de huisarts na een acute opname zonder voorafgaande
huisartsverwijzing. Gebruik daarvoor geen verwijsbrief en geen
specialistenbrief. Een reguliere 'Specialistenbrief' hoort daarentegen altijd
bij een eerdere verwijsbrief met hetzelfde traject-ID.

Als de ontslagbrief is gekoppeld aan een nevendiagnose die tijdens dezelfde
opname is vastgesteld, mag de briefepisode verschillen van de primaire episode
van het eerdere acute contact. Laat het opnameverband dan expliciet terugkomen
in de SOEP-regels en in de ontslagbrief. Een voorbeeld is atriumfibrilleren dat
tijdens een opname wegens een acuut CVA wordt ontdekt.

Zet in de eindstatus toekomstige episodes op 'Nog niet gestart', zonder start-
of einddatum. Zet na het laatste blok geen enkele episode meer op 'Nog niet
gestart'. De peildatum van de eindstatus is exact {plan.einddatum}. Neem in de
eindstatus exact de fictieve geboortedatum uit het leeftijdskader over en vul
'leeftijd_op_peildatum' in met de daar berekende leeftijd op de laatste dag van
dit blok.
""".strip()

    feedback = []
    laatste_beoordeling = None
    laatste_structureel_geldige_periode = None
    laatste_structureel_geldige_beoordeling = None
    for poging in range(1, MAX_HERSTELPOGINGEN + 2):
        resultaat = await Runner.run(
            schrijver_agent,
            basisinvoer + formatteer_feedback(feedback),
        )
        periode = resultaat.final_output
        if periode is None:
            feedback = ["de schrijver leverde geen gestructureerd dossierblok op"]
            continue
        schrijf_checkpoint(
            runmap
            / f"{plan.bloknummer:02d}_dossierperiode_poging_{poging}.json",
            periode,
        )

        feedback = valideer_dossierperiode(
            periode,
            plan,
            instellingen,
            vorige_status,
            laatste_blok,
        )
        if feedback:
            continue
        laatste_structureel_geldige_periode = periode

        review_invoer = f"""
Beoordeel dit geschreven dossierblok kritisch op klinische aannemelijkheid,
chronologie en samenhang tussen SOEP, medicatie, allergieën, laboratorium,
microbiologie en correspondentie. Controleer ook de overgang vanaf de vorige
status. Markeer alleen concrete medische of interne inconsistenties als
kritieke fout. Structurele aantallen en datums zijn al door Python gecontroleerd.

Controleer tevens de stijlscheiding: ingestelde ruis mag alleen in de S-, O-,
E- en P-regels voorkomen. Verwijsbrieven, specialistenbrieven en ontslagbrieven
moeten professioneel, grammaticaal verzorgd en zonder opzettelijke ruis zijn.
Als duidelijke journaalruis in een brief is toegepast, vraag dan concreet om
herstel van die brief.

Een 'Ontslagbrief' na een acute SEH-presentatie of ziekenhuisopname is bewust
een zelfstandig traject zonder voorafgaande huisartsverwijzing. Keur zo'n
traject niet af wegens het ontbreken van een verwijsbrief. Een
'Specialistenbrief' is wel uitsluitend de retourbrief van een regulier
verwijstraject.

De episode van een ontslagbrief mag verschillen van de primaire episode van het
eerdere acute contact als die diagnose tijdens dezelfde opname is vastgesteld.
Het opnameverband moet expliciet zijn en de brief moet binnen 90 dagen volgen.
Een CVA-contact met een ontslagbrief onder tijdens de opname ontdekt
atriumfibrilleren is dus toegestaan.

Het leeftijdskader is deterministisch door Python berekend. De instelling
'leeftijd' geldt alleen op de einddatum van het volledige dossier. Een lagere
leeftijd in een eerder blok is correct. Geboortedatum en leeftijd op de
blokpeildatum zijn bovendien structureel gecontroleerd; maak daarvan geen
kritieke fout als ze overeenkomen met dit kader.

Maak een verschil in redelijke klinische formulering, een verdedigbare
individuele streefwaarde of een stilistische voorkeur hoogstens tot een
waarschuwing. Markeer dit alleen als kritieke fout wanneer het blok zijn eigen
expliciete streefwaarde tegenspreekt, een concreet onveilig beleid bevat of de
medicatie- en SOEP-status elkaar feitelijk tegenspreken. Een voorkeur voor de
formulering van een reeds gestopt middel is bijvoorbeeld geen kritieke fout als
start, stopdatum, medicatielijst en beleid inhoudelijk consistent zijn.

PATIËNTINSTELLINGEN:
{model_naar_json(instellingen)}

DETERMINISTISCH LEEFTIJDSKADER:
{leeftijdskader}

BEGINSTATUS:
{vorige_status_tekst}

PLAN:
{model_naar_json(plan)}

GESCHREVEN DOSSIERBLOK:
{model_naar_json(periode)}
""".strip()
        review_resultaat = await Runner.run(reviewer_agent, review_invoer)
        beoordeling = review_resultaat.final_output
        if beoordeling is None:
            feedback = ["de reviewer leverde geen gestructureerde beoordeling op"]
            laatste_structureel_geldige_beoordeling = PeriodeBeoordeling(
                goedgekeurd=False,
                kritieke_fouten=[],
                waarschuwingen=feedback,
                herstelinstructies=[],
            )
            continue
        # Alleen expliciete kritieke fouten starten binnen het begrensde
        # herstelbudget een nieuwe poging. Na dat budget blijven reviewerpunten
        # zichtbaar, maar een structureel geldig dossierblok veroorzaakt geen
        # crash.
        if not beoordeling.kritieke_fouten:
            beoordeling.goedgekeurd = True
        laatste_beoordeling = beoordeling
        laatste_structureel_geldige_beoordeling = beoordeling
        schrijf_checkpoint(
            runmap
            / f"{plan.bloknummer:02d}_periodebeoordeling_poging_{poging}.json",
            beoordeling,
        )
        if beoordeling.goedgekeurd and not beoordeling.kritieke_fouten:
            schrijf_checkpoint(
                runmap / f"{plan.bloknummer:02d}_dossierperiode_goedgekeurd.json",
                periode,
            )
            return periode, beoordeling
        feedback = beoordeling.kritieke_fouten + beoordeling.herstelinstructies

    if laatste_structureel_geldige_periode is not None:
        beoordeling = (
            laatste_structureel_geldige_beoordeling
            or PeriodeBeoordeling(
                goedgekeurd=False,
                kritieke_fouten=[],
                waarschuwingen=[
                    "De dossierreview kon niet volledig worden afgerond; het "
                    "blok voldeed wel aan alle deterministische Pythoncontroles."
                ],
                herstelinstructies=[],
            )
        )
        periode = laatste_structureel_geldige_periode
        schrijf_checkpoint(
            runmap
            / f"{plan.bloknummer:02d}_dossierperiode_structureel_geldig.json",
            periode,
        )
        schrijf_checkpoint(
            runmap
            / f"{plan.bloknummer:02d}_periodebeoordeling_aandachtspunten.json",
            beoordeling,
        )
        print(
            "    Waarschuwing: de klinische dossierreview hield "
            "aandachtspunten, maar alle harde Pythoncontroles zijn geslaagd. "
            "De pipeline gaat door en bewaart de reviewpunten voor controle."
        )
        return periode, beoordeling

    details = feedback or (
        laatste_beoordeling.kritieke_fouten
        if laatste_beoordeling is not None
        else ["onbekende beoordelingsfout"]
    )
    raise RuntimeError(
        f"Dossierperiode {plan.bloknummer} bleef ongeldig:\n- "
        + "\n- ".join(details)
    )


async def genereer_eindbeoordeling(
    agent: Agent,
    dossier: SynthetischDossier,
    perioden: list[DossierPeriode],
    instellingen: PatientInstellingen,
    runmap: Path,
) -> Eindbeoordeling:
    """Laat een laatste agent alleen de compacte blokovergangen beoordelen."""
    samenvatting = maak_compacte_eindsamenvatting(
        dossier,
        perioden,
        instellingen,
    )
    dossierstart, dossiereinde = bepaal_dossierperiode(
        instellingen.aantal_jaren
    )
    leeftijdskader = maak_leeftijdskader(
        instellingen,
        dossierstart,
        dossiereinde,
    )
    invoer = f"""
Voer een laatste globale kwaliteitscontrole uit op deze compacte samenvatting
van een volledig synthetisch Nederlands huisartsendossier. Beoordeel vooral
langetermijnconsistentie, episodebeloop, medicatiehistorie, openstaande acties en
verwijstrajecten. Alle harde aantallen, episode-enums en datumformaten zijn al
door Python gecontroleerd. Een waarschuwing is geen kritieke fout. Deze controle
vervangt geen beoordeling door een huisarts.

Het onderstaande leeftijdskader is deterministisch door Python berekend. De
ingestelde leeftijd geldt uitsluitend op de einddatum van het volledige
dossier; lagere leeftijden in eerdere periodes zijn dus correct.

DETERMINISTISCH LEEFTIJDSKADER:
{leeftijdskader}

{model_naar_json(samenvatting)}
""".strip()
    resultaat = await Runner.run(agent, invoer)
    beoordeling = resultaat.final_output
    if beoordeling is None:
        beoordeling = Eindbeoordeling(
            goedgekeurd=False,
            kritieke_fouten=[],
            waarschuwingen=[
                "De automatische eindcontrole leverde geen gestructureerde "
                "beoordeling op; voer de handmatige checklist uit."
            ],
            samenvatting=(
                "Het dossier is wel door alle deterministische Pythoncontroles "
                "gekomen, maar kreeg geen volledige automatische eindreview."
            ),
        )
    if not beoordeling.kritieke_fouten:
        beoordeling.goedgekeurd = True
    schrijf_checkpoint(runmap / "99_eindbeoordeling.json", beoordeling)
    return beoordeling


async def main_async() -> None:
    instellingenpad = Path(__file__).with_name(SETTINGS_BESTANDSNAAM)
    instellingen = lees_patientinstellingen(instellingenpad)
    tijdsblokken = maak_tijdsblokken(instellingen)
    (
        blauwdrukmodel,
        periodeplanmodel,
        dossierperiodemodel,
    ) = maak_dynamische_iteratiemodellen(instellingen)

    api_key = getpass("Plak hier je API-key en druk op Enter: ").strip()
    if not api_key:
        raise ValueError("Er is geen API-key ingevoerd.")
    set_default_openai_key(api_key)

    tijdstip = datetime.now().strftime("%Y%m%d_%H%M%S")
    patient_id = veilige_bestandsnaam(instellingen.patient_id)
    runmap = Path(__file__).with_name("iteratieve_runs") / (
        f"{patient_id}_{tijdstip}"
    )
    runmap.mkdir(parents=True, exist_ok=False)

    metadata = {
        "pipeline_versie": PIPELINE_VERSIE,
        "model": MODEL_NAAM,
        "gegenereerd_op": datetime.now().isoformat(timespec="seconds"),
        "patient_id": instellingen.patient_id,
        "settings_bestand": SETTINGS_BESTANDSNAAM,
        "settings_sha256": hashlib.sha256(instellingenpad.read_bytes()).hexdigest(),
        "main_prompt_sha256": hashlib.sha256(
            MAIN_PROMPT.encode("utf-8")
        ).hexdigest(),
        "blokgrootte_jaren": BLOKGROOTTE_JAREN,
        "max_herstelpogingen": MAX_HERSTELPOGINGEN,
        "tijdsblokken": tijdsblokken,
    }
    schrijf_checkpoint(runmap / "00_metadata.json", metadata)
    schrijf_checkpoint(runmap / "00_patientinstellingen.json", instellingen)

    gemeenschappelijke_instructie = (
        MAIN_PROMPT
        + "\n\nGebruik waar relevant Nederlandse eerstelijnsstandaarden en "
        "realistische huisartsenzorg. Structurele correctheid betekent niet "
        "automatisch klinische correctheid; benoem bij beoordelingen alleen "
        "concrete inconsistenties."
    )

    blauwdruk_agent = Agent(
        name="Klinische dossierarchitect",
        instructions=(
            gemeenschappelijke_instructie
            + "\n\nJe taak is uitsluitend het ontwerpen van een globale "
            "longitudinale blauwdruk. Schrijf nog geen volledige SOEP-notities."
        ),
        model=MODEL_NAAM,
        output_type=blauwdrukmodel,
    )
    planner_agent = Agent(
        name="Periodeplanner",
        instructions=(
            gemeenschappelijke_instructie
            + "\n\nJe plant één afgebakend tijdsblok contact voor contact. "
            "Je schrijft nog geen volledige SOEP-tekst."
        ),
        model=MODEL_NAAM,
        output_type=periodeplanmodel,
    )
    plan_reviewer_agent = Agent(
        name="Klinische planreviewer",
        instructions=(
            gemeenschappelijke_instructie
            + "\n\nJe beoordeelt uitsluitend een contactplan. Maak geen nieuw "
            "dossier en verander niets zelf; rapporteer concrete fouten en "
            "gerichte herstelinstructies."
        ),
        model=MODEL_NAAM,
        output_type=PlanBeoordeling,
    )
    schrijver_agent = Agent(
        name="Huisartsdossierschrijver",
        instructions=(
            gemeenschappelijke_instructie
            + "\n\nJe werkt één goedgekeurd tijdsblok volledig uit en geeft "
            "daarna een compacte cumulatieve eindstatus voor het volgende blok."
        ),
        model=MODEL_NAAM,
        output_type=dossierperiodemodel,
    )
    periode_reviewer_agent = Agent(
        name="Klinische dossierreviewer",
        instructions=(
            gemeenschappelijke_instructie
            + "\n\nJe beoordeelt uitsluitend het geschreven tijdsblok en de "
            "overgang vanaf de vorige status. Je schrijft het blok niet zelf."
        ),
        model=MODEL_NAAM,
        output_type=PeriodeBeoordeling,
    )
    eind_reviewer_agent = Agent(
        name="Longitudinale eindreviewer",
        instructions=(
            gemeenschappelijke_instructie
            + "\n\nJe verricht een laatste globale controle op basis van een "
            "compacte samenvatting. Je genereert geen nieuwe dossierinhoud."
        ),
        model=MODEL_NAAM,
        output_type=Eindbeoordeling,
    )

    print("Stap 1/3: globale patiëntblauwdruk maken...")
    blauwdruk = await genereer_blauwdruk(
        blauwdruk_agent,
        instellingen,
        tijdsblokken,
        runmap,
    )

    perioden = []
    vorige_status = None
    eerste_contactnummer = 1
    planbeoordelingen = []
    periodebeoordelingen = []

    print(f"Stap 2/3: {len(blauwdruk.blokken)} tijdsblokken iteratief genereren...")
    for index, blokkader in enumerate(blauwdruk.blokken, start=1):
        openstaande_trajectdetails = bepaal_openstaande_trajectdetails(perioden)
        print(
            f"  Blok {index}/{len(blauwdruk.blokken)} "
            f"({blokkader.startdatum} t/m {blokkader.einddatum}): plannen..."
        )
        plan, planbeoordeling = await genereer_periodeplan(
            planner_agent,
            plan_reviewer_agent,
            blokkader,
            blauwdruk,
            instellingen,
            vorige_status,
            openstaande_trajectdetails,
            eerste_contactnummer,
            runmap,
        )
        planbeoordelingen.append(planbeoordeling)

        gebruikte_traject_ids = {
            brief.traject_id
            for eerdere_periode in perioden
            for brief in eerdere_periode.correspondentie
        }
        gebruikte_lab_ids = {
            uitslag.aanvraag_id
            for eerdere_periode in perioden
            for uitslag in eerdere_periode.laboratorium
        }

        print(
            f"  Blok {index}/{len(blauwdruk.blokken)}: "
            f"{len(plan.contacten)} contacten schrijven en controleren..."
        )
        periode, periodebeoordeling = await genereer_dossierperiode(
            schrijver_agent,
            periode_reviewer_agent,
            plan,
            blauwdruk,
            instellingen,
            vorige_status,
            openstaande_trajectdetails,
            laatste_blok=index == len(blauwdruk.blokken),
            gebruikte_traject_ids=gebruikte_traject_ids,
            gebruikte_lab_ids=gebruikte_lab_ids,
            runmap=runmap,
        )
        perioden.append(periode)
        # Traject-ID's worden door Python beheerd. Dit voorkomt dat het model
        # bijvoorbeeld T001 in een later blok opnieuw gebruikt. De actuele
        # openstaande trajecten worden tegelijk in de eindstatus vastgelegd.
        perioden = normaliseer_correspondentietrajecten(perioden)
        periode = perioden[-1]
        schrijf_checkpoint(
            runmap
            / f"{blokkader.bloknummer:02d}_dossierperiode_trajecten_genormaliseerd.json",
            periode,
        )
        periodebeoordelingen.append(periodebeoordeling)
        vorige_status = periode.eindstatus
        eerste_contactnummer += blokkader.aantal_deelcontacten

    dossier = voeg_dossierperioden_samen(blauwdruk, perioden)
    valideer_iteratief_einddossier(dossier, instellingen)

    print("Stap 3/3: globale eindcontrole uitvoeren...")
    eindbeoordeling = await genereer_eindbeoordeling(
        eind_reviewer_agent,
        dossier,
        perioden,
        instellingen,
        runmap,
    )
    schrijf_checkpoint(
        runmap / "99_reviewoverzicht.json",
        {
            "planbeoordelingen": [
                beoordeling.model_dump(mode="json")
                for beoordeling in planbeoordelingen
            ],
            "periodebeoordelingen": [
                beoordeling.model_dump(mode="json")
                for beoordeling in periodebeoordelingen
            ],
            "eindbeoordeling": eindbeoordeling.model_dump(mode="json"),
        },
    )

    uitvoerpad = Path(__file__).with_name(
        f"synthetisch_epd_{patient_id}_{tijdstip}_iteratief.xlsx"
    )
    dossier_naar_excel(dossier, instellingen, uitvoerpad)

    schrijf_checkpoint(
        runmap / "99_einddossier.json",
        dossier,
    )
    print(f"Klaar: {uitvoerpad.resolve()}")
    print(f"Tussenproducten en beoordelingen: {runmap.resolve()}")
    eerdere_reviewpunten = any(
        not beoordeling.goedgekeurd
        or bool(beoordeling.kritieke_fouten)
        or bool(beoordeling.waarschuwingen)
        for beoordeling in planbeoordelingen + periodebeoordelingen
    )
    if (
        eerdere_reviewpunten
        or not eindbeoordeling.goedgekeurd
        or eindbeoordeling.kritieke_fouten
    ):
        print(
            "LET OP: één of meer reviewer-agents zagen nog aandachtspunten. "
            "Het dossier is wel gegenereerd en voldeed aan de harde "
            "Pythoncontroles. Bekijk "
            f"{(runmap / '99_reviewoverzicht.json').resolve()} vóór de "
            "handmatige checklist."
        )
    elif eindbeoordeling.waarschuwingen:
        print(
            "De automatische eindcontrole is goedgekeurd met waarschuwingen; "
            "bekijk deze vóór de handmatige checklist."
        )


def main() -> None:
    argumenten = sys.argv[1:]
    if argumenten == ["--herstel-laatste-run"]:
        herstel_laatste_run()
        return
    if argumenten:
        raise ValueError(
            "Onbekende opdracht. Gebruik zonder argumenten voor een nieuwe run "
            "of gebruik --herstel-laatste-run om de nieuwste volledige run te "
            "herstellen."
        )
    asyncio.run(main_async())


if __name__ == "__main__":
    main()