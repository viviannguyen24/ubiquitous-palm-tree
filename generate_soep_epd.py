from getpass import getpass
from pathlib import Path
from typing import Literal
from datetime import date, datetime
import calendar
import hashlib
import re

import pandas as pd
from openai import OpenAI
from pydantic import BaseModel, Field, create_model


# Vaste hoofdprompt. Deze blijft voor iedere fictieve patiënt gelijk.
MAIN_PROMPT = """
Je maakt uitsluitend synthetische Nederlandse huisartsendossiers voor onderzoek.
Alle personen en gebeurtenissen moeten fictief zijn. Neem nooit gegevens van een
bestaande patiënt over.

Schrijf beknopt en realistisch in Nederlandse huisartsentaal. Maak een logisch,
chronologisch dossier zonder medische tegenstrijdigheden. Gebruik bij ieder
deelcontact de SOEP-structuur:
- S: klachten, ervaringen en informatie van de patiënt.
- O: objectieve bevindingen, lichamelijk onderzoek en relevante meetwaarden.
- E: beoordeling of diagnose, zo mogelijk met een passende ICPC-code tussen haken.
- P: beleid, medicatie, controles, adviezen en eventuele verwijzing.

Verdeel de contacten logisch over huisarts, POH-S en doktersassistente. Maak alleen
uitslagen, medicatie en verwijzingen die medisch passen bij de fictieve casus.

Genereer een overzicht met persoons- en achtergrondgegevens. Neem daarin een
rookstatus, alcoholgebruik, familieanamnese, woonsituatie, beroep of dagbesteding
en functionele context op. Laat deze gegevens aansluiten op leeftijd, episodes en
zorgbehoefte, maar vermijd stereotiepe of onnodig gedetailleerde invulling. Als
een achtergrondgegeven in de patiëntenprompt handmatig is opgegeven, moet die
waarde exact inhoudelijk worden gevolgd.

Genereer daarnaast een afzonderlijke episode-/probleemlijst. Neem iedere episode
uit de patiëntenprompt precies één keer op en behoud de opgegeven volgorde en
episodenaam. Vermeld per episode een passende ICPC-code, begin- en eventuele
einddatum, status, attentiewaarde, een korte samenvatting van het beloop en het
actuele of afsluitende beleid. Een actieve episode heeft geen einddatum. Een
afgesloten episode heeft een passende einddatum. Een begindatum mag vóór de
dossierperiode liggen als de aandoening aantoonbaar al eerder bestond.

Genereer daarnaast een medicatielijst die volledig aansluit op het journaal.
Neem actuele medicatie en relevante gestopte of eenmalige medicatie uit de
dossierperiode op. Als medicatie wordt gestart, gewijzigd of gestopt, moet dit
logisch terugkomen in de P-regels van het journaal en in de medicatielijst.
Gebruik geen medicatie zonder passende indicatie.

Genereer daarnaast volledige fictieve verwijsbrieven en bijbehorende
specialistenbrieven wanneer een verwijzing klinisch logisch is. Iedere verwijzing
en de bijbehorende specialistenbrief vormen samen een traject en krijgen hetzelfde
traject-ID, bijvoorbeeld T001. De verwijsbrief is altijd van de huisarts aan de
specialist. De specialistenbrief is van de specialist aan de huisarts en heeft
altijd een latere datum dan de bijbehorende verwijsbrief.

Een verwijsbrief bevat in beknopte maar volledige vorm de reden van verwijzing,
relevante voorgeschiedenis, relevante bevindingen en uitslagen, relevante
medicatie en een duidelijke vraagstelling aan de specialist. Een specialistenbrief
bevat in beknopte maar volledige vorm de reden van beoordeling, bevindingen,
eventueel verricht onderzoek, conclusie of diagnose, beleid en vervolgadvies.
Gebruik geen namen, adressen, geboortedata of andere identificerende gegevens in
de brieven.

Laat de correspondentie volledig aansluiten op het journaal en de medicatielijst.
Een verwijzing moet terugkomen in de P-regel van een passend deelcontact. Een
relevant advies, diagnose of medicatiewijziging uit een specialistenbrief moet
logisch terugkomen in een later deelcontact en zo nodig in de medicatielijst.

Volg voor allergieën, microbiologie en laboratorium altijd de gekozen modus uit
de patiëntenprompt: 'Automatisch', 'Zelf invoeren' of 'Geen'. Bij 'Zelf
invoeren' moeten de opgegeven items herkenbaar en volledig in de uitvoer worden
verwerkt. Bij 'Geen' moet de betreffende lijst leeg blijven.

Een allergieoverzicht vermeldt het allergeen, het type allergie, de reactie, de
ernst, de registratiedatum en de status. Als een geneesmiddelenallergie aanwezig
is, mag een geneesmiddel uit die groep nergens in het dossier worden
voorgeschreven.

Een microbiologische uitslag vermeldt datum, episode, materiaal, onderzoek,
uitslag, eventuele verwekker, hoeveelheid of groei,
gevoeligheid/resistentie en conclusie. Laat iedere uitslag logisch aansluiten
op de klachten, het journaal en het antibioticabeleid.

Groepeer laboratoriumbepalingen per realistisch prikmoment. Alle bepalingen uit
dezelfde aanvraag krijgen hetzelfde aanvraag-ID, dezelfde datum en dezelfde
aanvraagreden. Een laboratoriumuitslag vermeldt daarnaast episode, bepaling,
waarde, eenheid, referentiewaarde, afwijking en conclusie. Een aanvraag mag
meerdere klinisch samenhangende bepalingen bevatten en hoeft dus niet beperkt
te blijven tot één episode-specifieke waarde. Neem ook normale bepalingen op als
die bij hetzelfde pakket zijn aangevraagd.

Laat laboratoriumonderzoek aansluiten op de episodes, medicatie, controles,
leeftijd en klachten. Overweeg bij een passende cardiovasculaire controle onder
andere nierfunctie, elektrolyten, glucose, lipiden en urine-albuminurie, afhankelijk
van medicatie, risico en controle-interval. Een beperkt algemeen
laboratoriumonderzoek kan passen bij aspecifieke klachten, maar forceer geen
volledig bloedbeeld of breed pakket zonder indicatie. Een aanvraag of relevante
uitslag moet terugkomen in een passende O- of P-regel. Laat een afwijkende
uitslag die beleid vereist logisch doorwerken in een E- en P-regel en eventueel
in de medicatielijst. Gebruik alleen medisch passende combinaties van waarden,
eenheden en referentiewaarden.

Gebruik datums in de vorm JJJJ-MM-DD. Voeg geen uitleg buiten de gevraagde
gestructureerde uitvoer toe.
""".strip()


SETTINGS_BESTANDSNAAM = "patient_settings_uitgebreid.xlsx"


RuisNiveau = Literal["Geen", "Laag", "Gemiddeld"]


class EpisodeInstelling(BaseModel):
    """Een episode met chronologische volgorde en optionele patiëntspecificatie."""

    volgorde: int
    episode: str
    patient_specifieke_omschrijving: str = ""
    aantal_deelcontacten: int | None = None


class AllergieInstelling(BaseModel):
    """Een door de gebruiker opgegeven allergie uit het instellingenbestand."""

    allergeen: str
    type_allergie: str = "Model bepaalt"
    reactie: str = "Model bepaalt"
    ernst: str = "Model bepaalt"
    status: str = "Actief"


class MicrobiologieInstelling(BaseModel):
    """Een door de gebruiker gewenste microbiologische bepaling of uitslag."""

    episode: str = "Model bepaalt"
    materiaal: str = "Model bepaalt"
    onderzoek: str = "Model bepaalt"
    gewenste_uitslag: str = "Model bepaalt"
    verwekker: str = "Model bepaalt"


class LaboratoriumInstelling(BaseModel):
    """Een door de gebruiker gewenste laboratoriumbepaling."""

    episode: str = "Model bepaalt"
    aanvraagreden: str = "Model bepaalt"
    bepaling: str
    gewenste_waarde: str = "Model bepaalt"
    eenheid: str = "Model bepaalt"


class PatientInstellingen(BaseModel):
    patient_id: str
    geslacht: str
    leeftijd: int
    aantal_jaren: int
    aantal_deelcontacten: int
    ruis_spelfouten: RuisNiveau
    ruis_afkortingen: RuisNiveau
    ruis_telegramstijl: RuisNiveau
    ruis_interpunctie: RuisNiveau
    ruis_herhaling: RuisNiveau
    allergie_modus: Literal["Automatisch", "Zelf invoeren", "Geen"]
    microbiologie_modus: Literal["Automatisch", "Zelf invoeren", "Geen"]
    laboratorium_modus: Literal["Automatisch", "Zelf invoeren", "Geen"]
    rookstatus: str = "Automatisch"
    alcoholgebruik: str = "Automatisch"
    familieanamnese: str = "Automatisch"
    woonsituatie: str = "Automatisch"
    beroep_functionele_context: str = "Automatisch"
    episodes: list[EpisodeInstelling]
    handmatige_allergieen: list[AllergieInstelling]
    handmatige_microbiologie: list[MicrobiologieInstelling]
    handmatig_laboratorium: list[LaboratoriumInstelling]


def waarde_is_leeg(waarde: object) -> bool:
    """Controleer of een cel uit Excel leeg is."""
    return waarde is None or pd.isna(waarde) or str(waarde).strip() == ""


def lees_tekst(waarde: object) -> str:
    """Zet een optionele Excelwaarde veilig om naar tekst."""
    if waarde_is_leeg(waarde):
        return ""
    return str(waarde).strip()


def lees_geheel_getal(waarde: object, naam: str, minimum: int = 0) -> int:
    """Lees en valideer een geheel getal uit het instellingenbestand."""
    if waarde_is_leeg(waarde):
        raise ValueError(f"Instelling '{naam}' is verplicht.")

    try:
        getal = float(waarde)
    except (TypeError, ValueError) as fout:
        raise ValueError(f"Instelling '{naam}' moet een geheel getal zijn.") from fout

    if not getal.is_integer() or getal < minimum:
        raise ValueError(
            f"Instelling '{naam}' moet een geheel getal van minimaal {minimum} zijn."
        )
    return int(getal)


def valideer_episodecontactaantallen(
    episodes: list[EpisodeInstelling],
    totaal_deelcontacten: int,
) -> None:
    """Controleer de optionele verdeling van deelcontacten over episodes."""
    ingevulde_contactaantallen = [
        episode.aantal_deelcontacten
        for episode in episodes
        if episode.aantal_deelcontacten is not None
    ]
    if ingevulde_contactaantallen and len(ingevulde_contactaantallen) != len(episodes):
        raise ValueError(
            "Vul 'aantal_deelcontacten' op het tabblad 'Episodes' óf voor alle "
            "episodes in, óf laat de volledige kolom leeg voor automatische verdeling."
        )
    if ingevulde_contactaantallen:
        totaal_per_episode = sum(ingevulde_contactaantallen)
        if totaal_per_episode != totaal_deelcontacten:
            raise ValueError(
                "De aantallen deelcontacten op het tabblad 'Episodes' tellen op tot "
                f"{totaal_per_episode}, maar op het tabblad 'Patient' staat "
                f"{totaal_deelcontacten}. Maak deze totalen gelijk."
            )


def normaliseer_modus(waarde: object, naam: str) -> str:
    """Normaliseer een gebruikersvriendelijke keuzewaarde uit Excel."""
    invoer = lees_tekst(waarde).lower()
    omzetting = {
        "automatisch": "Automatisch",
        "zelf invoeren": "Zelf invoeren",
        "zelf": "Zelf invoeren",
        "geen": "Geen",
    }
    if invoer not in omzetting:
        raise ValueError(
            f"Instelling '{naam}' moet Automatisch, Zelf invoeren of Geen zijn."
        )
    return omzetting[invoer]


def normaliseer_ruisniveau(waarde: object, naam: str) -> str:
    """Normaliseer een afzonderlijk ruisniveau uit het instellingenbestand."""
    invoer = lees_tekst(waarde).lower()
    omzetting = {
        "geen": "Geen",
        "laag": "Laag",
        "gemiddeld": "Gemiddeld",
        # 'Matig' uit oudere instellingenbestanden wordt als Gemiddeld gelezen.
        "matig": "Gemiddeld",
    }
    if invoer not in omzetting:
        raise ValueError(
            f"Instelling '{naam}' moet Geen, Laag of Gemiddeld zijn."
        )
    return omzetting[invoer]


def lees_patientinstellingen(instellingenpad: Path) -> PatientInstellingen:
    """Lees patiëntkenmerken en episodes uit patient_settings.xlsx."""
    if not instellingenpad.exists():
        raise FileNotFoundError(
            f"Instellingenbestand niet gevonden: {instellingenpad}. "
            f"Plaats {SETTINGS_BESTANDSNAAM} in dezelfde map als dit Python-bestand."
        )

    try:
        patient_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Patient",
            header=2,
        )
        episodes_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Episodes",
        )
    except ValueError as fout:
        raise ValueError(
            "Het instellingenbestand moet de tabbladen 'Patient' en 'Episodes' bevatten."
        ) from fout

    # Deze twee tabbladen zijn alleen verplicht wanneer de bijbehorende modus
    # op 'Zelf invoeren' staat. Anders mogen ze leeg zijn of ontbreken.
    try:
        allergie_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Allergieen",
        )
    except ValueError:
        allergie_dataframe = pd.DataFrame()

    try:
        microbiologie_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Microbiologie",
        )
    except ValueError:
        microbiologie_dataframe = pd.DataFrame()

    try:
        laboratorium_dataframe = pd.read_excel(
            instellingenpad,
            sheet_name="Laboratorium",
        )
    except ValueError:
        laboratorium_dataframe = pd.DataFrame()

    vereiste_patientkolommen = {"Instelling", "Waarde"}
    if not vereiste_patientkolommen.issubset(patient_dataframe.columns):
        raise ValueError(
            "Het tabblad 'Patient' moet de kolommen 'Instelling' en 'Waarde' bevatten."
        )

    waarden = {}
    for _, rij in patient_dataframe.iterrows():
        sleutel = lees_tekst(rij.get("Instelling"))
        if sleutel:
            waarden[sleutel] = rij.get("Waarde")

    verplichte_sleutels = {
        "patient_id",
        "geslacht",
        "leeftijd",
        "aantal_jaren",
        "aantal_deelcontacten",
        "ruis_spelfouten",
        "ruis_afkortingen",
        "ruis_telegramstijl",
        "ruis_interpunctie",
        "ruis_herhaling",
        "allergie_modus",
        "microbiologie_modus",
        "laboratorium_modus",
    }
    ontbrekend = sorted(verplichte_sleutels - waarden.keys())
    if ontbrekend:
        raise ValueError(
            "Deze verplichte instellingen ontbreken op het tabblad 'Patient': "
            + ", ".join(ontbrekend)
        )

    totaal_deelcontacten = lees_geheel_getal(
        waarden["aantal_deelcontacten"],
        "aantal_deelcontacten",
        minimum=1,
    )

    vereiste_episodekolommen = {
        "volgorde",
        "episode",
        "patient_specifieke_omschrijving",
    }
    if not vereiste_episodekolommen.issubset(episodes_dataframe.columns):
        raise ValueError(
            "Het tabblad 'Episodes' moet de kolommen 'volgorde', 'episode' en "
            "'patient_specifieke_omschrijving' bevatten."
        )

    episodes = []
    for _, rij in episodes_dataframe.iterrows():
        episodenaam = lees_tekst(rij.get("episode"))
        if not episodenaam:
            continue
        episodes.append(
            EpisodeInstelling(
                volgorde=lees_geheel_getal(
                    rij.get("volgorde"),
                    "volgorde op het tabblad Episodes",
                    minimum=1,
                ),
                episode=episodenaam,
                patient_specifieke_omschrijving=lees_tekst(
                    rij.get("patient_specifieke_omschrijving")
                ),
                aantal_deelcontacten=(
                    None
                    if waarde_is_leeg(rij.get("aantal_deelcontacten"))
                    else lees_geheel_getal(
                        rij.get("aantal_deelcontacten"),
                        "aantal_deelcontacten op het tabblad Episodes",
                        minimum=1,
                    )
                ),
            )
        )

    if not episodes:
        raise ValueError("Vul minimaal één episode in op het tabblad 'Episodes'.")

    volgordes = [episode.volgorde for episode in episodes]
    if len(volgordes) != len(set(volgordes)):
        raise ValueError(
            "Iedere episode moet op het tabblad 'Episodes' een unieke volgorde hebben."
        )
    episodes.sort(key=lambda episode: episode.volgorde)

    valideer_episodecontactaantallen(episodes, totaal_deelcontacten)

    patient_id = lees_tekst(waarden["patient_id"])
    geslacht = lees_tekst(waarden["geslacht"])
    if not patient_id:
        raise ValueError("Instelling 'patient_id' mag niet leeg zijn.")
    if not geslacht:
        raise ValueError("Instelling 'geslacht' mag niet leeg zijn.")

    allergie_modus = normaliseer_modus(
        waarden["allergie_modus"],
        "allergie_modus",
    )
    microbiologie_modus = normaliseer_modus(
        waarden["microbiologie_modus"],
        "microbiologie_modus",
    )
    laboratorium_modus = normaliseer_modus(
        waarden["laboratorium_modus"],
        "laboratorium_modus",
    )

    handmatige_allergieen = []
    if allergie_modus == "Zelf invoeren":
        vereiste_kolom = {"allergeen"}
        if not vereiste_kolom.issubset(allergie_dataframe.columns):
            raise ValueError(
                "Bij allergie_modus 'Zelf invoeren' moet het tabblad "
                "'Allergieen' minimaal de kolom 'allergeen' bevatten."
            )

        for _, rij in allergie_dataframe.iterrows():
            allergeen = lees_tekst(rij.get("allergeen"))
            if not allergeen:
                continue
            handmatige_allergieen.append(
                AllergieInstelling(
                    allergeen=allergeen,
                    type_allergie=lees_tekst(rij.get("type_allergie"))
                    or "Model bepaalt",
                    reactie=lees_tekst(rij.get("reactie")) or "Model bepaalt",
                    ernst=lees_tekst(rij.get("ernst")) or "Model bepaalt",
                    status=lees_tekst(rij.get("status")) or "Actief",
                )
            )

        if not handmatige_allergieen:
            raise ValueError(
                "allergie_modus staat op 'Zelf invoeren', maar op het tabblad "
                "'Allergieen' is geen allergeen ingevuld."
            )

    handmatige_microbiologie = []
    if microbiologie_modus == "Zelf invoeren":
        mogelijke_kolommen = {
            "episode",
            "materiaal",
            "onderzoek",
            "gewenste_uitslag",
            "verwekker",
        }
        if microbiologie_dataframe.empty or not mogelijke_kolommen.intersection(
            microbiologie_dataframe.columns
        ):
            raise ValueError(
                "Bij microbiologie_modus 'Zelf invoeren' moet het tabblad "
                "'Microbiologie' de daarvoor bestemde kolommen bevatten."
            )

        for _, rij in microbiologie_dataframe.iterrows():
            waarden_rij = {
                kolom: lees_tekst(rij.get(kolom)) for kolom in mogelijke_kolommen
            }
            if not any(waarden_rij.values()):
                continue
            handmatige_microbiologie.append(
                MicrobiologieInstelling(
                    episode=waarden_rij["episode"] or "Model bepaalt",
                    materiaal=waarden_rij["materiaal"] or "Model bepaalt",
                    onderzoek=waarden_rij["onderzoek"] or "Model bepaalt",
                    gewenste_uitslag=waarden_rij["gewenste_uitslag"]
                    or "Model bepaalt",
                    verwekker=waarden_rij["verwekker"] or "Model bepaalt",
                )
            )

        if not handmatige_microbiologie:
            raise ValueError(
                "microbiologie_modus staat op 'Zelf invoeren', maar het tabblad "
                "'Microbiologie' bevat geen ingevulde regel."
            )

    handmatig_laboratorium = []
    if laboratorium_modus == "Zelf invoeren":
        if "bepaling" not in laboratorium_dataframe.columns:
            raise ValueError(
                "Bij laboratorium_modus 'Zelf invoeren' moet het tabblad "
                "'Laboratorium' minimaal de kolom 'bepaling' bevatten."
            )

        for _, rij in laboratorium_dataframe.iterrows():
            bepaling = lees_tekst(rij.get("bepaling"))
            if not bepaling:
                continue
            handmatig_laboratorium.append(
                LaboratoriumInstelling(
                    episode=lees_tekst(rij.get("episode")) or "Model bepaalt",
                    aanvraagreden=lees_tekst(rij.get("aanvraagreden"))
                    or "Model bepaalt",
                    bepaling=bepaling,
                    gewenste_waarde=lees_tekst(rij.get("gewenste_waarde"))
                    or "Model bepaalt",
                    eenheid=lees_tekst(rij.get("eenheid")) or "Model bepaalt",
                )
            )

        if not handmatig_laboratorium:
            raise ValueError(
                "laboratorium_modus staat op 'Zelf invoeren', maar op het tabblad "
                "'Laboratorium' is geen bepaling ingevuld."
            )

    return PatientInstellingen(
        patient_id=patient_id,
        geslacht=geslacht,
        leeftijd=lees_geheel_getal(
            waarden["leeftijd"],
            "leeftijd",
            minimum=0,
        ),
        aantal_jaren=lees_geheel_getal(
            waarden["aantal_jaren"],
            "aantal_jaren",
            minimum=1,
        ),
        aantal_deelcontacten=totaal_deelcontacten,
        ruis_spelfouten=normaliseer_ruisniveau(
            waarden["ruis_spelfouten"],
            "ruis_spelfouten",
        ),
        ruis_afkortingen=normaliseer_ruisniveau(
            waarden["ruis_afkortingen"],
            "ruis_afkortingen",
        ),
        ruis_telegramstijl=normaliseer_ruisniveau(
            waarden["ruis_telegramstijl"],
            "ruis_telegramstijl",
        ),
        ruis_interpunctie=normaliseer_ruisniveau(
            waarden["ruis_interpunctie"],
            "ruis_interpunctie",
        ),
        ruis_herhaling=normaliseer_ruisniveau(
            waarden["ruis_herhaling"],
            "ruis_herhaling",
        ),
        allergie_modus=allergie_modus,
        microbiologie_modus=microbiologie_modus,
        laboratorium_modus=laboratorium_modus,
        rookstatus=lees_tekst(waarden.get("rookstatus")) or "Automatisch",
        alcoholgebruik=lees_tekst(waarden.get("alcoholgebruik")) or "Automatisch",
        familieanamnese=(
            lees_tekst(waarden.get("familieanamnese")) or "Automatisch"
        ),
        woonsituatie=lees_tekst(waarden.get("woonsituatie")) or "Automatisch",
        beroep_functionele_context=(
            lees_tekst(waarden.get("beroep_functionele_context"))
            or "Automatisch"
        ),
        episodes=episodes,
        handmatige_allergieen=handmatige_allergieen,
        handmatige_microbiologie=handmatige_microbiologie,
        handmatig_laboratorium=handmatig_laboratorium,
    )


def bepaal_dossierperiode(aantal_jaren: int) -> tuple[str, str]:
    """Bereken automatisch een periode die eindigt op de uitvoerdatum."""
    einddatum = date.today()
    try:
        startdatum = einddatum.replace(year=einddatum.year - aantal_jaren)
    except ValueError:
        # Alleen relevant als het script op 29 februari wordt uitgevoerd.
        startdatum = einddatum.replace(
            year=einddatum.year - aantal_jaren,
            day=28,
        )
    return startdatum.isoformat(), einddatum.isoformat()


def bepaal_fictieve_geboortedatum(
    patient_id: str,
    leeftijd: int,
    einddatum: str,
) -> str:
    """Maak reproduceerbaar een fictieve geboortedatum die exact bij de leeftijd past."""
    einde = date.fromisoformat(einddatum)
    digest = hashlib.sha256(patient_id.encode("utf-8")).digest()
    maand = digest[0] % 12 + 1
    maximaal_aantal_dagen = calendar.monthrange(2000, maand)[1]
    dag = digest[1] % maximaal_aantal_dagen + 1

    geboortejaar = einde.year - leeftijd
    if (maand, dag) > (einde.month, einde.day):
        geboortejaar -= 1
    if maand == 2 and dag == 29 and not calendar.isleap(geboortejaar):
        dag = 28

    return date(geboortejaar, maand, dag).isoformat()


def maak_achtergrondregel(label: str, waarde: str) -> str:
    """Maak een promptregel voor een automatisch of handmatig achtergrondgegeven."""
    if waarde.strip().lower() == "automatisch":
        return (
            f"- {label}: bepaal automatisch een realistische waarde die past bij "
            "de patiënt en het klinische verloop"
        )
    return f"- {label}: gebruik verplicht de volgende specificatie: {waarde}"


def maak_patient_prompt(instellingen: PatientInstellingen) -> str:
    """Bouw de variabele patiëntenprompt uit het Excel-instellingenbestand."""
    startdatum, einddatum = bepaal_dossierperiode(instellingen.aantal_jaren)
    geboortedatum = bepaal_fictieve_geboortedatum(
        instellingen.patient_id,
        instellingen.leeftijd,
        einddatum,
    )
    episode_regels = []
    for episode in instellingen.episodes:
        regel = f"{episode.volgorde}. {episode.episode}"
        if episode.patient_specifieke_omschrijving:
            regel += (
                " | patiëntspecifieke omschrijving: "
                f"{episode.patient_specifieke_omschrijving}"
            )
        if episode.aantal_deelcontacten is not None:
            regel += (
                " | exact aantal deelcontacten: "
                f"{episode.aantal_deelcontacten}"
            )
        episode_regels.append(regel)

    contactaantallen_ingevuld = all(
        episode.aantal_deelcontacten is not None
        for episode in instellingen.episodes
    )
    if contactaantallen_ingevuld:
        contactverdeling_instructie = (
            "Gebruik voor iedere episode exact het opgegeven aantal deelcontacten. "
            f"Deze aantallen tellen samen op tot {instellingen.aantal_deelcontacten}. "
            "Koppel ieder deelcontact in het veld 'episode' aan precies één primaire "
            "episode uit de opgegeven lijst. Een contact waarin meerdere problemen "
            "worden besproken telt slechts eenmaal, bij de primaire episode. Maak in "
            f"totaal exact {instellingen.aantal_deelcontacten} deelcontacten."
        )
    else:
        contactverdeling_instructie = (
            "Er zijn geen aantallen per episode opgegeven. Verdeel daarom het totale "
            "aantal deelcontacten zelf medisch logisch over de episodes. Maak in "
            f"totaal exact {instellingen.aantal_deelcontacten} deelcontacten."
        )

    ruisinstructies = {
        "spelfouten": {
            "Geen": "Gebruik geen opzettelijke spelfouten of typefouten.",
            "Laag": (
                "Gebruik incidenteel een kleine, realistische type- of spelfout, "
                "terwijl ieder woord goed te reconstrueren blijft."
            ),
            "Gemiddeld": (
                "Gebruik verspreid door het dossier regelmatig kleine, realistische "
                "type- of spelfouten, zonder medische termen onherkenbaar te maken."
            ),
        },
        "afkortingen": {
            "Geen": "Schrijf termen in de vrije tekst zo veel mogelijk voluit.",
            "Laag": (
                "Gebruik incidenteel een gangbare Nederlandse medische afkorting."
            ),
            "Gemiddeld": (
                "Gebruik geregeld gangbare Nederlandse huisartsafkortingen, maar "
                "zorg dat de betekenis uit de context duidelijk blijft."
            ),
        },
        "telegramstijl": {
            "Geen": "Gebruik grammaticaal volledige zinnen.",
            "Laag": (
                "Gebruik incidenteel korte zinsdelen of weggelaten lidwoorden zoals "
                "in een realistische huisartsnotitie."
            ),
            "Gemiddeld": (
                "Gebruik geregeld beknopte telegramstijl en korte zinsdelen, zonder "
                "klinisch relevante informatie weg te laten."
            ),
        },
        "interpunctie en hoofdletters": {
            "Geen": "Gebruik consequente interpunctie en hoofdletters.",
            "Laag": (
                "Laat incidenteel een punt of hoofdletter weg in vrije tekst."
            ),
            "Gemiddeld": (
                "Gebruik geregeld licht inconsistente interpunctie of hoofdletters, "
                "maar behoud de leesbaarheid."
            ),
        },
        "herhaling": {
            "Geen": "Vermijd onnodige herhaling van eerder vastgelegde informatie.",
            "Laag": (
                "Herhaal incidenteel een korte relevante voorgeschiedenis of een "
                "eerder beleidspunt in een later deelcontact."
            ),
            "Gemiddeld": (
                "Laat regelmatig beperkte, realistische herhaling van relevante "
                "voorgeschiedenis of beleid voorkomen, zonder volledige notities te "
                "dupliceren."
            ),
        },
    }

    gekozen_ruis = [
        ("Spelfouten", instellingen.ruis_spelfouten, "spelfouten"),
        ("Afkortingen", instellingen.ruis_afkortingen, "afkortingen"),
        ("Telegramstijl", instellingen.ruis_telegramstijl, "telegramstijl"),
        (
            "Interpunctie en hoofdletters",
            instellingen.ruis_interpunctie,
            "interpunctie en hoofdletters",
        ),
        ("Herhaling", instellingen.ruis_herhaling, "herhaling"),
    ]
    ruis_regels = []
    for label, niveau, sleutel in gekozen_ruis:
        ruis_regels.append(
            f"- {label} — {niveau}: {ruisinstructies[sleutel][niveau]}"
        )

    achtergrond_regels = [
        maak_achtergrondregel("rookstatus", instellingen.rookstatus),
        maak_achtergrondregel("alcoholgebruik", instellingen.alcoholgebruik),
        maak_achtergrondregel("familieanamnese", instellingen.familieanamnese),
        maak_achtergrondregel("woonsituatie", instellingen.woonsituatie),
        maak_achtergrondregel(
            "beroep, dagbesteding en functionele context",
            instellingen.beroep_functionele_context,
        ),
    ]

    if instellingen.allergie_modus == "Automatisch":
        allergie_instructie = (
            "Bepaal zelf of een klinisch relevante allergie bij de casus past. "
            "Het allergieoverzicht mag leeg zijn. Als je een allergie opneemt, "
            "laat deze dan consistent terugkomen in het voorschrijfbeleid en de "
            "medicatielijst."
        )
    elif instellingen.allergie_modus == "Geen":
        allergie_instructie = (
            "Neem geen allergieën op. Geef voor allergieen een lege lijst terug."
        )
    else:
        allergieregels = []
        for nummer, allergie in enumerate(
            instellingen.handmatige_allergieen,
            start=1,
        ):
            allergieregels.append(
                f"{nummer}. allergeen={allergie.allergeen}; "
                f"type={allergie.type_allergie}; reactie={allergie.reactie}; "
                f"ernst={allergie.ernst}; status={allergie.status}"
            )
        allergie_instructie = (
            "Neem de onderstaande allergieën verplicht op en voeg geen andere "
            "allergieën toe. Kies alleen voor velden met 'Model bepaalt' zelf "
            "een passende waarde. Kies een passende registratiedatum binnen de "
            "dossierperiode en stem medicatie en beleid hierop af.\n"
            + "\n".join(allergieregels)
        )

    if instellingen.microbiologie_modus == "Automatisch":
        microbiologie_instructie = (
            "Genereer alleen microbiologische uitslagen als de episodes of het "
            "klinische beloop daar aanleiding toe geven. De microbiologielijst "
            "mag leeg zijn."
        )
    elif instellingen.microbiologie_modus == "Geen":
        microbiologie_instructie = (
            "Neem geen microbiologische uitslagen op. Geef voor microbiologie "
            "een lege lijst terug."
        )
    else:
        microbiologieregels = []
        for nummer, bepaling in enumerate(
            instellingen.handmatige_microbiologie,
            start=1,
        ):
            microbiologieregels.append(
                f"{nummer}. episode={bepaling.episode}; materiaal={bepaling.materiaal}; "
                f"onderzoek={bepaling.onderzoek}; "
                f"gewenste uitslag={bepaling.gewenste_uitslag}; "
                f"verwekker={bepaling.verwekker}"
            )
        microbiologie_instructie = (
            "Verwerk de onderstaande microbiologische bepalingen verplicht en "
            "voeg geen andere bepalingen toe. Kies alleen voor velden met "
            "'Model bepaalt' zelf een passende waarde. Kies passende datums, "
            "hoeveelheid/groei en gevoeligheid/resistentie en laat alles "
            "aansluiten op journaal en behandelbeleid.\n"
            + "\n".join(microbiologieregels)
        )

    if instellingen.laboratorium_modus == "Automatisch":
        laboratorium_instructie = (
            "Genereer laboratoriumuitslagen wanneer deze passen bij diagnostiek, "
            "medicatiebewaking of controles van de opgegeven episodes. Maak van "
            "ieder prikmoment een herkenbare aanvraag met een aanvraag-ID zoals "
            "L001. Gebruik voor alle bepalingen uit hetzelfde prikmoment dezelfde "
            "datum en aanvraagreden. Genereer per aanvraag een realistisch pakket "
            "van meerdere samenhangende bepalingen in plaats van uitsluitend één "
            "episode-specifieke waarde. Neem ook normale nevenbepalingen op die bij "
            "hetzelfde pakket plausibel zijn aangevraagd. Combineer controles voor "
            "meerdere aandoeningen zo nodig in één prikmoment. Laat bij chronische "
            "aandoeningen passende meetmomenten terugkomen, zonder onnodige "
            "bepalingen of onrealistisch veel uitslagen toe te voegen. Bij alleen "
            "hypertensie is een volledig bloedbeeld niet standaard verplicht; "
            "kies het pakket op basis van cardiovasculair risico, medicatie en "
            "eventuele bijkomende klachten. De laboratoriumlijst mag leeg zijn als "
            "geen enkele episode daar aanleiding toe geeft."
        )
    elif instellingen.laboratorium_modus == "Geen":
        laboratorium_instructie = (
            "Neem geen laboratoriumuitslagen op. Geef voor laboratorium een lege "
            "lijst terug."
        )
    else:
        laboratoriumregels = []
        for nummer, bepaling in enumerate(
            instellingen.handmatig_laboratorium,
            start=1,
        ):
            laboratoriumregels.append(
                f"{nummer}. episode={bepaling.episode}; "
                f"aanvraagreden={bepaling.aanvraagreden}; "
                f"bepaling={bepaling.bepaling}; "
                f"gewenste waarde={bepaling.gewenste_waarde}; "
                f"eenheid={bepaling.eenheid}"
            )
        laboratorium_instructie = (
            "Verwerk de onderstaande laboratoriumbepalingen verplicht en voeg "
            "geen andere bepalingen toe. Kies alleen voor velden met 'Model "
            "bepaalt' zelf een passende waarde. Groepeer bepalingen die medisch bij "
            "hetzelfde prikmoment passen onder één aanvraag-ID, datum en "
            "aanvraagreden. Kies passende datums en referentiewaarden en laat "
            "aanvragen, uitslagen en eventueel beleid aansluiten op het journaal.\n"
            + "\n".join(laboratoriumregels)
        )

    return f"""
Maak een synthetisch huisartsendossier voor:

- fictieve patiënt-ID: {instellingen.patient_id}
- geslacht: {instellingen.geslacht}
- fictieve geboortedatum: {geboortedatum}
- leeftijd aan het einde van de dossierperiode: {instellingen.leeftijd} jaar
- duur van het dossier: {instellingen.aantal_jaren} jaar
- automatisch berekende periode: {startdatum} tot en met {einddatum}
- gewenst aantal deelcontacten: {instellingen.aantal_deelcontacten}

Episodes:
{chr(10).join(episode_regels)}

Persoons- en achtergrondgegevens:
{chr(10).join(achtergrond_regels)}

Neem de patiënt-ID, geboortedatum, leeftijd en het formele geslacht exact over.
Genereer voor iedere opgegeven episode precies één regel in de episodelijst.
Behoud daarin exact de opgegeven volgorde en episodenaam. Kies een passende
ICPC-code, begin- en eventuele einddatum, status, attentiewaarde, samenvatting
en beleid die volledig aansluiten op het journaal.

De kolom 'volgorde' uit het instellingenbestand bepaalt de chronologische
volgorde waarin de episodes voor het eerst optreden of voor het eerst in het
dossier worden geregistreerd. Respecteer deze volgorde. Een patiëntspecifieke
omschrijving is een bindende specificatie van de betreffende episode en moet
consistent terugkomen in relevante SOEP-regels, bevindingen, medicatie,
correspondentie en uitslagen.

Verdeling van deelcontacten:
{contactverdeling_instructie}

Genereer een realistisch longitudinaal huisartsendossier over de volledige
automatisch berekende periode. Verdeel het gewenste aantal deelcontacten
chronologisch en medisch logisch. Het aantal contacten hoeft niet gelijk over
alle jaren verdeeld te zijn; bij een klein testaantal mogen sommige jaren
weinig contacten bevatten. Laat ontwikkelingen uit eerdere contacten logisch
terugkomen in latere contacten. Gebruik per deelcontact precies één S-, O-, E-
en P-regel.

BELANGRIJK: koppel ieder deelcontact aan precies één primaire episode. Vul in
het veld 'episode' letterlijk en ongewijzigd precies één episodenaam uit de
hierboven opgegeven episodelijst in. Combineer in dit veld nooit meerdere
episodenamen met een komma, puntkomma, schuine streep of het woord 'en'. Als
meerdere gezondheidsproblemen tijdens één werkelijk contact worden besproken,
registreer die als afzonderlijke deelcontacten, ieder gekoppeld aan één primaire
episode. Houd daarbij het totale gewenste aantal deelcontacten gelijk.

Bepaal zelf het realistische beloop van iedere opgegeven episode. Een episode
mag bijvoorbeeld chronisch actief, eenmalig of later afgesloten zijn. Laat
chronische episodes terugkomen in passende controles en voeg enkele passende
tussentijdse contacten toe.

Maak een medicatielijst die volledig aansluit op het journaal. Geef bij actieve
medicatie een lege einddatum en bij gestopte of eenmalige medicatie een passende
einddatum. Laat iedere start, wijziging of stop ook terugkomen in het journaal.

Genereer alleen verwijstrajecten als deze klinisch logisch volgen uit de
opgegeven episodes en het verloop van de klachten. Forceer geen vast aantal.
Maak bij iedere verwijzing één verwijsbrief en één latere specialistenbrief met
hetzelfde traject-ID. De lijst met correspondentie mag leeg zijn als geen
verwijzing nodig is.

Allergieën — gekozen modus: {instellingen.allergie_modus}
{allergie_instructie}

Microbiologie — gekozen modus: {instellingen.microbiologie_modus}
{microbiologie_instructie}

Laboratorium — gekozen modus: {instellingen.laboratorium_modus}
{laboratorium_instructie}

Stijl en ruis:
{chr(10).join(ruis_regels)}

Pas ruis uitsluitend toe op vrije tekst. Verander door ruis nooit patiënt-ID's,
datums, geneesmiddelnamen, doseringen, meetwaarden, eenheden, uitslagen,
diagnoses of traject-ID's. Ruis mag geen medische onjuistheden,
tegenstrijdigheden of klinische onduidelijkheid veroorzaken.

Zorg dat patiëntgegevens, episodelijst, journaal, medicatie, correspondentie,
allergieën, microbiologie en laboratorium onderling consistent zijn.
""".strip()


def veilige_bestandsnaam(waarde: str) -> str:
    """Maak van de fictieve patiënt-ID een veilige bestandsnaam."""
    veilig = re.sub(r"[^A-Za-z0-9_-]+", "_", waarde.strip())
    return veilig.strip("_") or "patient"


class PatientAchtergrond(BaseModel):
    rookstatus: str = Field(
        description="Rookstatus, bijvoorbeeld nooit gerookt, voormalig roker of roker"
    )
    alcoholgebruik: str = Field(
        description="Beknopte omschrijving van alcoholgebruik in eenheden per week"
    )
    familieanamnese: str = Field(
        description="Klinisch relevante familieanamnese of expliciet geen bijzonderheden"
    )
    woonsituatie: str = Field(
        description="Woonsituatie en relevante sociale ondersteuning"
    )
    beroep_dagbesteding: str = Field(
        description="Beroep, pensioenstatus of relevante dagelijkse bezigheden"
    )
    functionele_context: str = Field(
        description="Beknopte functionele context, bijvoorbeeld mobiliteit en zelfstandigheid"
    )


class EpisodeOverzichtRegel(BaseModel):
    volgorde: int = Field(description="Volgorde uit de opgegeven episodelijst")
    episode: str = Field(description="Exacte episodenaam uit de patiëntenprompt")
    icpc_code: str = Field(description="Passende ICPC-code, bijvoorbeeld T90.02")
    startdatum: str = Field(description="Begindatum in JJJJ-MM-DD")
    einddatum: str = Field(
        description="Einddatum in JJJJ-MM-DD; leeg laten als de episode actief is"
    )
    status: Literal["Actief", "Afgesloten"]
    attentiewaarde: Literal["Ja", "Nee"]
    samenvatting_beloop: str = Field(
        description="Beknopte samenvatting van het relevante klinische beloop"
    )
    beleid: str = Field(description="Actueel of afsluitend beleid voor deze episode")


class Deelcontact(BaseModel):
    contact_id: str = Field(description="Uniek contactnummer, bijvoorbeeld C001")
    datum: str = Field(description="Datum in JJJJ-MM-DD")
    zorgverlener: Literal["Huisarts", "POH-S", "Doktersassistente"]
    contactvorm: Literal[
        "Praktijkconsult", "Telefonisch", "Huisbezoek", "Administratief"
    ]
    episode: str
    s: str = Field(description="Subjectieve SOEP-regel")
    o: str = Field(description="Objectieve SOEP-regel")
    e: str = Field(description="Evaluatie met passende diagnose en eventueel ICPC-code")
    p: str = Field(description="Plan en beleid")


class MedicatieRegel(BaseModel):
    geneesmiddel: str = Field(description="Naam van het geneesmiddel")
    sterkte: str = Field(description="Sterkte, bijvoorbeeld 500 mg")
    dosering: str = Field(description="Dosering en frequentie, bijvoorbeeld 2dd1")
    indicatie: str = Field(description="Episode of reden waarvoor het middel wordt gebruikt")
    startdatum: str = Field(description="Startdatum in JJJJ-MM-DD")
    einddatum: str = Field(
        description="Einddatum in JJJJ-MM-DD; leeg laten als de medicatie actief is"
    )
    status: Literal["Actief", "Gestopt", "Eenmalig"]


class CorrespondentieRegel(BaseModel):
    traject_id: str = Field(
        description=(
            "Gedeeld trajectnummer voor een verwijsbrief en de bijbehorende "
            "specialistenbrief, bijvoorbeeld T001"
        )
    )
    datum: str = Field(description="Datum van de brief in JJJJ-MM-DD")
    type_brief: Literal["Verwijsbrief", "Specialistenbrief"]
    specialisme: str = Field(description="Betrokken specialisme, bijvoorbeeld Urologie")
    episode: str = Field(description="Episode waarop de correspondentie betrekking heeft")
    van: str = Field(description="Afzender, bijvoorbeeld Huisarts of Uroloog")
    aan: str = Field(description="Ontvanger, bijvoorbeeld Uroloog of Huisarts")
    onderwerp: str = Field(description="Beknopt onderwerp van de brief")
    inhoud: str = Field(
        description="Volledige fictieve brieftekst; niet alleen een samenvatting"
    )


class AllergieRegel(BaseModel):
    allergeen: str = Field(description="Stof of geneesmiddel waarvoor de allergie bestaat")
    type_allergie: Literal["Geneesmiddel", "Voedsel", "Omgeving", "Overig"]
    reactie: str = Field(description="Klinische reactie, bijvoorbeeld huiduitslag")
    ernst: Literal["Mild", "Matig", "Ernstig", "Onbekend"]
    registratiedatum: str = Field(description="Registratiedatum in JJJJ-MM-DD")
    status: Literal["Actief", "Inactief"]


class MicrobiologieRegel(BaseModel):
    datum: str = Field(description="Datum van de microbiologische uitslag in JJJJ-MM-DD")
    episode: str = Field(description="Episode waarop de uitslag betrekking heeft")
    materiaal: str = Field(description="Onderzocht materiaal, bijvoorbeeld urine")
    onderzoek: str = Field(description="Type onderzoek, bijvoorbeeld urinekweek")
    uitslag: str = Field(description="Hoofduitslag, bijvoorbeeld positief of negatief")
    verwekker: str = Field(
        description="Aangetoonde verwekker; leeg laten wanneer geen verwekker is gevonden"
    )
    hoeveelheid: str = Field(
        description="Hoeveelheid of groei indien relevant; anders leeg laten"
    )
    gevoeligheid_resistentie: str = Field(
        description="Beknopt en klinisch relevant gevoeligheids- of resistentiepatroon"
    )
    conclusie: str = Field(description="Beknopte klinische conclusie van de uitslag")


class LaboratoriumRegel(BaseModel):
    aanvraag_id: str = Field(
        description=(
            "Gedeeld nummer voor alle bepalingen uit hetzelfde prikmoment, "
            "bijvoorbeeld L001"
        )
    )
    datum: str = Field(description="Datum van de laboratoriumuitslag in JJJJ-MM-DD")
    aanvraagreden: str = Field(
        description=(
            "Reden of type aanvraag, bijvoorbeeld jaarlijkse CVRM-controle of "
            "algemeen onderzoek wegens vermoeidheid"
        )
    )
    episode: str = Field(description="Episode waarop de uitslag betrekking heeft")
    bepaling: str = Field(description="Naam van de bepaling, bijvoorbeeld HbA1c")
    waarde: str = Field(description="Gemeten waarde zonder eenheid")
    eenheid: str = Field(description="Bijpassende eenheid, bijvoorbeeld mmol/mol")
    referentiewaarde: str = Field(
        description="Passende referentiewaarde of streefwaarde als tekst"
    )
    afwijking: Literal["Laag", "Normaal", "Hoog", "Niet van toepassing"]
    conclusie: str = Field(description="Beknopte klinische duiding van de uitslag")


class SynthetischDossier(BaseModel):
    patient_achtergrond: PatientAchtergrond
    episodelijst: list[EpisodeOverzichtRegel]
    contacten: list[Deelcontact]
    medicatie: list[MedicatieRegel]
    correspondentie: list[CorrespondentieRegel]
    allergieen: list[AllergieRegel]
    microbiologie: list[MicrobiologieRegel]
    laboratorium: list[LaboratoriumRegel]


def maak_dynamisch_dossiermodel(
    instellingen: PatientInstellingen,
) -> type[SynthetischDossier]:
    """Beperk alle episodevelden tot de exacte namen uit het settingsbestand.

    De vaste Pydantic-modellen beschrijven de algemene uitvoerstructuur. Dit
    dynamische model voegt voor deze patiënt een enum toe met uitsluitend de
    toegestane episodenamen. Daardoor kan het taalmodel in een episodeveld geen
    samengestelde of anders geformuleerde episodenaam meer teruggeven.
    """
    episodenamen = tuple(episode.episode for episode in instellingen.episodes)
    episode_keuze = Literal[episodenamen]

    dynamisch_deelcontact = create_model(
        "DeelcontactMetVasteEpisode",
        __base__=Deelcontact,
        episode=(
            episode_keuze,
            Field(description="Exact één episodenaam uit het instellingenbestand"),
        ),
    )
    dynamische_episodeoverzichtregel = create_model(
        "EpisodeOverzichtRegelMetVasteEpisode",
        __base__=EpisodeOverzichtRegel,
        episode=(
            episode_keuze,
            Field(description="Exacte episodenaam uit het instellingenbestand"),
        ),
    )
    dynamische_correspondentieregel = create_model(
        "CorrespondentieRegelMetVasteEpisode",
        __base__=CorrespondentieRegel,
        episode=(
            episode_keuze,
            Field(description="Exacte gekoppelde episode uit het instellingenbestand"),
        ),
    )
    dynamische_microbiologieregel = create_model(
        "MicrobiologieRegelMetVasteEpisode",
        __base__=MicrobiologieRegel,
        episode=(
            episode_keuze,
            Field(description="Exacte gekoppelde episode uit het instellingenbestand"),
        ),
    )
    dynamische_laboratoriumregel = create_model(
        "LaboratoriumRegelMetVasteEpisode",
        __base__=LaboratoriumRegel,
        episode=(
            episode_keuze,
            Field(description="Exacte gekoppelde episode uit het instellingenbestand"),
        ),
    )

    return create_model(
        "SynthetischDossierMetVasteEpisodes",
        __base__=SynthetischDossier,
        episodelijst=(list[dynamische_episodeoverzichtregel], Field(...)),
        contacten=(list[dynamisch_deelcontact], Field(...)),
        correspondentie=(list[dynamische_correspondentieregel], Field(...)),
        microbiologie=(list[dynamische_microbiologieregel], Field(...)),
        laboratorium=(list[dynamische_laboratoriumregel], Field(...)),
    )


def datum_naar_nederlands(datum: str) -> str:
    """Zet JJJJ-MM-DD om naar DD-MM-JJJJ voor weergave in Excel."""
    if not datum:
        return ""

    try:
        return datetime.strptime(datum, "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        # Laat de oorspronkelijke waarde staan als het model onverwacht een
        # andere notatie heeft gebruikt.
        return datum


def normaliseer_episodenaam(waarde: str) -> str:
    """Normaliseer alleen voor een robuuste vergelijking van episodenamen."""
    return re.sub(r"\s+", " ", waarde.strip()).casefold()


def valideer_gegenereerd_dossier(
    dossier: SynthetischDossier,
    instellingen: PatientInstellingen,
) -> None:
    """Controleer essentiële aantallen, episodenamen en episodechronologie."""
    fouten = []

    if len(dossier.contacten) != instellingen.aantal_deelcontacten:
        fouten.append(
            f"verwacht {instellingen.aantal_deelcontacten} deelcontacten, "
            f"maar ontving {len(dossier.contacten)}"
        )

    verwachte_episodes = [episode.episode for episode in instellingen.episodes]
    ontvangen_episodes = [episode.episode for episode in dossier.episodelijst]
    if len(ontvangen_episodes) != len(verwachte_episodes):
        fouten.append(
            f"verwacht {len(verwachte_episodes)} regels in de episodelijst, "
            f"maar ontving {len(ontvangen_episodes)}"
        )
    else:
        for index, (verwacht, ontvangen) in enumerate(
            zip(verwachte_episodes, ontvangen_episodes),
            start=1,
        ):
            if normaliseer_episodenaam(verwacht) != normaliseer_episodenaam(ontvangen):
                fouten.append(
                    f"episode {index} heet '{ontvangen}' in plaats van '{verwacht}'"
                )

    toegestane_episodes = {
        normaliseer_episodenaam(episode) for episode in verwachte_episodes
    }
    onbekende_contactepisodes = sorted(
        {
            contact.episode
            for contact in dossier.contacten
            if normaliseer_episodenaam(contact.episode) not in toegestane_episodes
        }
    )
    if onbekende_contactepisodes:
        fouten.append(
            "onbekende episodenamen in het journaal: "
            + ", ".join(onbekende_contactepisodes)
        )

    _, dossier_einddatum = bepaal_dossierperiode(instellingen.aantal_jaren)
    uiterste_datum = date.fromisoformat(dossier_einddatum)
    for episode in dossier.episodelijst:
        try:
            start = date.fromisoformat(episode.startdatum)
        except ValueError:
            fouten.append(
                f"ongeldige startdatum bij episode '{episode.episode}': "
                f"{episode.startdatum}"
            )
            continue

        if start > uiterste_datum:
            fouten.append(
                f"startdatum van episode '{episode.episode}' ligt na de dossierperiode"
            )

        if episode.status == "Actief" and episode.einddatum:
            fouten.append(
                f"actieve episode '{episode.episode}' heeft toch een einddatum"
            )
        if episode.status == "Afgesloten" and not episode.einddatum:
            fouten.append(
                f"afgesloten episode '{episode.episode}' mist een einddatum"
            )
        if episode.einddatum:
            try:
                einde = date.fromisoformat(episode.einddatum)
            except ValueError:
                fouten.append(
                    f"ongeldige einddatum bij episode '{episode.episode}': "
                    f"{episode.einddatum}"
                )
                continue
            if einde < start:
                fouten.append(
                    f"einddatum van episode '{episode.episode}' ligt vóór de startdatum"
                )
            if einde > uiterste_datum:
                fouten.append(
                    f"einddatum van episode '{episode.episode}' ligt na de dossierperiode"
                )

    if fouten:
        raise RuntimeError(
            "Het model leverde een dossier op dat niet aan de basiscontroles voldoet:\n- "
            + "\n- ".join(fouten)
        )


def dossier_naar_excel(
    dossier: SynthetischDossier,
    instellingen: PatientInstellingen,
    uitvoerpad: Path,
) -> None:
    """Exporteer patiëntgegevens, episodelijst en dossieronderdelen naar Excel."""
    startdatum, einddatum = bepaal_dossierperiode(instellingen.aantal_jaren)
    geboortedatum = bepaal_fictieve_geboortedatum(
        instellingen.patient_id,
        instellingen.leeftijd,
        einddatum,
    )

    patient_dataframe = pd.DataFrame(
        [
            {"Onderdeel": "Patiënt-ID", "Waarde": instellingen.patient_id},
            {
                "Onderdeel": "Geboortedatum",
                "Waarde": datum_naar_nederlands(geboortedatum),
            },
            {
                "Onderdeel": "Leeftijd einde dossierperiode",
                "Waarde": f"{instellingen.leeftijd} jaar",
            },
            {"Onderdeel": "Formeel geslacht", "Waarde": instellingen.geslacht},
            {
                "Onderdeel": "Dossierperiode",
                "Waarde": (
                    f"{datum_naar_nederlands(startdatum)} t/m "
                    f"{datum_naar_nederlands(einddatum)}"
                ),
            },
            {
                "Onderdeel": "Rookstatus",
                "Waarde": dossier.patient_achtergrond.rookstatus,
            },
            {
                "Onderdeel": "Alcoholgebruik",
                "Waarde": dossier.patient_achtergrond.alcoholgebruik,
            },
            {
                "Onderdeel": "Familieanamnese",
                "Waarde": dossier.patient_achtergrond.familieanamnese,
            },
            {
                "Onderdeel": "Woonsituatie",
                "Waarde": dossier.patient_achtergrond.woonsituatie,
            },
            {
                "Onderdeel": "Beroep/dagbesteding",
                "Waarde": dossier.patient_achtergrond.beroep_dagbesteding,
            },
            {
                "Onderdeel": "Functionele context",
                "Waarde": dossier.patient_achtergrond.functionele_context,
            },
        ],
        columns=["Onderdeel", "Waarde"],
    )

    episode_rijen = []
    for episode in sorted(dossier.episodelijst, key=lambda regel: regel.volgorde):
        episode_rijen.append(
            {
                "Volgorde": episode.volgorde,
                "Episode": episode.episode,
                "ICPC-code": episode.icpc_code,
                "Startdatum": datum_naar_nederlands(episode.startdatum),
                "Einddatum": datum_naar_nederlands(episode.einddatum),
                "Status": episode.status,
                "Attentiewaarde": episode.attentiewaarde,
                "Samenvatting beloop": episode.samenvatting_beloop,
                "Beleid": episode.beleid,
            }
        )

    episodelijst_dataframe = pd.DataFrame(
        episode_rijen,
        columns=[
            "Volgorde",
            "Episode",
            "ICPC-code",
            "Startdatum",
            "Einddatum",
            "Status",
            "Attentiewaarde",
            "Samenvatting beloop",
            "Beleid",
        ],
    )

    rijen = []

    for contact in dossier.contacten:
        metadata = (
            f"[{contact.contact_id} | {datum_naar_nederlands(contact.datum)} | "
            f"{contact.zorgverlener} | "
            f"{contact.contactvorm} | Episode: {contact.episode}]"
        )
        for soep_code, tekst in (
            ("S", contact.s),
            ("O", contact.o),
            ("E", contact.e),
            ("P", contact.p),
        ):
            rijen.append({"SOEP": soep_code, "Inhoud": f"{metadata} {tekst}"})

    dataframe = pd.DataFrame(rijen, columns=["SOEP", "Inhoud"])

    medicatie_rijen = []
    for medicijn in dossier.medicatie:
        medicatie_rijen.append(
            {
                "Geneesmiddel": medicijn.geneesmiddel,
                "Sterkte": medicijn.sterkte,
                "Dosering": medicijn.dosering,
                "Indicatie": medicijn.indicatie,
                "Startdatum": datum_naar_nederlands(medicijn.startdatum),
                "Einddatum": datum_naar_nederlands(medicijn.einddatum),
                "Status": medicijn.status,
            }
        )

    medicatie_dataframe = pd.DataFrame(
        medicatie_rijen,
        columns=[
            "Geneesmiddel",
            "Sterkte",
            "Dosering",
            "Indicatie",
            "Startdatum",
            "Einddatum",
            "Status",
        ],
    )

    correspondentie_rijen = []
    # Sorteer eerst op de interne JJJJ-MM-DD-notatie. Daarna zetten we de datum
    # pas om naar DD-MM-JJJJ voor de zichtbare Excel-uitvoer.
    gesorteerde_correspondentie = sorted(
        dossier.correspondentie,
        key=lambda brief: (brief.datum, brief.traject_id),
    )
    for brief in gesorteerde_correspondentie:
        correspondentie_rijen.append(
            {
                "Traject-ID": brief.traject_id,
                "Datum": datum_naar_nederlands(brief.datum),
                "Type": brief.type_brief,
                "Specialisme": brief.specialisme,
                "Episode": brief.episode,
                "Van": brief.van,
                "Aan": brief.aan,
                "Onderwerp": brief.onderwerp,
                "Brieftekst": brief.inhoud,
            }
        )

    correspondentie_dataframe = pd.DataFrame(
        correspondentie_rijen,
        columns=[
            "Traject-ID",
            "Datum",
            "Type",
            "Specialisme",
            "Episode",
            "Van",
            "Aan",
            "Onderwerp",
            "Brieftekst",
        ],
    )

    allergie_rijen = []
    for allergie in dossier.allergieen:
        allergie_rijen.append(
            {
                "Allergeen": allergie.allergeen,
                "Type": allergie.type_allergie,
                "Reactie": allergie.reactie,
                "Ernst": allergie.ernst,
                "Registratiedatum": datum_naar_nederlands(allergie.registratiedatum),
                "Status": allergie.status,
            }
        )

    allergie_dataframe = pd.DataFrame(
        allergie_rijen,
        columns=[
            "Allergeen",
            "Type",
            "Reactie",
            "Ernst",
            "Registratiedatum",
            "Status",
        ],
    )

    microbiologie_rijen = []
    for uitslag in sorted(dossier.microbiologie, key=lambda regel: regel.datum):
        microbiologie_rijen.append(
            {
                "Datum": datum_naar_nederlands(uitslag.datum),
                "Episode": uitslag.episode,
                "Materiaal": uitslag.materiaal,
                "Onderzoek": uitslag.onderzoek,
                "Uitslag": uitslag.uitslag,
                "Verwekker": uitslag.verwekker,
                "Hoeveelheid/groei": uitslag.hoeveelheid,
                "Gevoeligheid/resistentie": uitslag.gevoeligheid_resistentie,
                "Conclusie": uitslag.conclusie,
            }
        )

    microbiologie_dataframe = pd.DataFrame(
        microbiologie_rijen,
        columns=[
            "Datum",
            "Episode",
            "Materiaal",
            "Onderzoek",
            "Uitslag",
            "Verwekker",
            "Hoeveelheid/groei",
            "Gevoeligheid/resistentie",
            "Conclusie",
        ],
    )

    laboratorium_rijen = []
    for uitslag in sorted(
        dossier.laboratorium,
        key=lambda regel: (regel.datum, regel.aanvraag_id, regel.bepaling),
    ):
        laboratorium_rijen.append(
            {
                "Aanvraag-ID": uitslag.aanvraag_id,
                "Datum": datum_naar_nederlands(uitslag.datum),
                "Aanvraagreden": uitslag.aanvraagreden,
                "Episode": uitslag.episode,
                "Bepaling": uitslag.bepaling,
                "Waarde": uitslag.waarde,
                "Eenheid": uitslag.eenheid,
                "Referentiewaarde": uitslag.referentiewaarde,
                "Afwijking": uitslag.afwijking,
                "Conclusie": uitslag.conclusie,
            }
        )

    laboratorium_dataframe = pd.DataFrame(
        laboratorium_rijen,
        columns=[
            "Aanvraag-ID",
            "Datum",
            "Aanvraagreden",
            "Episode",
            "Bepaling",
            "Waarde",
            "Eenheid",
            "Referentiewaarde",
            "Conclusie",
            "Afwijking",
        ],
    )

    # Plaats de correspondentie enkele rijen onder de medicatielijst.
    # Deze variabele is een Excel-rijnummer (Excel telt vanaf 1).
    correspondentie_titelrij = len(medicatie_dataframe) + 5
    correspondentie_eindrij = correspondentie_titelrij + 1 + len(
        correspondentie_dataframe
    )
    allergie_titelrij = correspondentie_eindrij + 3
    allergie_eindrij = allergie_titelrij + 1 + len(allergie_dataframe)
    microbiologie_titelrij = allergie_eindrij + 3
    microbiologie_eindrij = microbiologie_titelrij + 1 + len(
        microbiologie_dataframe
    )
    laboratorium_titelrij = microbiologie_eindrij + 3
    laboratorium_eindrij = laboratorium_titelrij + 1 + len(
        laboratorium_dataframe
    )

    with pd.ExcelWriter(uitvoerpad, engine="openpyxl") as writer:
        patient_dataframe.to_excel(
            writer,
            sheet_name="Patiëntgegevens",
            index=False,
            startrow=2,
        )
        episodelijst_dataframe.to_excel(
            writer,
            sheet_name="Episodelijst",
            index=False,
            startrow=2,
        )
        dataframe.to_excel(writer, sheet_name="Journaal", index=False)
        medicatie_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=1,
            startcol=3,
        )
        correspondentie_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=correspondentie_titelrij,
            startcol=3,
        )
        allergie_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=allergie_titelrij,
            startcol=3,
        )
        microbiologie_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=microbiologie_titelrij,
            startcol=3,
        )
        laboratorium_dataframe.to_excel(
            writer,
            sheet_name="Journaal",
            index=False,
            startrow=laboratorium_titelrij,
            startcol=3,
        )

        patient_werkblad = writer.book["Patiëntgegevens"]
        patient_werkblad.merge_cells("A1:B1")
        patient_werkblad["A1"] = "PATIËNT- EN ACHTERGRONDGEGEVENS"
        patient_werkblad.freeze_panes = "A4"
        patient_werkblad.auto_filter.ref = f"A3:B{len(patient_dataframe) + 3}"
        patient_werkblad.column_dimensions["A"].width = 32
        patient_werkblad.column_dimensions["B"].width = 85

        episode_werkblad = writer.book["Episodelijst"]
        episode_werkblad.merge_cells("A1:I1")
        episode_werkblad["A1"] = "EPISODE-/PROBLEEMLIJST"
        episode_werkblad.freeze_panes = "A4"
        episode_werkblad.auto_filter.ref = (
            f"A3:I{len(episodelijst_dataframe) + 3}"
        )
        episode_breedtes = {
            "A": 10,
            "B": 32,
            "C": 13,
            "D": 14,
            "E": 14,
            "F": 14,
            "G": 16,
            "H": 50,
            "I": 55,
        }
        for kolom, breedte in episode_breedtes.items():
            episode_werkblad.column_dimensions[kolom].width = breedte

        werkblad = writer.book["Journaal"]
        werkblad.freeze_panes = "A2"
        werkblad.auto_filter.ref = f"A1:B{len(dataframe) + 1}"
        werkblad.column_dimensions["A"].width = 10
        werkblad.column_dimensions["B"].width = 115

        werkblad.merge_cells("D1:J1")
        werkblad["D1"] = "MEDICATIELIJST"
        werkblad.column_dimensions["D"].width = 24
        werkblad.column_dimensions["E"].width = 14
        werkblad.column_dimensions["F"].width = 22
        werkblad.column_dimensions["G"].width = 32
        werkblad.column_dimensions["H"].width = 14
        werkblad.column_dimensions["I"].width = 14
        werkblad.column_dimensions["J"].width = 12
        werkblad.column_dimensions["K"].width = 30
        werkblad.column_dimensions["L"].width = 100
        werkblad.column_dimensions["M"].width = 14

        werkblad.merge_cells(
            start_row=correspondentie_titelrij,
            start_column=4,
            end_row=correspondentie_titelrij,
            end_column=12,
        )
        correspondentie_titelcel = werkblad.cell(
            row=correspondentie_titelrij,
            column=4,
        )
        correspondentie_titelcel.value = (
            "CORRESPONDENTIE – VERWIJS- EN SPECIALISTENBRIEVEN"
        )

        werkblad.merge_cells(
            start_row=allergie_titelrij,
            start_column=4,
            end_row=allergie_titelrij,
            end_column=9,
        )
        allergie_titelcel = werkblad.cell(row=allergie_titelrij, column=4)
        allergie_titelcel.value = "ALLERGIEËN"

        werkblad.merge_cells(
            start_row=microbiologie_titelrij,
            start_column=4,
            end_row=microbiologie_titelrij,
            end_column=12,
        )
        microbiologie_titelcel = werkblad.cell(
            row=microbiologie_titelrij,
            column=4,
        )
        microbiologie_titelcel.value = "MICROBIOLOGIE"

        werkblad.merge_cells(
            start_row=laboratorium_titelrij,
            start_column=4,
            end_row=laboratorium_titelrij,
            end_column=13,
        )
        laboratorium_titelcel = werkblad.cell(
            row=laboratorium_titelrij,
            column=4,
        )
        laboratorium_titelcel.value = "LABORATORIUMAANVRAGEN EN -UITSLAGEN"

        from openpyxl.styles import Alignment, Font, PatternFill

        patient_werkblad["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        patient_werkblad["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        patient_werkblad["A1"].alignment = Alignment(horizontal="center")
        for cel in patient_werkblad[3][0:2]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="5B9BD5")
            cel.alignment = Alignment(horizontal="center", vertical="center")
        for rij in patient_werkblad.iter_rows(
            min_row=4,
            max_row=len(patient_dataframe) + 3,
            min_col=1,
            max_col=2,
        ):
            rij[0].font = Font(bold=True, color="1F1F1F")
            rij[0].fill = PatternFill("solid", fgColor="DDEBF7")
            rij[1].fill = PatternFill("solid", fgColor="FFF2CC")
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        episode_werkblad["A1"].font = Font(bold=True, color="FFFFFF", size=14)
        episode_werkblad["A1"].fill = PatternFill("solid", fgColor="1F4E78")
        episode_werkblad["A1"].alignment = Alignment(horizontal="center")
        for cel in episode_werkblad[3][0:9]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="5B9BD5")
            cel.alignment = Alignment(horizontal="center", vertical="center")
        for rij in episode_werkblad.iter_rows(
            min_row=4,
            max_row=len(episodelijst_dataframe) + 3,
            min_col=1,
            max_col=9,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")
            rij[0].alignment = Alignment(horizontal="center", vertical="top")
            rij[2].alignment = Alignment(horizontal="center", vertical="top")
            rij[5].alignment = Alignment(horizontal="center", vertical="top")
            rij[6].alignment = Alignment(horizontal="center", vertical="top")

        for cel in werkblad[1][0:2]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="1F4E78")
            cel.alignment = Alignment(vertical="center")

        werkblad["D1"].font = Font(bold=True, color="FFFFFF")
        werkblad["D1"].fill = PatternFill("solid", fgColor="548235")
        werkblad["D1"].alignment = Alignment(horizontal="center")

        for cel in werkblad[2][3:10]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="70AD47")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        correspondentie_titelcel.font = Font(bold=True, color="FFFFFF")
        correspondentie_titelcel.fill = PatternFill("solid", fgColor="7030A0")
        correspondentie_titelcel.alignment = Alignment(horizontal="center")

        for cel in werkblad[correspondentie_titelrij + 1][3:12]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="9E66B4")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        allergie_titelcel.font = Font(bold=True, color="FFFFFF")
        allergie_titelcel.fill = PatternFill("solid", fgColor="C65911")
        allergie_titelcel.alignment = Alignment(horizontal="center")

        for cel in werkblad[allergie_titelrij + 1][3:9]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="ED7D31")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        microbiologie_titelcel.font = Font(bold=True, color="FFFFFF")
        microbiologie_titelcel.fill = PatternFill("solid", fgColor="008C95")
        microbiologie_titelcel.alignment = Alignment(horizontal="center")

        for cel in werkblad[microbiologie_titelrij + 1][3:12]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="00A6B2")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        laboratorium_titelcel.font = Font(bold=True, color="FFFFFF")
        laboratorium_titelcel.fill = PatternFill("solid", fgColor="2F5597")
        laboratorium_titelcel.alignment = Alignment(horizontal="center")

        for cel in werkblad[laboratorium_titelrij + 1][3:13]:
            cel.font = Font(bold=True, color="FFFFFF")
            cel.fill = PatternFill("solid", fgColor="4472C4")
            cel.alignment = Alignment(horizontal="center", vertical="center")

        for rij in werkblad.iter_rows(min_row=2, max_col=2):
            rij[0].alignment = Alignment(horizontal="center", vertical="top")
            rij[1].alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=3,
            max_row=len(medicatie_dataframe) + 2,
            min_col=4,
            max_col=10,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=correspondentie_titelrij + 2,
            max_row=correspondentie_eindrij,
            min_col=4,
            max_col=12,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=allergie_titelrij + 2,
            max_row=allergie_eindrij,
            min_col=4,
            max_col=9,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=microbiologie_titelrij + 2,
            max_row=microbiologie_eindrij,
            min_col=4,
            max_col=12,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")

        for rij in werkblad.iter_rows(
            min_row=laboratorium_titelrij + 2,
            max_row=laboratorium_eindrij,
            min_col=4,
            max_col=13,
        ):
            for cel in rij:
                cel.alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    instellingenpad = Path(__file__).with_name(SETTINGS_BESTANDSNAAM)
    instellingen = lees_patientinstellingen(instellingenpad)
    patient_prompt = maak_patient_prompt(instellingen)
    dossiermodel = maak_dynamisch_dossiermodel(instellingen)

    api_key = getpass("Plak hier je API-key en druk op Enter: ").strip()
    if not api_key:
        raise ValueError("Er is geen API-key ingevoerd.")

    client = OpenAI(api_key=api_key)

    response = client.responses.parse(
        model="gpt-5.6",
        input=[
            {"role": "system", "content": MAIN_PROMPT},
            {"role": "user", "content": patient_prompt},
        ],
        text_format=dossiermodel,
    )

    dossier = response.output_parsed
    if dossier is None:
        raise RuntimeError("Het model leverde geen bruikbaar gestructureerd dossier op.")

    valideer_gegenereerd_dossier(dossier, instellingen)

    # Elke run krijgt een uniek tijdstip in de bestandsnaam, zodat eerdere
    # gegenereerde dossiers niet worden overschreven.
    tijdstip = datetime.now().strftime("%Y%m%d_%H%M%S")
    patient_id = veilige_bestandsnaam(instellingen.patient_id)
    uitvoerpad = Path(__file__).with_name(
        f"synthetisch_epd_{patient_id}_{tijdstip}.xlsx"
    )
    dossier_naar_excel(dossier, instellingen, uitvoerpad)
    print(f"Klaar: {uitvoerpad.resolve()}")


if __name__ == "__main__":
    main()